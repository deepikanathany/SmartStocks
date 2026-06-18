"""
final_5_app.py
==============
SmartStock — Final 5 Picker  (web UI version)

Self-contained Flask app. Run in the same folder as your SmartStock app.py.

Usage:
    python final_5_app.py
    Open http://localhost:5006

Requires SmartStock running on localhost:5000
pip install flask requests openpyxl yfinance
"""

from flask import Flask, render_template_string, request, jsonify, Response, send_file
import threading, time, json, io
from datetime import datetime
from collections import defaultdict
import requests as _req

app  = Flask(__name__)
PORT = 5006
HOST = "http://localhost:5000"

# ── Shared job state ──────────────────────────────────────────────────────────
_job   = {"running": False, "log": [], "screener_results": {},
          "cross": {}, "top5": [], "done": False, "error": None,
          "config": {}, "start_time": None}
_lock  = threading.Lock()

# ── Universe + sector definitions ─────────────────────────────────────────────
UNIVERSES = [
    {"key": "SP500_TOP50",   "label": "S&P 500 Top 50",    "count": 50,   "time": "~20 min"},
    {"key": "SP500_TOP100",  "label": "S&P 500 Top 100",   "count": 100,  "time": "~40 min"},
    {"key": "SP500_TOP200",  "label": "S&P 500 Top 200",   "count": 200,  "time": "~70 min"},
    {"key": "SP500",         "label": "S&P 500 Full",       "count": 503,  "time": "~2.5 hrs"},
    {"key": "NASDAQ100",     "label": "NASDAQ 100",         "count": 100,  "time": "~40 min"},
    {"key": "DOWJONES",      "label": "Dow Jones 30",       "count": 30,   "time": "~12 min"},
    {"key": "RUSSELL1000",   "label": "Russell 1000",       "count": 1000, "time": "~6 hrs ⚠"},
]

SECTORS = [
    "Energy","Industrials","Financials","Financial Services",
    "Technology","Healthcare","Utilities","Consumer Defensive",
    "Consumer Cyclical","Basic Materials","Real Estate",
    "Communication Services",
]

SCREENER_WEIGHTS = {
    "Fixed Triple":              1.5,
    "Triple Screen":             1.4,
    "SEPA Minervini":            1.3,
    "US Investing Championship": 1.3,
    "Swing Pro":                 1.0,
    "Combined":                  1.0,
}
RANK_SCORES = {1:100, 2:85, 3:70, 4:55, 5:40}
TOP_N = 5

# ═══════════════════════════════════════════════════════════════════════════════
# SCREENER ENGINE (adapted from final_5_screener.py)
# ═══════════════════════════════════════════════════════════════════════════════

def _log(msg, kind="info", screener=None):
    with _lock:
        _job["log"].append({
            "ts":       datetime.now().strftime("%H:%M:%S"),
            "msg":      msg,
            "kind":     kind,
            "screener": screener,
        })

def _start(endpoint, payload):
    try:
        r = _req.post(f"{HOST}{endpoint}", json=payload,
                      headers={"Content-Type":"application/json"}, timeout=30)
        return r.json().get("job_id")
    except Exception as e:
        _log(f"Failed to start {endpoint}: {e}", "error")
        return None

def _poll(job_id, screener_name):
    start   = time.time()
    results = []; seen = set()
    while True:
        elapsed = int(time.time() - start)
        if elapsed > 1800:
            _log(f"{screener_name}: timeout", "error", screener_name)
            return results, []
        try:
            d = _req.get(f"{HOST}/api/progress/{job_id}", timeout=10).json()
        except Exception:
            time.sleep(5); continue

        for entry in (d.get("log") or []):
            if isinstance(entry, dict) and entry.get("kind") == "match_result":
                mr  = entry.get("match_result", {})
                sym = mr.get("ticker") or mr.get("symbol")
                if sym and sym not in seen:
                    seen.add(sym); results.append(mr)

        for mr in (d.get("results") or []):
            sym = mr.get("ticker") or mr.get("symbol")
            if sym and sym not in seen:
                seen.add(sym); results.append(mr)

        pct = min(99, int(elapsed / 1200 * 100))
        _log(f"{'▰' * (pct//10)}{'▱' * (10-pct//10)} {elapsed}s · {len(results)} found",
             "progress", screener_name)

        if d.get("done"):
            rej = []
            try:
                rej = _req.get(f"{HOST}/api/get_rejections/{job_id}",
                                timeout=15).json().get("rejections", [])
            except Exception:
                pass
            _log(f"Complete — {len(results)} results, {len(rej)} rejections",
                 "done", screener_name)
            return results, rej

        time.sleep(5)

def _run_screener(name, endpoint, payload):
    cfg = _job["config"]
    payload["exchanges"]    = [cfg["universe"]]
    payload["customTickers"] = []
    if cfg.get("sectors"):
        payload["sectors"] = cfg["sectors"]
    _log(f"Starting {name}...", "phase", name)
    with _lock:
        _job["screener_results"][name] = {
            "status": "running", "results": [], "rejections": [], "ranked": []
        }
    job_id = _start(endpoint, payload)
    if not job_id:
        with _lock:
            _job["screener_results"][name]["status"] = "error"
        return [], []
    results, rej = _poll(job_id, name)
    return results, rej


def run_all_screeners(cfg):
    """Main job — runs in background thread."""
    with _lock:
        _job.update({"running": True, "log": [], "screener_results": {},
                     "cross": {}, "top5": [], "done": False, "error": None,
                     "config": cfg, "start_time": datetime.now().isoformat()})

    _log(f"Starting Final 5 Screener", "title")
    _log(f"Universe: {cfg['universe']}  |  Sectors: {cfg.get('sectors') or 'All'}", "info")
    _log(f"Running 6 screeners sequentially...", "info")

    try:
        screener_fns = [
            ("Combined", "/api/run", {
                "mode": "combined", "conditions": {}, "minSwingScore": 4, "minValueScore": 3,
                "thresholds": {"min_market_cap_b":1,"min_roe_pct":8,"max_pe":50},
            }),
            ("SEPA Minervini", "/api/run_sepa", {
                "thresholds": {"sma200_min_days":20,"min_pct_above_52w_low":25,
                               "max_pct_below_52w_high":30,"min_rs_rating":65,
                               "min_eps_growth_pct":15,"min_rev_growth_pct":10,
                               "vcp_contractions":2},
            }),
            ("Triple Screen", "/api/run_triple", {
                "sepaThresholds": {"sma200_min_days":20,"min_pct_above_52w_low":25,
                                   "max_pct_below_52w_high":30,"min_rs_rating":65,
                                   "min_eps_growth_pct":15,"min_rev_growth_pct":10,
                                   "vcp_contractions":2},
                "swingConditions": {"market_cap":True,"div_yield":False,"pe_ratio":False,
                                    "debt_equity":False,"monthly_trend":True,"weekly_trend":False,
                                    "swing_rsi":True,"swing_ema":True,"swing_volume":False,
                                    "swing_macd":False,"swing_atr":False},
                "swingThresholds": {"min_market_cap_b":1,"swing_rsi_min":25,"swing_rsi_max":75},
                "valueConditions": {"pe_ratio":True,"debt_equity":True,"roe":True,"analyst_rating":True},
                "valueThresholds": {"max_pe":60,"max_debt_equity_pct":150,
                                    "min_roe_pct":5,"max_analyst_mean":3.5},
            }),
            ("Swing Pro", "/api/run_swing_pro", {
                "strategies": {"ema_reversion":True,"cons_breakout":True,"xlp_reversion":True,
                                "seasonal_nasdaq":True,"seasonal_tom":True},
                "thresholds": {"min_score":30,"ema20_proximity_pct":3.0,"ema_rsi_low":40,
                                "ema_rsi_high":55,"cons_range_pct":10.0},
                "run_backtest":True,"run_sentiment":True,"run_verdict":True,
            }),
            ("US Investing Championship", "/api/run_usic", {
                "thresholds": {"sma200_min_days":20,"min_pct_above_52w_low":25,
                               "max_pct_below_52w_high":30,"min_eps_growth_pct":20,
                               "min_rev_growth_pct":15,"vcp_contractions":2,
                               "eps_accel_required":True,"surprise_required":False,
                               "sector_rs_required":True,"min_rs_rating":70,
                               "min_champion_score":5.0},
            }),
            ("Fixed Triple", "/api/run_fixed_triple", {
                "market":"US","useRegimeGate":True,"sepaThresholds":{},
                "swingThresholds":{"min_monthly_chg_pct":-5.0},
            }),
        ]

        all_ranked = {}

        for screener_name, endpoint, base_payload in screener_fns:
            results, rejections = _run_screener(screener_name, endpoint, base_payload)
            ranked = _rank(results, screener_name)

            # Pad with near-miss rejects if < TOP_N
            if len(ranked) < TOP_N and rejections:
                rej_sorted = sorted(
                    [r for r in rejections if r.get("rs_rating")],
                    key=lambda x: x.get("rs_rating",0), reverse=True
                )
                for rj in rej_sorted[:TOP_N - len(ranked)]:
                    rj.update({"_rank": len(ranked)+1, "_screener": screener_name,
                                "_rank_pts": 15 * SCREENER_WEIGHTS[screener_name],
                                "_is_reject": True, "_composite": 0})
                    ranked.append(rj)

            all_ranked[screener_name] = ranked

            with _lock:
                _job["screener_results"][screener_name] = {
                    "status":     "done",
                    "results":    results,
                    "rejections": rejections,
                    "ranked":     [_serialise(r) for r in ranked[:TOP_N]],
                }

        # Cross-score
        _log("Computing cross-screener scores...", "phase")
        top_n_only = {k: v[:TOP_N] for k,v in all_ranked.items()}
        cross = _cross_score(top_n_only)

        # Live entry trigger check
        _log("Checking live entry triggers (yfinance)...", "phase")
        try:
            import yfinance as yf
            for sym, c in list(cross.items())[:20]:
                trig = _entry_trigger(sym, yf)
                c["entry_triggered"] = c["entry_triggered"] or trig["triggered"]
                c["trigger_note"]    = trig["note"]
                c["price"]           = trig.get("price") or c.get("price")
        except ImportError:
            _log("yfinance not installed — skipping live trigger check", "warn")

        sorted_cross = sorted(cross.items(), key=lambda x: -x[1]["score"])
        top5         = sorted_cross[:TOP_N]

        with _lock:
            _job["cross"] = {s: _c_serial(c) for s,c in sorted_cross[:30]}
            _job["top5"]  = [{"symbol":s, **_c_serial(c)} for s,c in top5]
            _job["done"]  = True

        _log(f"Done! Top 5: {' · '.join(s for s,_ in top5)}", "done")

    except Exception as e:
        import traceback
        _log(f"Fatal error: {e}", "error")
        _log(traceback.format_exc(), "error")
        with _lock:
            _job["done"]  = True
            _job["error"] = str(e)


def _get_ticker(r):
    return (r.get("ticker") or r.get("symbol") or "").upper().strip()

def _rank(results, name):
    def key(r):
        if name == "Combined":
            return (r.get("combined_score") or 0,
                    r.get("swing_score") or 0,
                    r.get("value_score") or 0)
        if name in ("SEPA Minervini","Triple Screen"):
            rs  = r.get("rs_rating") or 0
            eps = min(r.get("eps_growth_pct") or 0, 200)
            rev = min(r.get("rev_growth_pct") or 0, 200)
            ai  = r.get("ai_score") or 0
            vcp = 10 if r.get("near_pivot") else (5 if r.get("vcp") else 0)
            return (rs*0.4 + eps*0.2 + rev*0.2 + ai*2 + vcp,)
        if name == "Swing Pro":
            return (r.get("rank_score") or 0,)
        if name == "US Investing Championship":
            return (r.get("champion_score") or 0,)
        if name == "Fixed Triple":
            return ((1 if r.get("entry_triggered") else 0)*50
                    + (r.get("rs_rating") or 0)*0.4
                    + (100-(r.get("rsi2") or 50))*0.1,)
        return (0,)

    ranked = sorted(results, key=key, reverse=True)
    w = SCREENER_WEIGHTS[name]
    for i,r in enumerate(ranked):
        r["_rank"]      = i+1
        r["_screener"]  = name
        r["_is_reject"] = False
        r["_rank_pts"]  = RANK_SCORES.get(i+1, max(5,40-(i-4)*5)) * w
        rs  = r.get("rs_rating") or 0
        eps = min(r.get("eps_growth_pct") or 0, 200)
        rev = min(r.get("rev_growth_pct") or 0, 200)
        ai  = r.get("ai_score") or 0
        r["_composite"] = round(rs*0.4+eps*0.2+rev*0.2+ai*2, 1)
    return ranked

def _cross_score(all_ranked):
    cross = defaultdict(lambda: {
        "score":0.0,"frequency":0,"screeners":[],"ranks":{},
        "entry_triggered":False,"rs_rating":0,"name":"","sector":"",
        "price":None,"monthly_chg":None,"eps_growth":None,"rev_growth":None,
        "champion_score":None,"combined_score":None,"ai_score":None,
        "vcp":False,"near_pivot":False,"rsi2":None,"trigger_note":"",
        "is_reject":False,
    })
    for sname, ranked in all_ranked.items():
        for r in ranked:
            sym = _get_ticker(r)
            if not sym: continue
            c = cross[sym]
            c["score"]     += r.get("_rank_pts", 0)
            c["frequency"] += 1
            c["screeners"].append(sname)
            c["ranks"][sname] = r.get("_rank", 99)
            c["is_reject"]    = c["is_reject"] or r.get("_is_reject", False)
            if r.get("entry_triggered"): c["entry_triggered"] = True
            if r.get("rs_rating"):       c["rs_rating"]       = max(c["rs_rating"], r["rs_rating"])
            for k in ["name","sector","price_now","monthly_chg_pct","eps_growth_pct",
                      "rev_growth_pct","champion_score","combined_score","ai_score","rsi2"]:
                if r.get(k) and not c[k.replace("_pct","").replace("price_now","price")
                                        .replace("monthly_chg_pct","monthly_chg")
                                        .replace("eps_growth_pct","eps_growth")
                                        .replace("rev_growth_pct","rev_growth")]:
                    fk = (k.replace("price_now","price")
                           .replace("monthly_chg_pct","monthly_chg")
                           .replace("eps_growth_pct","eps_growth")
                           .replace("rev_growth_pct","rev_growth"))
                    c[fk] = r[k]
            if r.get("vcp"):       c["vcp"]       = True
            if r.get("near_pivot"):c["near_pivot"] = True

    for sym,c in cross.items():
        f = c["frequency"]
        c["score"] += (25 if f>=4 else 15 if f==3 else 5 if f==2 else 0)
        c["score"]  = round(c["score"], 1)
    return dict(cross)

def _entry_trigger(sym, yf):
    try:
        h = yf.Ticker(sym).history(period="10d", interval="1d", auto_adjust=True)
        if h is None or len(h) < 3:
            return {"triggered":False,"note":"No data"}
        ph   = float(h["High"].iloc[-2])
        cp   = float(h["Close"].iloc[-1])
        va   = float(h["Volume"].iloc[:-1].mean())
        vt   = float(h["Volume"].iloc[-1])
        vr   = round(vt/va,2) if va>0 else 0
        trig = cp>ph and vr>=1.2
        return {"triggered":trig,"price":round(cp,2),"prev_high":round(ph,2),
                "vol_ratio":vr,
                "note":(f"${cp:.2f} > ${ph:.2f} + vol {vr}×" if trig
                        else f"${cp:.2f} vs ${ph:.2f} · vol {vr}×")}
    except Exception as e:
        return {"triggered":False,"note":str(e)[:50]}

def _serialise(r):
    """Make a result dict JSON-safe."""
    out = {}
    for k,v in r.items():
        if k.startswith("_") and k not in ("_rank","_screener","_is_reject",
                                             "_rank_pts","_composite"):
            continue
        try:
            json.dumps(v)
            out[k] = v
        except TypeError:
            out[k] = str(v)
    return out

def _c_serial(c):
    d = dict(c)
    d["screeners"] = list(d["screeners"])
    d["ranks"]     = dict(d["ranks"])
    return {k:(list(v) if isinstance(v,set) else v) for k,v in d.items()}


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def build_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb  = openpyxl.Workbook()
    hf  = PatternFill("solid", fgColor="0D1520")
    hft = Font(bold=True, color="00D68F", size=10)
    wf  = PatternFill("solid", fgColor="0A2018")
    lf  = PatternFill("solid", fgColor="200808")
    nf  = PatternFill("solid", fgColor="1A1200")
    ctr = Alignment(horizontal="center", vertical="center")

    def hrow(ws, cols, row=1):
        for c,h in enumerate(cols,1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill=hf; cell.font=hft; cell.alignment=ctr
        ws.freeze_panes = f"A{row+1}"
        ws.sheet_view.showGridLines = False

    def afit(ws):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(w+2,8),35)

    # Top 5
    ws1 = wb.active; ws1.title = "Final Top 5"
    hrow(ws1, ["#","Ticker","Company","Sector","Quality Score","Screener Count",
               "RS Rating","Entry NOW?","Trigger Note","Mo%","EPS%","Rev%",
               "VCP","Near Pivot","Champion Score","Appears In"])
    top5 = _job.get("top5",[])
    for i,r in enumerate(top5,2):
        fill = nf if i==2 else wf
        vals=[i-1,r.get("symbol"),r.get("name",""),r.get("sector",""),
              r.get("score"),r.get("frequency"),r.get("rs_rating"),
              "YES ✓" if r.get("entry_triggered") else "NO",
              r.get("trigger_note",""),r.get("monthly_chg"),
              r.get("eps_growth"),r.get("rev_growth"),
              "Yes" if r.get("vcp") else "No",
              "Yes ★" if r.get("near_pivot") else "No",
              r.get("champion_score"),
              ", ".join(r.get("screeners",[]))]
        for ci,v in enumerate(vals,1):
            cell=ws1.cell(row=i,column=ci,value=v)
            cell.fill=fill; cell.alignment=ctr
    afit(ws1)

    # All Cross Scores
    ws2 = wb.create_sheet("All Cross Scores")
    cross = _job.get("cross",{})
    hrow(ws2, ["Ticker","Company","Score","Freq","RS","Entry",
               "Combined","SEPA","Triple","SwingPro","USIC","Fixed Triple"])
    for i,(sym,c) in enumerate(sorted(cross.items(),key=lambda x:-x[1]["score"]),2):
        ranks=c.get("ranks",{})
        fill=wf if i<=6 else PatternFill("solid",fgColor="0D1520")
        vals=[sym,c.get("name",""),c.get("score"),c.get("frequency"),
              c.get("rs_rating"),"YES" if c.get("entry_triggered") else "—",
              ranks.get("Combined","—"),ranks.get("SEPA Minervini","—"),
              ranks.get("Triple Screen","—"),ranks.get("Swing Pro","—"),
              ranks.get("US Investing Championship","—"),ranks.get("Fixed Triple","—")]
        for ci,v in enumerate(vals,1):
            cell=ws2.cell(row=i,column=ci,value=v)
            cell.fill=fill; cell.alignment=ctr
    afit(ws2)

    # Per-screener sheets
    for sname,sdata in _job.get("screener_results",{}).items():
        ws = wb.create_sheet(sname[:25])
        hrow(ws,["Rank","Ticker","Company","Sector","RS","EPS%","Rev%",
                 "Price","Mo%","Metric","Entry","Near Reject"])
        for i,r in enumerate(sdata.get("ranked",[]),2):
            fill=lf if r.get("_is_reject") else wf
            metric=(r.get("champion_score") or r.get("rank_score") or
                    r.get("combined_score") or r.get("_composite") or "—")
            vals=[r.get("_rank","—"),
                  r.get("ticker") or r.get("symbol",""),
                  (r.get("name") or "")[:30],
                  r.get("sector",""),r.get("rs_rating","—"),
                  r.get("eps_growth_pct","—"),r.get("rev_growth_pct","—"),
                  r.get("price_now","—"),r.get("monthly_chg_pct","—"),
                  metric,"YES" if r.get("entry_triggered") else "—",
                  "YES" if r.get("_is_reject") else "—"]
            for ci,v in enumerate(vals,1):
                cell=ws.cell(row=i,column=ci,value=v)
                cell.fill=fill; cell.alignment=ctr
        afit(ws)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
# ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template_string(HTML,
        universes=UNIVERSES, sectors=SECTORS)

@app.route("/api/run", methods=["POST"])
def api_run():
    if _job["running"] and not _job["done"]:
        return jsonify({"error":"Already running"}), 400
    cfg = request.get_json() or {}
    t   = threading.Thread(target=run_all_screeners, args=(cfg,), daemon=True)
    t.start()
    return jsonify({"status":"started"})

@app.route("/api/state")
def api_state():
    with _lock:
        return jsonify({
            "running":          _job["running"],
            "done":             _job["done"],
            "error":            _job["error"],
            "log":              _job["log"][-80:],
            "screener_results": {k: {"status":v["status"],
                                     "count":  len(v.get("results",[])),
                                     "ranked": v.get("ranked",[])}
                                 for k,v in _job["screener_results"].items()},
            "cross":            dict(list(_job["cross"].items())[:30]),
            "top5":             _job["top5"],
            "config":           _job["config"],
            "start_time":       _job["start_time"],
        })

@app.route("/api/stream")
def api_stream():
    """SSE — pushes state updates every 3 seconds."""
    def generate():
        last_log = 0
        while True:
            with _lock:
                new_logs = _job["log"][last_log:]
                last_log = len(_job["log"])
                done     = _job["done"]
                top5     = _job["top5"]
                s_res    = {k:{"status":v["status"],"count":len(v.get("results",[]))}
                            for k,v in _job["screener_results"].items()}
            payload = json.dumps({
                "logs":    new_logs,
                "done":    done,
                "top5":    top5,
                "screens": s_res,
            })
            yield f"data: {payload}\n\n"
            if done:
                break
            time.sleep(3)
    return Response(generate(), mimetype="text/event-stream")

@app.route("/api/export")
def api_export():
    if not _job["done"]:
        return jsonify({"error":"Run not complete"}), 400
    try:
        buf   = build_excel()
        fname = f"final5_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=fname)
    except ImportError:
        return jsonify({"error":"pip install openpyxl"}), 500

@app.route("/api/check_smartstock")
def api_check():
    try:
        r = _req.get(f"{HOST}/", timeout=5)
        return jsonify({"ok": r.status_code == 200})
    except Exception:
        return jsonify({"ok": False})


# ═══════════════════════════════════════════════════════════════════════════════
# HTML TEMPLATE
# ═══════════════════════════════════════════════════════════════════════════════

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>SmartStock — Final 5 Picker</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{--bg:#08090a;--s1:#0f1215;--s2:#141a1e;--bd:#1d2730;--bd2:#243040;
      --acc:#00d68f;--acc2:#0095ff;--red:#ff4444;--gold:#ffb700;
      --txt:#dde4ec;--mut:#4a6070;--mono:'JetBrains Mono',monospace;
      --sans:'Syne','Inter',sans-serif}
html{scroll-behavior:smooth}
body{background:var(--bg);color:var(--txt);font-family:var(--sans);min-height:100vh}
a{color:var(--acc);text-decoration:none}

/* Layout */
.wrap{max-width:1400px;margin:0 auto;padding:28px 24px}
.page-title{font-size:1.8rem;font-weight:800;letter-spacing:-0.5px;margin-bottom:4px}
.page-title span{color:var(--acc)}
.page-sub{font-family:var(--mono);font-size:0.6rem;color:var(--mut);
          letter-spacing:3px;text-transform:uppercase;margin-bottom:28px}
.grid2{display:grid;grid-template-columns:360px 1fr;gap:20px;align-items:start}
.grid3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px}

/* Cards */
.card{background:var(--s1);border:1px solid var(--bd);border-radius:6px;overflow:hidden;margin-bottom:16px}
.card-head{padding:12px 18px;background:var(--s2);border-bottom:1px solid var(--bd);
           font-family:var(--mono);font-size:0.6rem;letter-spacing:2px;
           color:var(--mut);display:flex;justify-content:space-between;align-items:center}
.card-body{padding:18px}

/* Universe pills */
.univ-grid{display:flex;flex-direction:column;gap:6px}
.univ-pill{display:flex;align-items:center;justify-content:space-between;
           padding:10px 14px;border:1px solid var(--bd2);border-radius:4px;
           cursor:pointer;transition:all 0.15s;user-select:none}
.univ-pill:hover{border-color:var(--acc2);color:var(--acc2)}
.univ-pill.sel{background:rgba(0,149,255,0.1);border-color:var(--acc2)}
.univ-pill .name{font-weight:600;font-size:0.82rem}
.univ-pill .meta{font-family:var(--mono);font-size:0.62rem;color:var(--mut);text-align:right}
.univ-pill.sel .meta{color:var(--acc2)}

/* Sector checkboxes */
.sect-grid{display:grid;grid-template-columns:1fr 1fr;gap:5px}
.sect-item{display:flex;align-items:center;gap:7px;padding:6px 8px;
           border:1px solid var(--bd);border-radius:3px;cursor:pointer;
           font-size:0.72rem;transition:all 0.12s;user-select:none}
.sect-item:hover{border-color:var(--acc);color:var(--acc)}
.sect-item.on{background:rgba(0,214,143,0.08);border-color:var(--acc);color:var(--acc)}
.sect-item input{accent-color:var(--acc);cursor:pointer}
.sect-actions{display:flex;gap:6px;margin-bottom:8px}
.sect-btn{font-family:var(--mono);font-size:0.6rem;padding:4px 10px;
          background:transparent;border:1px solid var(--bd2);
          color:var(--mut);cursor:pointer;border-radius:2px;transition:all 0.12s}
.sect-btn:hover{color:var(--txt);border-color:var(--txt)}

/* Run button */
.run-wrap{padding:0;margin-top:4px}
.run-btn{width:100%;padding:14px;background:var(--acc);color:#0a0a0a;
         border:none;font-family:var(--sans);font-size:1rem;font-weight:800;
         letter-spacing:2px;cursor:pointer;border-radius:4px;transition:all 0.2s;
         text-transform:uppercase}
.run-btn:hover:not(:disabled){background:#00ffaa;transform:translateY(-1px);
         box-shadow:0 6px 24px rgba(0,214,143,0.35)}
.run-btn:disabled{opacity:0.4;cursor:not-allowed}
.run-btn.running{background:var(--gold);color:#0a0a0a;animation:pulse 1.4s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:0.7}}

/* Status + server check */
.status-row{display:flex;align-items:center;gap:10px;margin-bottom:16px;flex-wrap:wrap}
.badge{font-family:var(--mono);font-size:0.6rem;padding:3px 10px;
       border-radius:2px;letter-spacing:2px;text-transform:uppercase}
.badge-idle   {background:rgba(74,96,112,0.2);color:var(--mut);border:1px solid var(--bd2)}
.badge-run    {background:rgba(255,183,0,0.12);color:var(--gold);border:1px solid rgba(255,183,0,0.3);animation:pulse 1.2s infinite}
.badge-done   {background:rgba(0,214,143,0.1);color:var(--acc);border:1px solid rgba(0,214,143,0.3)}
.badge-err    {background:rgba(255,68,68,0.1);color:var(--red);border:1px solid rgba(255,68,68,0.3)}
.server-dot   {width:8px;height:8px;border-radius:50%;display:inline-block}
.dot-ok  {background:var(--acc)}
.dot-err {background:var(--red)}
.dot-chk {background:var(--mut);animation:pulse 1s infinite}
.server-lbl{font-family:var(--mono);font-size:0.62rem;color:var(--mut)}

/* Screener status row */
.screen-pills{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:16px}
.screen-pill{font-family:var(--mono);font-size:0.6rem;padding:4px 10px;
             border-radius:20px;border:1px solid var(--bd2);color:var(--mut);
             transition:all 0.3s;white-space:nowrap}
.screen-pill.running{border-color:var(--gold);color:var(--gold);animation:pulse 1.2s infinite}
.screen-pill.done   {border-color:var(--acc);color:var(--acc)}
.screen-pill.error  {border-color:var(--red);color:var(--red)}

/* Log */
.log-body{padding:12px 16px;max-height:280px;overflow-y:auto;
          font-family:var(--mono);font-size:0.65rem;line-height:1.9}
.log-body::-webkit-scrollbar{width:3px}
.log-body::-webkit-scrollbar-thumb{background:var(--bd2)}
.ll{display:block}
.ll.info{color:var(--mut)}.ll.phase{color:var(--acc2);font-weight:600}
.ll.done{color:var(--acc);font-weight:700}.ll.error{color:var(--red)}
.ll.warn{color:var(--gold)}.ll.title{color:var(--txt);font-weight:800}
.ll.progress{color:#2a3a48}

/* Final 5 table */
.f5-table{width:100%;border-collapse:collapse;font-size:0.8rem}
.f5-table thead th{background:var(--s2);padding:10px 12px;
                   font-family:var(--mono);font-size:0.58rem;letter-spacing:2px;
                   text-transform:uppercase;color:var(--mut);border-bottom:1px solid var(--bd);
                   text-align:left;white-space:nowrap}
.f5-table thead th.right{text-align:right}
.f5-table tbody tr{border-bottom:1px solid var(--bd);transition:background 0.1s}
.f5-table tbody tr:hover{background:rgba(255,255,255,0.02)}
.f5-table td{padding:11px 12px;vertical-align:middle}
.rank-cell{font-family:var(--mono);font-weight:800;font-size:1.1rem;color:var(--acc)}
.ticker-cell{font-family:var(--mono);font-weight:700;font-size:0.95rem;color:var(--acc)}
.name-cell{color:var(--mut);font-size:0.76rem;max-width:180px;
           overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.score-cell{font-family:var(--mono);font-weight:700;font-size:0.88rem;color:var(--gold);text-align:right}
.rs-pill{display:inline-block;font-family:var(--mono);font-size:0.6rem;
         padding:2px 8px;border-radius:2px;
         background:rgba(255,183,0,0.1);color:var(--gold);border:1px solid rgba(255,183,0,0.25)}
.rs-high{background:rgba(0,214,143,0.1);color:var(--acc);border-color:rgba(0,214,143,0.25)}
.trig-yes{font-family:var(--mono);font-size:0.7rem;font-weight:700;color:var(--acc)}
.trig-no {font-family:var(--mono);font-size:0.7rem;color:var(--mut)}
.freq-pill{font-family:var(--mono);font-size:0.6rem;padding:2px 8px;border-radius:20px;
           background:rgba(0,149,255,0.1);color:var(--acc2);border:1px solid rgba(0,149,255,0.2)}
.freq-high{background:rgba(0,214,143,0.12);color:var(--acc);border-color:rgba(0,214,143,0.3)}
.screens-cell{font-family:var(--mono);font-size:0.6rem;color:var(--mut);line-height:1.6}
.action-buy{color:var(--acc);font-weight:600;font-size:0.75rem}
.action-watch{color:var(--gold);font-size:0.75rem}

/* Cross table */
.cross-table{width:100%;border-collapse:collapse;font-size:0.72rem}
.cross-table thead th{background:var(--s2);padding:8px 10px;font-family:var(--mono);
                      font-size:0.56rem;letter-spacing:2px;text-transform:uppercase;
                      color:var(--mut);border-bottom:1px solid var(--bd);white-space:nowrap}
.cross-table tbody tr{border-bottom:1px solid rgba(29,39,48,0.5);transition:background 0.1s}
.cross-table tbody tr:hover{background:rgba(255,255,255,0.02)}
.cross-table td{padding:8px 10px;font-family:var(--mono);font-size:0.68rem;vertical-align:middle}
.rank-badge{display:inline-block;padding:1px 7px;border-radius:2px;
            font-size:0.6rem;font-weight:700;background:rgba(0,214,143,0.08);
            border:1px solid rgba(0,214,143,0.2);color:var(--acc)}
.rank-rej  {background:rgba(255,68,68,0.07);border-color:rgba(255,68,68,0.2);color:var(--red)}
.rank-none {color:var(--mut)}

/* Export btn */
.export-btn{padding:8px 20px;background:rgba(0,149,255,0.1);border:1px solid var(--acc2);
            color:var(--acc2);font-family:var(--mono);font-size:0.68rem;font-weight:700;
            border-radius:3px;cursor:pointer;transition:all 0.15s;letter-spacing:1px}
.export-btn:hover{background:rgba(0,149,255,0.2)}
.export-btn:disabled{opacity:0.3;cursor:not-allowed}

/* Empty */
.empty{text-align:center;padding:60px;color:var(--mut);
       font-family:var(--mono);font-size:0.72rem;line-height:2}
.empty-icon{font-size:2rem;margin-bottom:12px;opacity:0.3}

/* Timer */
.timer{font-family:var(--mono);font-size:0.65rem;color:var(--mut)}
</style>
</head>
<body>
<div class="wrap">

  <div class="page-title">SMART<span>STOCK</span> — Final 5 Picker</div>
  <div class="page-sub">// Cross-screener intelligence · 6 screeners · 2 dimensions</div>

  <div class="grid2">

    <!-- ── LEFT: Config ── -->
    <div>

      <div class="card">
        <div class="card-head">
          <span>UNIVERSE</span>
          <span id="univLabel" style="color:var(--acc2);font-size:0.65rem">S&P 500 Top 100</span>
        </div>
        <div class="card-body">
          <div class="univ-grid" id="univGrid">
            {% for u in universes %}
            <div class="univ-pill {% if u.key == 'SP500_TOP100' %}sel{% endif %}"
                 data-key="{{ u.key }}" onclick="pickUniverse(this)">
              <span class="name">{{ u.label }}</span>
              <span class="meta">{{ u.count }} stocks<br>{{ u.time }}</span>
            </div>
            {% endfor %}
          </div>
        </div>
      </div>

      <div class="card">
        <div class="card-head">
          <span>SECTORS</span>
          <span style="font-size:0.62rem;color:var(--acc)">All = no filter</span>
        </div>
        <div class="card-body">
          <div class="sect-actions">
            <button class="sect-btn" onclick="sectAll()">All</button>
            <button class="sect-btn" onclick="sectNone()">None</button>
          </div>
          <div class="sect-grid" id="sectGrid">
            {% for s in sectors %}
            <div class="sect-item" onclick="toggleSect(this)">
              <input type="checkbox" checked>
              <span>{{ s }}</span>
            </div>
            {% endfor %}
          </div>
        </div>
      </div>

      <div class="run-wrap">
        <button class="run-btn" id="runBtn" onclick="runScreener()">▶ RUN ALL 6 SCREENERS</button>
      </div>

      <div style="margin-top:10px;font-family:var(--mono);font-size:0.6rem;
                  color:var(--mut);line-height:1.8;padding:10px 0">
        ⚡ Tip: Use SP500_TOP50 for a quick 20-min run.<br>
        The 6 screeners run sequentially — come back when done.<br>
        Results auto-save to Excel when the run finishes.
      </div>
    </div>

    <!-- ── RIGHT: Results ── -->
    <div>

      <!-- Status bar -->
      <div class="status-row">
        <span class="badge badge-idle" id="statusBadge">IDLE</span>
        <span class="server-dot dot-chk" id="serverDot"></span>
        <span class="server-lbl" id="serverLbl">Checking SmartStock...</span>
        <span class="timer" id="timerEl"></span>
        <button class="export-btn" id="exportBtn" onclick="doExport()" disabled>⬇ Export Excel</button>
      </div>

      <!-- Screener status pills -->
      <div class="screen-pills" id="screenPills">
        <span class="screen-pill" id="pill_Combined">⚡📊 Combined</span>
        <span class="screen-pill" id="pill_SEPA Minervini">🏆 SEPA</span>
        <span class="screen-pill" id="pill_Triple Screen">⚡📊🏆 Triple</span>
        <span class="screen-pill" id="pill_Swing Pro">📈 Swing Pro</span>
        <span class="screen-pill" id="pill_US Investing Championship">🥇 USIC</span>
        <span class="screen-pill" id="pill_Fixed Triple">🔧 Fixed Triple</span>
      </div>

      <!-- Log terminal -->
      <div class="card">
        <div class="card-head">
          <span>EXECUTION LOG</span>
          <span id="logCount" style="color:var(--mut)"></span>
        </div>
        <div class="log-body" id="logBody">
          <span class="ll info">// Ready. Configure universe and sectors, then press RUN.</span>
        </div>
      </div>

      <!-- Final 5 -->
      <div class="card">
        <div class="card-head">
          <span>★ FINAL TOP 5 PICKS</span>
          <span style="font-family:var(--mono);font-size:0.6rem;color:var(--mut)">Quality score + Entry trigger</span>
        </div>
        <div id="top5Wrap">
          <div class="empty"><div class="empty-icon">◎</div>Run the screener to see your top 5 picks</div>
        </div>
      </div>

      <!-- Cross-screener table -->
      <div class="card">
        <div class="card-head">
          <span>CROSS-SCREENER RANKINGS</span>
          <span style="font-family:var(--mono);font-size:0.6rem;color:var(--mut)">All stocks · top 20</span>
        </div>
        <div id="crossWrap">
          <div class="empty" style="padding:30px"><div class="empty-icon" style="font-size:1.5rem">◉</div>Results appear here</div>
        </div>
      </div>

    </div>
  </div>
</div>

<script>
// ── State ─────────────────────────────────────────────────────────────────────
var _universe = 'SP500_TOP100';
var _sectors  = [];   // empty = all
var _running  = false;
var _done     = false;
var _es       = null;
var _timer    = null;
var _startTs  = null;

// ── Universe picker ───────────────────────────────────────────────────────────
function pickUniverse(el) {
  document.querySelectorAll('.univ-pill').forEach(function(p){ p.classList.remove('sel'); });
  el.classList.add('sel');
  _universe = el.getAttribute('data-key');
  document.getElementById('univLabel').textContent = el.querySelector('.name').textContent;
}

// ── Sector picker ─────────────────────────────────────────────────────────────
function toggleSect(el) {
  var cb = el.querySelector('input');
  cb.checked = !cb.checked;
  el.classList.toggle('on', cb.checked);
}
function sectAll()  { document.querySelectorAll('.sect-item').forEach(function(el){ el.querySelector('input').checked=true; el.classList.add('on'); }); }
function sectNone() { document.querySelectorAll('.sect-item').forEach(function(el){ el.querySelector('input').checked=false; el.classList.remove('on'); }); }
function getSelectedSectors() {
  var checked=[];
  document.querySelectorAll('.sect-item').forEach(function(el){
    if (el.querySelector('input').checked) checked.push(el.querySelector('span').textContent.trim());
  });
  var total = document.querySelectorAll('.sect-item').length;
  return checked.length === total ? [] : checked;   // empty = all
}

// ── Server check ──────────────────────────────────────────────────────────────
function checkServer() {
  fetch('/api/check_smartstock').then(function(r){ return r.json(); }).then(function(d){
    var dot = document.getElementById('serverDot');
    var lbl = document.getElementById('serverLbl');
    if (d.ok) {
      dot.className = 'server-dot dot-ok';
      lbl.textContent = 'SmartStock connected (localhost:5000)';
    } else {
      dot.className = 'server-dot dot-err';
      lbl.textContent = 'SmartStock not found — start app.py first';
    }
  }).catch(function(){
    document.getElementById('serverDot').className = 'server-dot dot-err';
    document.getElementById('serverLbl').textContent = 'SmartStock not reachable';
  });
}
checkServer();
setInterval(checkServer, 15000);

// ── Run screener ──────────────────────────────────────────────────────────────
function runScreener() {
  if (_running) return;
  _running = true; _done = false;
  _sectors = getSelectedSectors();

  document.getElementById('runBtn').disabled    = true;
  document.getElementById('runBtn').classList.add('running');
  document.getElementById('runBtn').textContent = '⟳ RUNNING...';
  document.getElementById('statusBadge').className  = 'badge badge-run';
  document.getElementById('statusBadge').textContent = 'RUNNING';
  document.getElementById('exportBtn').disabled = true;
  document.getElementById('logBody').innerHTML  = '';
  document.getElementById('top5Wrap').innerHTML = '<div class="empty"><div class="empty-icon">⟳</div>Screeners running...</div>';
  document.getElementById('crossWrap').innerHTML= '<div class="empty" style="padding:30px">Computing...</div>';

  // Reset pills
  document.querySelectorAll('.screen-pill').forEach(function(p){ p.className='screen-pill'; });

  // Start timer
  _startTs = Date.now();
  clearInterval(_timer);
  _timer = setInterval(updateTimer, 1000);

  fetch('/api/run', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({universe: _universe, sectors: _sectors})
  })
  .then(function(r){ return r.json(); })
  .then(function(d){
    if (d.error) { logLine(d.error,'error'); resetBtn(); return; }
    startSSE();
  })
  .catch(function(e){ logLine('Error: '+e.message,'error'); resetBtn(); });
}

function updateTimer() {
  if (!_startTs) return;
  var s  = Math.floor((Date.now()-_startTs)/1000);
  var m  = Math.floor(s/60); var ss = s%60;
  document.getElementById('timerEl').textContent = m+'m '+ss+'s elapsed';
}

function startSSE() {
  if (_es) _es.close();
  _es = new EventSource('/api/stream');
  _es.onmessage = function(e) {
    var d = JSON.parse(e.data);

    // Logs
    (d.logs || []).forEach(function(l){ logLine(l.msg, l.kind, l.screener); });

    // Screener pills
    for (var name in (d.screens||{})) {
      var pill = document.getElementById('pill_'+name);
      if (pill) pill.className = 'screen-pill ' + (d.screens[name].status || '');
    }

    // Top 5
    if (d.top5 && d.top5.length) renderTop5(d.top5);

    if (d.done) {
      _es.close();
      // Fetch full cross
      fetch('/api/state').then(function(r){return r.json();}).then(function(s){
        if (s.cross) renderCross(s.cross);
        if (s.top5 && s.top5.length) renderTop5(s.top5);
      });
      finishRun(true);
    }
  };
  _es.onerror = function(){ _es.close(); finishRun(false); };
}

function finishRun(success) {
  _running = false; _done = success;
  clearInterval(_timer);
  resetBtn(success);
  document.getElementById('statusBadge').className   = success ? 'badge badge-done' : 'badge badge-err';
  document.getElementById('statusBadge').textContent  = success ? 'DONE' : 'ERROR';
  if (success) document.getElementById('exportBtn').disabled = false;
}

function resetBtn(success) {
  var btn = document.getElementById('runBtn');
  btn.disabled = false;
  btn.classList.remove('running');
  btn.textContent = success ? '↺ RUN AGAIN' : '▶ RUN ALL 6 SCREENERS';
}

// ── Logging ───────────────────────────────────────────────────────────────────
function logLine(msg, kind, screener) {
  kind = kind || 'info';
  if (kind === 'progress') return;   // skip noisy progress bars
  var body = document.getElementById('logBody');
  var el   = document.createElement('span');
  el.className = 'll ' + kind;
  var prefix = screener ? '['+screener+'] ' : '';
  el.textContent = prefix + msg;
  body.appendChild(el);
  body.scrollTop = body.scrollHeight;
  document.getElementById('logCount').textContent = body.children.length + ' lines';
}

// ── Render Top 5 ──────────────────────────────────────────────────────────────
function renderTop5(top5) {
  if (!top5 || !top5.length) return;
  var html = '<div style="overflow-x:auto"><table class="f5-table"><thead><tr>'
    + '<th>#</th><th>Ticker</th><th>Company</th><th>Sector</th>'
    + '<th class="right">Quality</th><th>RS</th>'
    + '<th>Entry NOW?</th><th>Screeners</th><th>Action</th>'
    + '</tr></thead><tbody>';

  top5.forEach(function(r,i) {
    var rsCls  = (r.rs_rating||0)>=80 ? 'rs-pill rs-high' : 'rs-pill';
    var freqCls= (r.frequency||0)>=3  ? 'freq-pill freq-high' : 'freq-pill';
    var scrList= (r.screeners||[]).map(function(s){
      return '<span style="display:inline-block;font-family:var(--mono);font-size:0.56rem;'
        +'padding:1px 5px;border-radius:2px;background:rgba(0,149,255,0.08);'
        +'border:1px solid rgba(0,149,255,0.2);color:var(--acc2);margin:1px">'+s+'</span>';
    }).join(' ');
    var action = r.entry_triggered
      ? '<span class="action-buy">BUY NOW ✓</span>'
      : '<span class="action-watch">Watch — wait for entry</span>';
    var trig = r.entry_triggered
      ? '<span class="trig-yes">✓ YES</span>'
      : '<span class="trig-no">— NO</span>';
    var rankColor = i===0?'color:var(--gold)':'color:var(--acc)';

    html += '<tr>'
      + '<td><span class="rank-cell" style="'+rankColor+'">'+(i+1)+'</span></td>'
      + '<td><span class="ticker-cell">'+r.symbol+'</span></td>'
      + '<td><span class="name-cell" title="'+(r.name||'')+'">'+((r.name||'—').substring(0,22))+'</span></td>'
      + '<td style="font-size:0.72rem;color:var(--mut)">'+(r.sector||'—')+'</td>'
      + '<td class="score-cell">'+( r.score||0).toFixed(1)+'</td>'
      + '<td><span class="'+rsCls+'">'+(r.rs_rating||'—')+'</span></td>'
      + '<td>'+trig+'</td>'
      + '<td><span class="'+freqCls+'">'+(r.frequency||0)+'×</span> '+scrList+'</td>'
      + '<td>'+action+'</td>'
      + '</tr>'
      + '<tr><td colspan="9" style="padding:3px 12px 10px;font-family:var(--mono);'
      + 'font-size:0.6rem;color:var(--mut);border-bottom:1px solid var(--bd)">'
      + (r.trigger_note ? '→ '+r.trigger_note : '') + '</td></tr>';
  });
  html += '</tbody></table></div>';
  document.getElementById('top5Wrap').innerHTML = html;
}

// ── Render Cross Table ────────────────────────────────────────────────────────
function renderCross(cross) {
  var entries = Object.entries(cross).sort(function(a,b){ return b[1].score-a[1].score; }).slice(0,20);
  var screenNames = ['Combined','SEPA Minervini','Triple Screen','Swing Pro',
                     'US Investing Championship','Fixed Triple'];
  var shortNames  = ['Comb','SEPA','Triple','SwPro','USIC','FixTri'];

  var html = '<div style="overflow-x:auto"><table class="cross-table"><thead><tr>'
    + '<th>Ticker</th><th>Company</th><th class="right">Score</th><th>RS</th>'
    + '<th>Freq</th><th>Entry</th>';
  shortNames.forEach(function(n){ html += '<th>'+n+'</th>'; });
  html += '</tr></thead><tbody>';

  entries.forEach(function(entry, idx) {
    var sym = entry[0]; var c = entry[1];
    var rowStyle = idx<5 ? 'background:rgba(0,214,143,0.03)' : '';
    var rk = c.ranks || {};
    html += '<tr style="'+rowStyle+'">'
      + '<td style="font-weight:700;color:var(--acc);font-family:var(--mono)">'+sym+'</td>'
      + '<td style="color:var(--mut);font-size:0.65rem">'+((c.name||'').substring(0,20))+'</td>'
      + '<td style="text-align:right;font-weight:700;color:var(--gold)">'+((c.score||0).toFixed(1))+'</td>'
      + '<td><span class="rs-pill '+(c.rs_rating>=80?'rs-high':'')+'">'+c.rs_rating+'</span></td>'
      + '<td style="text-align:center"><span class="freq-pill '+(c.frequency>=3?'freq-high':'')+'">'
      +   c.frequency+'×</span></td>'
      + '<td style="text-align:center;color:'+(c.entry_triggered?'var(--acc)':'var(--mut)')
      +   ';font-weight:700">'+(c.entry_triggered?'✓':'—')+'</td>';
    screenNames.forEach(function(s){
      var rank = rk[s];
      if (rank === undefined || rank === '—') {
        html += '<td class="rank-none" style="text-align:center;color:var(--bd2)">—</td>';
      } else if (rank >= 90) {
        html += '<td style="text-align:center"><span class="rank-badge rank-rej">R</span></td>';
      } else {
        html += '<td style="text-align:center"><span class="rank-badge">#'+rank+'</span></td>';
      }
    });
    html += '</tr>';
  });
  html += '</tbody></table></div>';
  document.getElementById('crossWrap').innerHTML = html;
}

// ── Export ────────────────────────────────────────────────────────────────────
function doExport() {
  window.location.href = '/api/export';
}

// Init
sectAll();
</script>
</body>
</html>"""


if __name__ == "__main__":
    print(f"\n  SmartStock — Final 5 Picker")
    print(f"  Open http://localhost:{PORT}")
    print(f"  SmartStock must be running on localhost:5000\n")
    app.run(port=PORT, debug=False, threaded=True)