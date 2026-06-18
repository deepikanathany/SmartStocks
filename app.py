"""
app.py  —  SmartStock US Screener
Run with: python app.py
Then open: http://localhost:5000
"""

import time
import json
import io
import threading
import requests as _requests
from flask import Flask, render_template, request, jsonify, Response
try:
    from intermarket_engine import IntermarketEngine as _IME
    _im_us = _IME(market="US")
    _IM_AVAILABLE = True
except (ImportError, Exception):
    _IM_AVAILABLE = False
    print("[IM] Engine not available:", __import__("traceback").format_exc()[:200])
from swing_pro_engine import run_swing_pro
from smma_screener import run_smma_screen
from screener_engine import (
    fetch_tickers_for_exchanges,
    screen_stock,
    export_to_excel,
    load_industry_pe_map,
    run_sepa_screen,
    run_triple_screen,
    run_fixed_triple_screen,
    get_market_regime,
    check_entry_trigger,
    get_top_etfs_for_ticker,          # ETF holder lookup — Final 5 table
)

# ── Safe import of golden_combo (needs groq package + golden_combo.py in same folder) ──
try:
    from golden_combo import run_golden_combo as _run_golden_combo
    _golden_combo_available = True
    _golden_combo_err = ""
except Exception as _gc_err:
    _golden_combo_available = False
    _golden_combo_err = str(_gc_err)

# ── USIC Screener ─────────────────────────────────────────────────────────────
try:
    from usic_screener import run_usic_screen as _run_usic_screen
    _usic_available = True
except Exception as _usic_err:
    _usic_available = False
    _usic_err_msg   = str(_usic_err)

# ── Minervini Value Ranker ────────────────────────────────────────────────────
try:
    from minervini_ranker import rank_sepa_results as _rank_sepa
    _ranker_available = True
except Exception as _rank_err:
    _ranker_available = False
    _rank_err_msg = str(_rank_err)

# ── JSON encoder — handles numpy/pandas types ──────────────────────────────────
class _SafeEncoder(json.JSONEncoder):
    def default(self, obj):
        try:
            import numpy as np
            if isinstance(obj, np.bool_):    return bool(obj)
            if isinstance(obj, np.integer):  return int(obj)
            if isinstance(obj, np.floating): return float(obj)
            if isinstance(obj, np.ndarray):  return obj.tolist()
        except ImportError:
            pass
        try:
            return float(obj)
        except (TypeError, ValueError):
            pass
        return str(obj)


def _dumps(obj):
    import re as _re
    raw = json.dumps(obj, cls=_SafeEncoder, allow_nan=True)
    raw = _re.sub(r'\bNaN\b',       'null', raw)
    raw = _re.sub(r'\bInfinity\b',  'null', raw)
    raw = _re.sub(r'\b-Infinity\b', 'null', raw)
    return raw


app        = Flask(__name__)
_jobs      = {}
_jobs_lock = threading.Lock()


def _compute_ai_score(info):
    """Sector-agnostic earnings acceleration + analyst conviction. Score 0-10."""
    score = 0
    rev = info.get("revenueGrowth")
    if rev is not None:
        if rev > 0.15: score += 3
        elif rev > 0.08: score += 1
    eg = info.get("earningsGrowth")
    if eg is not None:
        if eg > 0.20: score += 2
        elif eg > 0.10: score += 1
    price  = info.get("regularMarketPrice") or info.get("currentPrice")
    target = info.get("targetMeanPrice")
    if price and target and float(price) > 0:
        upside = (float(target) - float(price)) / float(price) * 100
        if upside > 10: score += 2
        elif upside > 0: score += 1
    if info.get("recommendationKey","") in ("strong_buy","buy"): score += 1
    tpe = info.get("trailingPE"); fpe = info.get("forwardPE")
    if tpe and fpe and float(tpe) > 0 and float(fpe) < float(tpe) * 0.85: score += 1
    pm = info.get("profitMargins")
    if pm and float(pm) > 0.10: score += 1
    return min(score, 10)

# ── Groq config ────────────────────────────────────────────────────────────────
try:
    from groq import Groq as _Groq
    GROQ_API_KEY  = "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"
    GROQ_MODEL    = "llama-3.3-70b-versatile"
    _groq_client  = _Groq(api_key=GROQ_API_KEY)
    _groq_available = True
except ImportError:
    _groq_available = False
    _groq_client    = None
    GROQ_MODEL      = ""


def _fetch_news_content(ticker_symbol: str, max_articles: int = 3) -> str:
    import yfinance as yf
    from bs4 import BeautifulSoup
    try:
        news = yf.Ticker(ticker_symbol).news or []
    except Exception:
        return "No news available."
    if not news:
        return "No recent news found."
    articles_text = []
    for item in news[:max_articles]:
        title     = item.get("title", "No title")
        url       = item.get("link") or item.get("url", "")
        publisher = item.get("publisher", "")
        body = ""
        if url:
            try:
                r = _requests.get(url, timeout=8,
                    headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                             "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"},
                    allow_redirects=True)
                if r.ok and "text/html" in r.headers.get("Content-Type", ""):
                    soup = BeautifulSoup(r.text, "html.parser")
                    for tag in soup(["script","style","nav","header","footer","aside","form","iframe","noscript"]):
                        tag.decompose()
                    article_tag = (soup.find("article") or
                                   soup.find(class_=lambda c: c and "article" in c.lower()) or
                                   soup.find("main") or soup.body)
                    if article_tag:
                        body = " ".join(article_tag.get_text(separator=" ", strip=True).split())[:2000]
            except Exception:
                pass
        if body:
            articles_text.append(f"[{publisher}] {title}\n{body}")
        else:
            articles_text.append(f"[{publisher}] {title}\n(Full text unavailable — headline only)")
    return "\n\n---\n\n".join(articles_text) if articles_text else "No news available."


def _groq_verdict(stock: dict) -> dict:
    ticker = stock.get("ticker", "")

    if not _groq_available:
        return {"ticker": ticker, "verdict": "N/A",
                "reason": "groq package not installed — run: pip install groq", "error": True}

    name    = stock.get("name", ticker)
    sector  = stock.get("sector", "Unknown")
    mktcap  = stock.get("market_cap_b", "N/A")
    pe      = stock.get("pe", "N/A")
    div     = stock.get("div_yield_pct", "N/A")
    de      = stock.get("debt_equity_pct", "N/A")
    monthly = stock.get("monthly_chg_pct", "N/A")
    weekly  = stock.get("weekly_chg_pct", "N/A")
    analyst = stock.get("analyst_label", "N/A")
    roce    = stock.get("roce", "N/A")
    roe     = stock.get("roe", "N/A")
    pe_vi   = stock.get("pe_vs_industry", "N/A")
    failed  = stock.get("failed_filters", [])

    rsi         = stock.get("rsi")
    ema_trend   = stock.get("ema_trend")
    vol_ratio   = stock.get("vol_ratio")
    atr         = stock.get("atr")
    macd_signal = stock.get("macd_signal")
    bb_pos      = stock.get("price_vs_bb")
    swing_score = stock.get("swing_score")
    swing_sigs  = stock.get("swing_signals", [])

    ema_labels  = {2:"Price>EMA20>EMA50 (strong uptrend)",1:"Price>EMA20 (mild uptrend)",
                   0:"Mixed (no clear trend)",-1:"Price<EMA20<EMA50 (downtrend)"}
    macd_labels = {1:"Bullish crossover (buy signal)",0:"No crossover",-1:"Bearish crossover"}

    has_swing  = swing_score is not None
    swing_block = ""
    if has_swing:
        swing_block = f"""
=== SWING TRADING TECHNICALS (1-2 week hold) ===
RSI(14):        {rsi if rsi is not None else 'N/A'}
EMA Trend:      {ema_labels.get(ema_trend, 'N/A') if ema_trend is not None else 'N/A'}
Volume Ratio:   {f'{vol_ratio}x 20-day avg' if vol_ratio is not None else 'N/A'}
ATR(14):        {f'{atr}% of price' if atr is not None else 'N/A'}
MACD Signal:    {macd_labels.get(macd_signal, 'N/A') if macd_signal is not None else 'N/A'}
BB Position:    {f'{bb_pos}% (0=lower band, 100=upper band)' if bb_pos is not None else 'N/A'}
Swing Score:    {swing_score}/10  — signals: {'; '.join(swing_sigs) if swing_sigs else 'none'}
================================================"""

    news_content = _fetch_news_content(ticker)
    screener_status = (
        f"This stock FAILED quantitative screening on: {', '.join(failed)}. "
        "Assess whether recent news/sentiment outweighs these weaknesses."
        if failed else
        "This stock PASSED all quantitative screening filters."
    )

    if has_swing:
        task = """You are a swing trading analyst. The investor wants to BUY and sell for a 2-3% profit within 1-2 weeks.
1. Read the technicals AND news — is there a near-term catalyst in the next 2 weeks?
2. Is the technical setup aligned for a move up? (RSI entry zone, EMA trend, MACD crossover, volume)
3. Identify the 1-2 most significant SHORT-TERM RISKS
4. Identify the 1-2 most significant SHORT-TERM STRENGTHS
5. Give a verdict: Strong Buy / Buy / Hold / Avoid — for a 1-2 week swing trade.
6. One sentence (max 25 words) on WHY — mention the key technical signal or catalyst."""
    else:
        task = """You are a concise equity analyst evaluating this stock for medium-term value investing.
1. Read the news articles carefully — identify material events, earnings surprises, management changes.
2. Identify the 1-2 most significant RISKS (combining news + fundamentals).
3. Identify the 1-2 most significant STRENGTHS.
4. Give a verdict: Strong Buy / Buy / Hold / Avoid.
5. One sentence (max 25 words) explaining the verdict."""

    prompt = f"""{task}

Stock: {name} ({ticker})
Sector: {sector}
Market Cap: ${mktcap}B | P/E: {pe}x | Div Yield: {div}% | D/E: {de}%
Monthly Delta: {monthly}% | Weekly Delta: {weekly}% | Analyst: {analyst}
ROCE: {roce}% | ROE: {roe}% | P/E vs Industry: {pe_vi}x
{swing_block}
Screener status: {screener_status}

=== LATEST NEWS ===
{news_content}
===================

Respond in this exact JSON format, no markdown, no extra text:
{{"verdict": "Buy", "reason": "Your one sentence here.", "risks": "Key risk.", "strengths": "Key strength.", "news_summary": "One sentence on what the news says."}}"""

    for attempt in range(3):
        try:
            response = _groq_client.chat.completions.create(
                model=GROQ_MODEL,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3,
                max_tokens=400,
            )
            raw    = response.choices[0].message.content.strip()
            raw    = raw.replace("```json", "").replace("```", "").strip()
            parsed = json.loads(raw)
            return {
                "ticker":       ticker,
                "verdict":      parsed.get("verdict", "Hold"),
                "reason":       parsed.get("reason", ""),
                "risks":        parsed.get("risks", ""),
                "strengths":    parsed.get("strengths", ""),
                "news_summary": parsed.get("news_summary", ""),
                "error":        False,
            }
        except Exception as ex:
            if attempt == 2:
                return {"ticker": ticker, "verdict": "N/A", "reason": str(ex)[:80], "error": True}
            time.sleep(5)
    return {"ticker": ticker, "verdict": "N/A", "reason": "Groq rate limit — retry shortly", "error": True}


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route("/api/ai_verdicts", methods=["POST"])
def ai_verdicts():
    data     = request.get_json()
    stocks   = data.get("results", [])
    verdicts = []
    for s in stocks:
        verdicts.append(_groq_verdict(s))
        if len(stocks) > 1:
            time.sleep(4)
    return jsonify({"verdicts": verdicts})


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/run", methods=["POST"])
def run_screen():
    import uuid
    cfg    = request.get_json()
    job_id = str(uuid.uuid4())[:8]
    with _jobs_lock:
        _jobs[job_id] = {"log": [], "results": [], "rejections": [], "done": False, "error": None}
    t = threading.Thread(target=_run_job, args=(job_id, cfg), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})


@app.route("/api/progress/<job_id>")
def stream_progress(job_id):
    def generate():
        import time as _t
        sent = 0
        while True:
            with _jobs_lock:
                job = _jobs.get(job_id)
            if not job:
                yield f"data: {_dumps({'error': 'Job not found'})}\n\n"; break
            lines = job["log"]
            while sent < len(lines):
                yield f"data: {_dumps(lines[sent])}\n\n"; sent += 1
            if job["done"]:
                yield f"data: {_dumps({'done': True, 'results': job['results']})}\n\n"; break
            _t.sleep(0.25)
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/job_status/<job_id>")
def job_status(job_id):
    with _jobs_lock:
        job = _jobs.get(job_id, {})
    return Response(_dumps({
        "done":       job.get("done", False),
        "error":      job.get("error"),
        "results":    job.get("results", []),
        "rejections": job.get("rejections", []),
        "log":        job.get("log", []),
    }), mimetype="application/json")


@app.route("/api/get_rejections/<job_id>")
def get_rejections(job_id):
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        return Response(_dumps({"rejections": [], "error": "Job not found"}), mimetype="application/json")
    return Response(_dumps({"rejections": job.get("rejections", [])}), mimetype="application/json")


@app.route("/api/export", methods=["POST"])
def export():
    data    = request.get_json()
    results = data.get("results", [])
    config  = data.get("config", {})
    buf     = io.BytesIO()
    export_to_excel(results, buf, config)
    buf.seek(0)
    return Response(buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=alpha_screen_results.xlsx"})


@app.route("/api/export_rejections", methods=["POST"])
def export_rejections():
    data       = request.get_json()
    rejections = data.get("rejections", [])
    lines      = ["Symbol,Company,Sector,Filter,Status,Actual Value,Threshold,Active"]
    for r in rejections:
        ticker = r.get("ticker", "")
        name   = '"' + r.get("name", "").replace('"', "'") + '"'
        sector = '"' + r.get("sector", "").replace('"', "'") + '"'
        for s in r.get("scorecard", []):
            status = "PASS" if s.get("passed") else ("SKIP" if not s.get("active") else "FAIL")
            row = [ticker, name, sector, s.get("filter",""), status,
                   s.get("actual",""), s.get("threshold",""),
                   "Yes" if s.get("active") else "No"]
            lines.append(",".join(f'"{v}"' if "," in str(v) else str(v) for v in row))
    return Response("\n".join(lines), mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=alpha_screen_rejections.csv"})


@app.route("/api/funds", methods=["POST"])
def fetch_funds():
    """Fetch ETF/REIT data for the Funds tab."""
    import yfinance as yf
    import concurrent.futures as _cf
    data     = request.get_json() or {}
    category = data.get("category", "all")

    FUND_UNIVERSE = {
        "dividend": [
            "SCHD","VYM","DVY","HDV","DGRO","VIG","NOBL","SDY","SPHD","SPYD",
            "FDVV","DGRW","RDVY","FVD","PFF","JEPI","JEPQ","QYLD","XYLD","RYLD"
        ],
        "sector": [
            "SPY","QQQ","IWM","DIA","VOO","VTI","VGT","XLK","XLF","XLE",
            "XLV","XLI","XLB","XLC","XLY","XLP","XLU","XLRE","ARKK","SOXX",
            "SMH","IBB","GLD","SLV","USO","TLT","HYG","LQD","IEMG","EFA"
        ],
        "reit": [
            "VNQ","SCHH","IYR","REM","MORT","O","AMT","PLD","EQIX","PSA",
            "DLR","SPG","AVB","EQR","WPC","STAG","NNN","VICI","MPW","IIPR"
        ],
        "bond": [
            "TLT","IEF","SHY","BND","AGG","LQD","HYG","JNK","TIP","VTIP",
            "VCIT","VCSH","MUB","EMB","BNDX","BSV","BIV","BLV","SCHO","SCHR"
        ],
    }

    if category == "all":
        tickers = []
        for v in FUND_UNIVERSE.values():
            for t in v:
                if t not in tickers:
                    tickers.append(t)
    else:
        tickers = FUND_UNIVERSE.get(category, [])

    cat_map = {}
    for cat, syms in FUND_UNIVERSE.items():
        for s in syms:
            cat_map[s] = cat

    def _fetch_one(sym):
        try:
            tk   = yf.Ticker(sym)
            info = tk.info or {}
            if not info or len(info) < 3:
                return None

            hist = None
            try:
                hist = tk.history(period="1y", interval="1d")
            except Exception:
                pass

            price      = info.get("regularMarketPrice") or info.get("currentPrice") or info.get("navPrice")
            prev_close = info.get("regularMarketPreviousClose")
            w52_high   = info.get("fiftyTwoWeekHigh")
            w52_low    = info.get("fiftyTwoWeekLow")
            aum        = info.get("totalAssets")
            expense    = info.get("annualReportExpenseRatio") or info.get("expenseRatio")
            div_yield  = info.get("yield") or info.get("dividendYield") or 0
            volume     = info.get("regularMarketVolume") or info.get("averageVolume")
            name       = info.get("longName") or info.get("shortName") or sym
            fund_type  = info.get("quoteType","ETF")
            category_label = info.get("category") or cat_map.get(sym,"—").replace("_"," ").title()

            ytd_ret = yr_ret = day_ret = None
            if hist is not None and not hist.empty and price:
                try:
                    from datetime import datetime
                    import pandas as pd
                    now = datetime.today()
                    ytd_start = hist[hist.index >= pd.Timestamp(f"{now.year}-01-01", tz=hist.index.tz)]["Close"]
                    if len(ytd_start):
                        ytd_ret = round((price - float(ytd_start.iloc[0])) / float(ytd_start.iloc[0]) * 100, 2)
                    if len(hist) >= 252:
                        yr_ret = round((price - float(hist["Close"].iloc[-252])) / float(hist["Close"].iloc[-252]) * 100, 2)
                    elif len(hist) > 1:
                        yr_ret = round((price - float(hist["Close"].iloc[0])) / float(hist["Close"].iloc[0]) * 100, 2)
                except Exception:
                    pass

            if prev_close and price:
                day_ret = round((price - prev_close) / prev_close * 100, 2)

            from screener_engine import _get_swing_technicals
            swing = _get_swing_technicals(tk)

            dist_freq = "—"
            try:
                divs = tk.dividends
                if divs is not None and len(divs) >= 2:
                    import pandas as pd
                    divs = divs.sort_index()
                    gaps = divs.index.to_series().diff().dropna().dt.days
                    avg_gap = gaps.mean()
                    if avg_gap < 40:    dist_freq = "Monthly"
                    elif avg_gap < 100: dist_freq = "Quarterly"
                    elif avg_gap < 200: dist_freq = "Semi-Annual"
                    else:               dist_freq = "Annual"
            except Exception:
                pass

            if not price:
                return None

            pct_from_high = round((price - w52_high) / w52_high * 100, 2) if w52_high and price else None

            return {
                "ticker":         sym,
                "name":           name[:45],
                "fund_category":  cat_map.get(sym, "sector"),
                "category_label": category_label,
                "fund_type":      fund_type,
                "price":          round(float(price), 2),
                "day_chg_pct":    day_ret,
                "w52_high":       round(float(w52_high), 2) if w52_high else None,
                "w52_low":        round(float(w52_low), 2) if w52_low else None,
                "pct_from_high":  pct_from_high,
                "div_yield_pct":  round(float(div_yield) * 100, 2) if div_yield else 0,
                "dist_freq":      dist_freq,
                "expense_ratio":  round(float(expense) * 100, 3) if expense else None,
                "aum_b":          round(float(aum) / 1e9, 2) if aum else None,
                "volume":         int(volume) if volume else None,
                "ytd_ret":        ytd_ret,
                "yr_ret":         yr_ret,
                "swing_score":    swing["swing_score"],
                "swing_signals":  swing["swing_signals"],
                "rsi":            swing["rsi"],
                "ema_trend":      swing["ema_trend"],
                "macd_signal":    swing["macd_signal"],
            }
        except Exception:
            return None

    results = []
    with _cf.ThreadPoolExecutor(max_workers=8) as ex:
        futures = {ex.submit(_fetch_one, sym): sym for sym in tickers}
        for fut in _cf.as_completed(futures):
            r = fut.result()
            if r:
                results.append(r)

    results.sort(key=lambda x: x["div_yield_pct"] or 0, reverse=True)
    return Response(_dumps({"funds": results, "count": len(results)}),
                    mimetype="application/json")


# ── Universe catalogue ────────────────────────────────────────────────────────
@app.route("/api/universes", methods=["GET"])
def list_universes():
    """Return the full universe catalogue so the frontend can build the selector UI."""
    catalogue = {
        "groups": [
            {
                "id": "sp500",
                "label": "S&P 500",
                "options": [
                    {"key": "SP500_TOP100", "label": "Top 100",      "sublabel": "mega cap",  "count": 100},
                    {"key": "SP500_TOP200", "label": "Top 200",      "sublabel": "",           "count": 200},
                    {"key": "SP500",        "label": "Full S&P 500", "sublabel": "",           "count": 503},
                ],
            },
            {
                "id": "nasdaq",
                "label": "NASDAQ",
                "options": [
                    {"key": "NASDAQ100",    "label": "NASDAQ 100",   "sublabel": "QQQ",        "count": 100},
                    {"key": "NASDAQ",       "label": "Full NASDAQ",  "sublabel": "",           "count": 3300},
                ],
            },
            {
                "id": "other",
                "label": "Other indices",
                "options": [
                    {"key": "DOWJONES",     "label": "Dow Jones 30", "sublabel": "",           "count": 30},
                    {"key": "RUSSELL1000",  "label": "Russell 1000", "sublabel": "large+mid",  "count": 1000},
                    {"key": "RUSSELL2000",  "label": "Russell 2000", "sublabel": "small cap",  "count": 2000},
                    {"key": "NYSE",         "label": "NYSE",         "sublabel": "",           "count": 3000},
                ],
            },
            {
                "id": "sector",
                "label": "By sector (S&P 500 stocks)",
                "options": [
                    {"key": "SECTOR_TECH",        "label": "Technology",   "sublabel": "",    "count": 60},
                    {"key": "SECTOR_HEALTHCARE",  "label": "Healthcare",   "sublabel": "",    "count": 60},
                    {"key": "SECTOR_FINANCIALS",  "label": "Financials",   "sublabel": "",    "count": 60},
                    {"key": "SECTOR_ENERGY",      "label": "Energy",       "sublabel": "",    "count": 40},
                    {"key": "SECTOR_CONSUMER",    "label": "Consumer",     "sublabel": "",    "count": 50},
                    {"key": "SECTOR_INDUSTRIALS", "label": "Industrials",  "sublabel": "",    "count": 40},
                ],
            },
        ]
    }
    return jsonify(catalogue)



# ── SEPA Screener ──────────────────────────────────────────────────────────────
@app.route("/api/run_sepa", methods=["POST"])
def run_sepa():
    import uuid
    cfg    = request.get_json() or {}
    job_id = str(uuid.uuid4())[:8]
    with _jobs_lock:
        _jobs[job_id] = {"log": [], "results": [], "rejections": [], "done": False, "error": None}
    t = threading.Thread(target=_run_sepa_job, args=(job_id, cfg), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})


def _run_sepa_job(job_id, cfg):
    import time as _time
    try:
        exchanges      = cfg.get("exchanges", ["SP500_TOP100"])
        custom_tickers = cfg.get("customTickers", [])
        thresholds     = cfg.get("thresholds", {})

        _log(job_id, "SEPA Screen — Minervini Trend Template + Fundamentals + VCP", "title")

        if custom_tickers:
            tickers = list(dict.fromkeys(custom_tickers))
            _log(job_id, f"Custom list: {len(tickers)} tickers", "info")
        else:
            _log(job_id, f"Universe: {', '.join(exchanges)}", "info")
            tickers = fetch_tickers_for_exchanges(exchanges, lambda m: _log(job_id, m, "info"))
            _log(job_id, f"Universe: {len(tickers)} tickers", "ok")

        def sepa_log(m):
            if m.startswith("__sepa_progress__"):
                parts = m.split("__")
                pct   = parts[2] if len(parts) > 2 else "0"
                sym   = parts[3] if len(parts) > 3 else ""
                found = parts[4] if len(parts) > 4 else "0"
                _log(job_id, f"__progress__{pct}__{sym}__{found}", "progress_raw")
            else:
                _log(job_id, m, "info")

        results, rejections = run_sepa_screen(tickers, thresholds, log_cb=sepa_log)

        # Compute AI catalyst score for each passing stock
        import yfinance as _yf
        for r in results:
            try:
                _info      = _yf.Ticker(r["ticker"]).info or {}
                r["ai_score"] = _compute_ai_score(_info)
            except Exception:
                r["ai_score"] = None

        # Sanitize and store
        results    = [_sanitize(r) for r in results]
        rejections = [_sanitize(r) for r in rejections]

        for r in results:
            with _jobs_lock:
                _jobs[job_id]["log"].append({"match_result": r, "kind": "match_result"})
            _log(job_id,
                 f"PASS {r['ticker']} — RS:{r.get('rs_rating','?')} "
                 f"EPS:{r.get('eps_growth_pct','?')}% AI:{r.get('ai_score','?')}/10",
                 "match")

        results = _run_im_scoring(job_id, results, _log)
        with _jobs_lock:
            _jobs[job_id]["results"]    = results
            _jobs[job_id]["rejections"] = rejections
            _jobs[job_id]["done"]       = True

    except Exception as e:
        import traceback
        _log(job_id, f"SEPA error: {e}", "error")
        _log(job_id, traceback.format_exc(), "error")
        with _jobs_lock:
            _jobs[job_id]["done"]  = True
            _jobs[job_id]["error"] = str(e)



# ── Triple Screen — Swing → Value → SEPA ──────────────────────────────────────
@app.route("/api/run_triple", methods=["POST"])
def run_triple():
    import uuid
    cfg    = request.get_json() or {}
    job_id = str(uuid.uuid4())[:8]
    with _jobs_lock:
        _jobs[job_id] = {"log": [], "results": [], "rejections": [],
                         "agent_stats": {}, "done": False, "error": None}
    t = threading.Thread(target=_run_triple_job, args=(job_id, cfg), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})


def _run_triple_job(job_id, cfg):
    import time as _time
    try:
        exchanges      = cfg.get("exchanges", ["SP500_TOP100"])
        custom_tickers = cfg.get("customTickers", [])
        sectors        = cfg.get("sectors", [])
        swing_thresholds = cfg.get("swingThresholds", {})
        value_thresholds = cfg.get("valueThresholds", {})
        sepa_thresholds  = cfg.get("sepaThresholds", {})

        _log(job_id, "TRIPLE SCREEN — Swing → Value → SEPA Minervini", "title")
        _log(job_id, "Agent 1: Swing  |  Agent 2: Value  |  Agent 3: SEPA", "info")

        # Build universe from selected exchanges/custom list
        if custom_tickers:
            tickers = list(dict.fromkeys(custom_tickers))
            _log(job_id, f"Universe: {len(tickers)} custom tickers", "info")
        else:
            _log(job_id, f"Universe: {', '.join(exchanges)}", "info")
            tickers = fetch_tickers_for_exchanges(exchanges, lambda m: _log(job_id, m, "info"))
            _log(job_id, f"Universe: {len(tickers)} tickers loaded", "ok")

        # Build swing conditions — all IDs hardcoded (CONDITIONS lives in the frontend)
        _ALL_IDS = [
            "market_cap","div_yield","pe_ratio","debt_equity","monthly_trend",
            "weekly_trend","analyst_rating","roce","roe","pe_vs_industry",
            "swing_rsi","swing_ema","swing_volume","swing_macd","swing_atr",
        ]
        # Agent 1: only RSI + EMA (most permissive — volume/MACD/ATR removed)
        # This ensures stocks aren't filtered out just because they had low volume today
        _SWING_IDS = {"swing_rsi", "swing_ema"}           # removed: swing_volume, swing_macd, swing_atr
        _VALUE_IDS = {"market_cap", "div_yield", "pe_ratio", "debt_equity", "monthly_trend",
                      "weekly_trend", "analyst_rating", "roce", "roe", "pe_vs_industry"}

        # Use swingConditions from frontend if provided, else use defaults
        if "swingConditions" in cfg:
            swing_conds = cfg["swingConditions"]
        else:
            swing_conds = {cid: (cid in _SWING_IDS) for cid in _ALL_IDS}

        if "valueConditions" in cfg:
            value_conds = cfg["valueConditions"]
        else:
            value_conds = {cid: (cid in _VALUE_IDS) for cid in _ALL_IDS}

        # Default thresholds — loosened so Agent 1 actually passes stocks
        default_thresh = {
            "min_market_cap_b":    1,
            "min_div_yield_pct":   1,    # was 3% — almost no growth stock yields 3%
            "max_pe":              60,   # was 15 — kills all growth/tech stocks
            "max_debt_equity_pct": 150,  # was 20 — too strict for financials
            "max_analyst_mean":    3.5,  # was 2.5
            "min_roce_pct":        5,    # was 18
            "min_roe_pct":         5,    # was 15
            "max_pe_vs_industry":  2.0,  # was 1.0
            "swing_rsi_min":       25,   # RSI(14) between 25-75 (pullback zone)
            "swing_rsi_max":       75,
            "swing_min_vol_ratio": 1.0,  # was 1.5 — 1.5x kills most days
            "swing_min_atr":       0.5,  # was 1.5
        }
        default_thresh.update(swing_thresholds)
        default_thresh.update(value_thresholds)

        default_sepa = {
            "sma200_min_days": 20, "min_pct_above_52w_low": 30,
            "max_pct_below_52w_high": 25, "min_rs_rating": 70,
            "min_eps_growth_pct": 20, "min_rev_growth_pct": 20,
            "vcp_contractions": 3,
        }
        default_sepa.update(sepa_thresholds)

        def triple_log(m):
            if m.startswith("__agent__"):
                # __agent__N__start__total
                parts = m.split("__")
                agent = parts[2] if len(parts) > 2 else "?"
                total = parts[4] if len(parts) > 4 else "?"
                _log(job_id, f"__agent_start__{agent}__{total}", "agent_event")
            elif m.startswith("__triple_progress__"):
                # __triple_progress__AGENT__PCT__SYM__FOUND
                parts = m.split("__")
                agent = parts[2] if len(parts) > 2 else "1"
                pct   = parts[3] if len(parts) > 3 else "0"
                sym   = parts[4] if len(parts) > 4 else ""
                found = parts[5] if len(parts) > 5 else "0"
                _log(job_id, f"__progress__{pct}__{sym}__{found}__agent{agent}", "progress_raw")
            else:
                _log(job_id, m, "info")

        result = run_triple_screen(
            tickers          = tickers,
            swing_conditions = swing_conds,
            swing_thresholds = default_thresh,
            value_conditions = value_conds,
            value_thresholds = default_thresh,
            sepa_thresholds  = default_sepa,
            allowed_sectors  = list(sectors) if sectors else None,
            log_cb           = triple_log,
        )

        final      = [_sanitize(r) for r in result["final"]]
        # All rejections combined across agents
        all_rejects = (
            [_sanitize(r) for r in result["agent1"]["rejected"]] +
            [_sanitize(r) for r in result["agent2"]["rejected"]] +
            [_sanitize(r) for r in result["agent3"]["rejected"]]
        )

        agent_stats = {
            "agent1_passed": len(result["agent1"]["passed"]),
            "agent2_passed": len(result["agent2"]["passed"]),
            "agent3_passed": len(result["agent3"]["passed"]),
            "total_started": len(tickers),
        }

        for r in final:
            with _jobs_lock:
                _jobs[job_id]["log"].append({"match_result": r, "kind": "match_result"})
            _log(job_id,
                 f"PASS {r['ticker']} — RS:{r.get('rs_rating','?')} "
                 f"Swing:{r.get('swing_score','?')} "
                 f"VCP:{r.get('vcp','?')}",
                 "match")

        with _jobs_lock:
            _jobs[job_id]["results"]     = final
            _jobs[job_id]["rejections"]  = all_rejects
            _jobs[job_id]["agent_stats"] = agent_stats
            _jobs[job_id]["done"]        = True

    except Exception as e:
        import traceback
        _log(job_id, f"Triple screen error: {e}", "error")
        _log(job_id, traceback.format_exc(), "error")
        with _jobs_lock:
            _jobs[job_id]["done"]  = True
            _jobs[job_id]["error"] = str(e)


# Also expose agent stats in the progress endpoint
@app.route("/api/triple_stats/<job_id>")
def triple_stats(job_id):
    with _jobs_lock:
        job = _jobs.get(job_id, {})
    return Response(_dumps({
        "agent_stats": job.get("agent_stats", {}),
        "done":        job.get("done", False),
    }), mimetype="application/json")



# ── Swing Pro ──────────────────────────────────────────────────────────────────
@app.route("/api/run_swing_pro", methods=["POST"])
def run_swing_pro_route():
    import uuid
    cfg    = request.get_json() or {}
    job_id = str(uuid.uuid4())[:8]
    with _jobs_lock:
        _jobs[job_id] = {"log":[], "results":[], "rejections":[], "done":False, "error":None}
    t = threading.Thread(target=_run_swing_pro_job, args=(job_id, cfg), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})


def _run_swing_pro_job(job_id, cfg):
    try:
        exchanges       = cfg.get("exchanges", ["SP500_TOP100"])
        custom_tickers  = cfg.get("customTickers", [])
        sectors         = cfg.get("sectors", [])
        strategies      = cfg.get("strategies", {
            "ema_reversion":   True,
            "cons_breakout":   True,
            "xlp_reversion":   True,
            "seasonal_nasdaq": True,
            "seasonal_tom":    True,
        })
        thresh          = cfg.get("thresholds", {})
        run_backtest    = cfg.get("run_backtest", True)
        run_sentiment   = cfg.get("run_sentiment", True)
        run_verdict     = cfg.get("run_verdict", True)

        _log(job_id, "📈 Swing Pro Screen starting...", "title")

        # Build universe from selected exchanges — same as every other mode
        if custom_tickers:
            tickers = list(dict.fromkeys(custom_tickers))
            _log(job_id, f"Universe: {len(tickers)} custom tickers", "info")
        else:
            _log(job_id, f"Universe: {', '.join(exchanges)}", "info")
            tickers = fetch_tickers_for_exchanges(exchanges, lambda m: _log(job_id, m, "info"))
            _log(job_id, f"Universe: {len(tickers)} tickers loaded", "ok")

        enabled = [k for k, v in strategies.items() if v]
        _log(job_id, f"Strategies: {', '.join(enabled)}", "info")
        _log(job_id, f"AI sentiment: {'on' if run_sentiment else 'off'} | "
                     f"AI verdict: {'on' if run_verdict else 'off'} | "
                     f"Backtest: {'on' if run_backtest else 'off'}", "info")

        reject_count_sp = [0]  # mutable counter for reject_count events

        def sp_log(m):
            if m.startswith("__progress__"):
                _log(job_id, m, "progress_raw")
            elif m.startswith("PASS "):
                _log(job_id, m, "match")
            elif m.startswith("Swing Pro error"):
                _log(job_id, m, "error")
            elif "rejected" in m.lower() or "skipped" in m.lower():
                _log(job_id, m, "info")
            else:
                _log(job_id, m, "info")

        def sp_reject_cb(rejection):
            """Called each time a stock is rejected — emits reject_count event."""
            reject_count_sp[0] += 1
            with _jobs_lock:
                _jobs[job_id]["log"].append({
                    "kind": "reject_count",
                    "count": reject_count_sp[0],
                })

        output = run_swing_pro(
            tickers         = tickers,
            strategies      = strategies,
            thresh          = thresh,
            allowed_sectors = list(sectors) if sectors and len(sectors) > 0 else None,
            run_backtest    = run_backtest,
            run_sentiment   = run_sentiment,
            run_verdict     = run_verdict,
            log_cb          = sp_log,
            reject_cb       = sp_reject_cb,
        )

        results    = [_sanitize(r) for r in output["results"]]
        rejections = [_sanitize(r) for r in output["rejections"]]

        for r in results:
            with _jobs_lock:
                _jobs[job_id]["log"].append({"match_result": r, "kind": "match_result"})

        results = _run_im_scoring(job_id, results, _log)
        with _jobs_lock:
            _jobs[job_id]["results"]    = results
            _jobs[job_id]["rejections"] = rejections
            _jobs[job_id]["done"]       = True

    except Exception as e:
        import traceback
        _log(job_id, f"Swing Pro error: {e}", "error")
        _log(job_id, traceback.format_exc(), "error")
        with _jobs_lock:
            _jobs[job_id]["done"]  = True
            _jobs[job_id]["error"] = str(e)


# ── Golden Combo Flip Analysis ─────────────────────────────────────────────────
@app.route("/api/golden_combo", methods=["POST"])
def golden_combo():
    """
    Golden Combo Flip Analysis — 5 AI agents.
    Requires golden_combo.py in the same folder and 'groq' package installed.
    """
    import traceback
    try:
        if not _golden_combo_available:
            return Response(
                _dumps({"error": f"Could not load golden_combo.py: {_golden_combo_err}. "
                        "Make sure golden_combo.py is in the same folder as app.py, "
                        "and run:  pip install groq"}),
                mimetype="application/json"
            )

        data        = request.get_json() or {}
        stocks      = data.get("stocks", [])
        target_date = data.get("target_date", "")
        min_filters = int(data.get("min_filters", 7))

        if not stocks:
            return Response(
                _dumps({"error": "No stocks found. Please run a screen first, then open Golden Combo."}),
                mimetype="application/json"
            )
        if not target_date:
            return Response(
                _dumps({"error": "Please pick a target date."}),
                mimetype="application/json"
            )

        result = _run_golden_combo(
            all_stocks  = stocks,
            target_date = target_date,
            min_filters = min_filters,
            market      = "US",
        )
        return Response(_dumps(result), mimetype="application/json")

    except Exception as e:
        return Response(
            _dumps({"error": str(e), "detail": traceback.format_exc()}),
            mimetype="application/json"
        )


# ── Internal helpers ───────────────────────────────────────────────────────────
def _sanitize(obj):
    return json.loads(_dumps(obj))


def _run_im_scoring(job_id, results, log_fn):
    """Attach IM scores to results just before done=True."""
    if not _IM_AVAILABLE or not results:
        return results
    try:
        log_fn(job_id, "IM: Fetching US market basket...", "info")
        basket = _im_us.fetch_basket()
        def _im_log(m): log_fn(job_id, m, "info")
        _im_us.score_results(results, basket, log_cb=_im_log)
        log_fn(job_id, "IM: Scoring complete", "ok")
    except Exception as e:
        log_fn(job_id, f"IM error (non-fatal): {e}", "warn")
    return results


def _log(job_id, msg, kind="info"):
    with _jobs_lock:
        _jobs[job_id]["log"].append({"msg": msg, "kind": kind})


def _run_job(job_id, cfg):
    import time as _time
    try:
        exchanges      = cfg.get("exchanges", ["SP500_TOP100"])
        custom_tickers = cfg.get("customTickers", [])
        sectors        = cfg.get("sectors", [])
        conditions     = cfg.get("conditions", {})
        thresholds     = cfg.get("thresholds", {})

        _log(job_id, "Alpha Screen starting...", "title")
        if custom_tickers:
            _log(job_id, f"Mode: Custom list — {len(custom_tickers)} tickers", "info")
        else:
            _log(job_id, f"Universe: {', '.join(exchanges)}", "info")
        _log(job_id, f"Sectors: {', '.join(sectors) if sectors else 'All bullish sectors'}", "info")
        if "NYSE" in exchanges:
            _log(job_id, "NYSE selected — universe ~3000 stocks. May take 30-90 min.", "info")

        load_industry_pe_map(log_cb=lambda m: _log(job_id, m, "info"))

        _log(job_id, "PHASE 1 — Building ticker universe", "phase")
        if custom_tickers:
            tickers = list(dict.fromkeys(custom_tickers))  # deduplicate, preserve order
            _log(job_id, f"Custom list: {len(tickers)} tickers", "ok")
        else:
            tickers = fetch_tickers_for_exchanges(exchanges, lambda m: _log(job_id, m, "info"))
            _log(job_id, f"Universe: {len(tickers)} unique tickers", "ok")

        survivors = tickers
        if sectors:
            _log(job_id, f"PHASE 2 — Sector filter active ({len(sectors)} sectors, checked inline)", "phase")
        else:
            _log(job_id, "PHASE 2 — No sector filter (all sectors selected)", "info")

        _log(job_id, "PHASE 3 — Full screening", "phase")
        results       = []
        rejections    = []
        filter_counts = {}
        total         = len(survivors)

        for i, symbol in enumerate(survivors, 1):
            pct = int(i / total * 100)
            _log(job_id, f"__progress__{pct}__{symbol}__{len(results)}", "progress_raw")

            result, rejection = screen_stock(symbol, conditions, thresholds, sectors or [])
            _time.sleep(0.1)

            if result:
                result = _sanitize(result)
                results.append(result)
                with _jobs_lock:
                    _jobs[job_id]["log"].append({"match_result": result, "kind": "match_result"})
                _log(job_id, f"PASS {symbol} — {result['name']} passed all criteria", "match")
            elif rejection:
                rejection = _sanitize(rejection)
                rejections.append(rejection)
                with _jobs_lock:
                    _jobs[job_id]["log"].append({"kind": "reject_count", "count": len(rejections)})
                failed = rejection.get("failed_filters", [])
                for f in failed:
                    filter_counts[f] = filter_counts.get(f, 0) + 1
                if not set(failed) <= {"Error", "Data Fetch", "Sector"}:
                    _log(job_id, f"FAIL {symbol} ({rejection['name'][:22]}) — failed: {', '.join(failed)}", "reject")
                else:
                    filter_counts["__data_error"] = filter_counts.get("__data_error", 0) + 1

        _log(job_id, "-" * 48, "info")
        _log(job_id, f"DONE — {len(survivors)} screened · {len(results)} passed · {len(rejections)} rejected", "done_msg")
        data_errors = filter_counts.pop("__data_error", 0)
        if data_errors:
            _log(job_id, f"  {data_errors} tickers skipped — no data from Yahoo Finance", "info")
        if filter_counts:
            _log(job_id, "Rejections by filter:", "phase")
            for f, count in sorted(filter_counts.items(), key=lambda x: -x[1]):
                _log(job_id, f"  {f:20s} -> {count} stocks filtered out", "info")

        results = _run_im_scoring(job_id, results, _log)
        with _jobs_lock:
            _jobs[job_id]["results"]    = results
            _jobs[job_id]["rejections"] = rejections
            _jobs[job_id]["done"]       = True

    except Exception as e:
        _log(job_id, f"Fatal error: {e}", "error")
        with _jobs_lock:
            _jobs[job_id]["done"]  = True
            _jobs[job_id]["error"] = str(e)


# ── Value Ranker — rank SEPA results by value quality score ──────────────────
@app.route("/api/rank_sepa", methods=["POST"])
def rank_sepa():
    """
    Takes a list of SEPA-screened tickers + their SEPA data,
    fetches value metrics for each, scores using weighted value system,
    returns ranked list (best fundamental quality first).

    Request body:
        {
          "tickers":   ["AAPL", "MSFT", ...],   // required
          "sepa_data": { "AAPL": {rs_rating, eps_growth_pct, ...}, ... }  // optional
        }

    Response:
        { ranked: [...], top5: [...], knocked: [...], total, scored, weights }
    """
    import uuid
    if not _ranker_available:
        return Response(
            _dumps({"error": f"minervini_ranker.py not found: {_rank_err_msg}"}),
            mimetype="application/json"
        ), 500

    data      = request.get_json() or {}
    tickers   = data.get("tickers", [])
    sepa_data = data.get("sepa_data", {})

    if not tickers:
        return Response(
            _dumps({"error": "No tickers provided. Run SEPA screen first."}),
            mimetype="application/json"
        ), 400

    job_id = str(uuid.uuid4())[:8]
    with _jobs_lock:
        _jobs[job_id] = {"log": [], "done": False, "error": None, "rank_result": None}

    def _run():
        def log_cb(msg, kind="info"):
            _log(job_id, msg, kind)
        try:
            result = _rank_sepa(
                tickers    = tickers,
                sepa_data  = sepa_data,
                log_cb     = log_cb,
                max_workers = 6,
            )
            with _jobs_lock:
                _jobs[job_id]["rank_result"] = result
                _jobs[job_id]["done"]        = True
        except Exception as e:
            _log(job_id, f"Ranker error: {e}", "error")
            with _jobs_lock:
                _jobs[job_id]["done"]  = True
                _jobs[job_id]["error"] = str(e)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/api/rank_sepa_progress/<job_id>")
def rank_sepa_progress(job_id):
    """SSE stream for rank_sepa job — same pattern as /api/progress."""
    def generate():
        sent = 0
        while True:
            with _jobs_lock:
                job = _jobs.get(job_id)
            if not job:
                payload = _dumps({"error": "Job not found"})
                yield "data: " + payload + "\n\n"
                break
            lines = job["log"]
            while sent < len(lines):
                payload = _dumps(lines[sent])
                yield "data: " + payload + "\n\n"
                sent += 1
            if job["done"]:
                payload = _dumps({"done": True, "result": job.get("rank_result")})
                yield "data: " + payload + "\n\n"
                break
            time.sleep(0.25)
    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── USIC Screen route ─────────────────────────────────────────────────────────
@app.route("/api/run_usic", methods=["POST"])
def run_usic():
    import uuid
    if not _usic_available:
        return jsonify({"error": f"usic_screener.py not loaded: {_usic_err_msg}"}), 500
    cfg    = request.get_json() or {}
    job_id = str(uuid.uuid4())[:8]
    with _jobs_lock:
        _jobs[job_id] = {"log":[], "results":[], "rejections":[], "done":False, "error":None}
    threading.Thread(target=_run_usic_job, args=(job_id, cfg), daemon=True).start()
    return jsonify({"job_id": job_id})


def _run_usic_job(job_id, cfg):
    try:
        exchanges      = cfg.get("exchanges", ["SP500_TOP100"])
        custom_tickers = cfg.get("customTickers", [])
        thresholds     = cfg.get("thresholds", {})

        _log(job_id, "US Investing Championship Screen", "title")
        _log(job_id, "7 stages: Trend · Fundamentals · VCP · EPS Accel · EPS Surprise · Sector RS · Champion Score", "info")

        if custom_tickers:
            tickers = list(dict.fromkeys(custom_tickers))
            _log(job_id, f"Custom list: {len(tickers)} tickers", "info")
        else:
            _log(job_id, f"Universe: {', '.join(exchanges)}", "info")
            tickers = fetch_tickers_for_exchanges(exchanges, lambda m: _log(job_id, m, "info"))
            _log(job_id, f"Universe: {len(tickers)} tickers", "ok")

        def usic_log(m):
            if m.startswith("__usic_progress__"):
                parts = m.split("__")
                pct   = parts[2] if len(parts)>2 else "0"
                sym   = parts[3] if len(parts)>3 else ""
                found = parts[4] if len(parts)>4 else "0"
                _log(job_id, f"__progress__{pct}__{sym}__{found}", "progress_raw")
            else:
                _log(job_id, m, "info")

        results, rejections = _run_usic_screen(tickers, thresholds, log_cb=usic_log)

        # AI score via existing _compute_ai_score
        import yfinance as _yf
        for r in results:
            try:
                info = _yf.Ticker(r["ticker"]).info or {}
                r["ai_score"] = _compute_ai_score(info)
            except Exception:
                r["ai_score"] = None

        results    = [_sanitize(r) for r in results]
        rejections = [_sanitize(r) for r in rejections]

        for r in results:
            with _jobs_lock:
                _jobs[job_id]["log"].append({"match_result": r, "kind": "match_result"})
            _log(job_id,
                 f"✓ {r['ticker']} — {r.get('name','')[:28]} "
                 f"RS={r.get('rs_rating','?')} "
                 f"Champ={r.get('champion_score','?')}/10 "
                 f"EPS Accel={'✓' if r.get('eps_accelerating') else '✗'} "
                 f"Sector={'✓' if r.get('sector_outperforming') else '✗'}",
                 "match")

        results = _run_im_scoring(job_id, results, _log)
        with _jobs_lock:
            _jobs[job_id]["results"]    = results
            _jobs[job_id]["rejections"] = rejections
            _jobs[job_id]["done"]       = True

        _log(job_id, f"━"*48, "info")
        _log(job_id, f"USIC COMPLETE — {len(results)} stocks passed all 7 stages", "done_msg")

    except Exception as e:
        _log(job_id, f"Fatal error: {e}", "error")
        with _jobs_lock:
            _jobs[job_id]["done"]  = True
            _jobs[job_id]["error"] = str(e)


# ── Market Regime Gate ────────────────────────────────────────────────────────
@app.route("/api/market_regime", methods=["GET"])
def market_regime():
    market = request.args.get("market", "US")
    r = get_market_regime(market)
    return Response(_dumps(r), mimetype="application/json")


# ── Fixed Triple Screen ───────────────────────────────────────────────────────
@app.route("/api/run_fixed_triple", methods=["POST"])
def run_fixed_triple():
    import uuid
    cfg    = request.get_json() or {}
    job_id = str(uuid.uuid4())[:8]
    with _jobs_lock:
        _jobs[job_id] = {"log": [], "results": [], "rejections": [],
                         "agent_stats": {}, "done": False, "error": None}
    t = threading.Thread(target=_run_fixed_triple_job, args=(job_id, cfg), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})


def _run_fixed_triple_job(job_id, cfg):
    try:
        exchanges      = cfg.get("exchanges", ["SP500_TOP100"])
        custom_tickers = cfg.get("customTickers", [])
        sepa_thresh    = cfg.get("sepaThresholds", {})
        market         = cfg.get("market", "US")
        use_regime     = cfg.get("useRegimeGate", True)

        _log(job_id, "FIXED TRIPLE SCREEN — Technical → Fundamental → SEPA", "title")
        _log(job_id, "New: Regime Gate + RSI(25-75) + Sane PE threshold + Entry Trigger", "info")

        if custom_tickers:
            tickers = list(dict.fromkeys(custom_tickers))
        else:
            tickers = fetch_tickers_for_exchanges(exchanges, lambda m: _log(job_id, m, "info"))
            _log(job_id, f"Universe: {len(tickers)} tickers", "ok")

        def ft_log(m):
            if m.startswith("__agent__"):
                parts = m.split("__")
                agent = parts[2]; total = parts[4] if len(parts) > 4 else "?"
                labels = {"1": "Technical", "2": "Fundamental", "3": "SEPA"}
                _log(job_id, f"__agent__{agent}__start__{total}", "agent_event")
            elif m.startswith("__triple_progress__"):
                parts = m.split("__")
                agent = parts[2]; pct = parts[3]; sym = parts[4]; found = parts[5]
                _log(job_id, f"__progress__{pct}__{sym}__{found}", "progress_raw")
            else:
                _log(job_id, m, "info")

        result = run_fixed_triple_screen(
            tickers         = tickers,
            sepa_thresholds = sepa_thresh,
            market          = market,
            use_regime_gate = use_regime,
            log_cb          = ft_log,
        )

        final    = [_sanitize(r) for r in result.get("final", [])]
        regime   = result.get("regime", {})

        # Build unified rejections list with agent label and full scorecard
        all_rej = []
        for agent_num, key in [(1, "agent1"), (2, "agent2"), (3, "agent3")]:
            for r in result.get(key, {}).get("rejected", []):
                r2 = _sanitize(r)
                r2["_agent_rejected"] = agent_num
                r2["_agent_label"]    = {1:"Technical",2:"Fundamental",3:"SEPA"}[agent_num]
                all_rej.append(r2)

        for r in final:
            with _jobs_lock:
                _jobs[job_id]["log"].append({"match_result": r, "kind": "match_result"})
            _log(job_id,
                 f"✓ {r['ticker']} — {r.get('name','')[:25]} "
                 f"RS={r.get('rs_rating','?')} "
                 f"RSI={r.get('rsi','?')} "
                 f"Trigger={'✓' if r.get('entry_triggered') else '—'}",
                 "match")

        _log(job_id, f"Regime: {regime.get('regime','?')} "
             f"({regime.get('index','?')} {regime.get('pct_above','?')}% vs 200MA)", "info")
        _log(job_id, f"FIXED TRIPLE COMPLETE — {len(final)} stocks passed all 3 stages", "done_msg")

        with _jobs_lock:
            _jobs[job_id]["results"]    = final
            _jobs[job_id]["rejections"] = all_rej
            _jobs[job_id]["agent_stats"] = {
                "total_started":  len(tickers),
                "agent1_passed":  len(result.get("agent1",{}).get("passed",[])),
                "agent2_passed":  len(result.get("agent2",{}).get("passed",[])),
                "agent3_passed":  len(final),
                "regime":         regime,
            }
            _jobs[job_id]["done"] = True

    except Exception as e:
        _log(job_id, f"Fatal: {e}", "error")
        with _jobs_lock:
            _jobs[job_id]["done"]  = True
            _jobs[job_id]["error"] = str(e)


# ── US SMMA Screen ───────────────────────────────────────────────────────────
@app.route("/api/run_smma", methods=["POST"])
def run_smma():
    import uuid
    cfg    = request.get_json() or {}
    job_id = str(uuid.uuid4())[:8]
    with _jobs_lock:
        _jobs[job_id] = {"log": [], "results": [], "rejections": [],
                         "done": False, "error": None}
    t = threading.Thread(target=_run_smma_job, args=(job_id, cfg), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})


def _run_smma_job(job_id, cfg):
    try:
        exchanges      = cfg.get("exchanges", ["SP500_TOP100"])
        custom_tickers = cfg.get("customTickers", [])

        _log(job_id, "US SMMA Screen — 5-min crossover + 15-min + 1-hour HTF", "title")
        _log(job_id, "Filters: RS≥65 · UDTS≥40 · Vol dry-up+expansion 11-13 EST", "info")
        _log(job_id, "Entry: 10:30 AM–2:30 PM · Dead zone 1:00–1:30 PM", "info")

        if custom_tickers:
            tickers = list(dict.fromkeys(custom_tickers))
        else:
            tickers = fetch_tickers_for_exchanges(
                exchanges, lambda m: _log(job_id, m, "info"))
            _log(job_id, f"Universe: {len(tickers)} tickers", "ok")

        def smma_log(m):
            _log(job_id, m, "info")

        result  = run_smma_screen(tickers, log_cb=smma_log, sessions=5)
        results = [_sanitize(r) for r in result.get("results", [])]
        rejects = [_sanitize(r) for r in result.get("rejections", [])]

        for r in results:
            with _jobs_lock:
                _jobs[job_id]["log"].append({"match_result": r, "kind": "match_result"})
            _log(job_id,
                 f"✓ {r['ticker']} — RS:{r.get('rs_rating','?')} "
                 f"UDTS:{r.get('udts_score','?')} "
                 f"Entry:${r.get('entry','?')} "
                 f"{'[ENTRY ACTIVE]' if r.get('entry_triggered') else ''}",
                 "match")

        # Intermarket scoring
        results = _run_im_scoring(job_id, results, _log)

        _log(job_id,
             f"US SMMA complete — {len(results)} setups found "
             f"({sum(1 for r in results if r.get('entry_triggered'))} entry active)",
             "done_msg")

        with _jobs_lock:
            _jobs[job_id]["results"]    = results
            _jobs[job_id]["rejections"] = rejects
            _jobs[job_id]["done"]       = True

    except Exception as e:
        import traceback
        _log(job_id, f"Fatal: {e}", "error")
        _log(job_id, traceback.format_exc(), "error")
        with _jobs_lock:
            _jobs[job_id]["done"]  = True
            _jobs[job_id]["error"] = str(e)


# ── Export to Excel (results + rejects with full scorecard) ──────────────────
@app.route("/api/export_excel", methods=["POST"])
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    data       = request.get_json() or {}
    results    = data.get("results", [])
    rejections = data.get("rejections", [])
    config     = data.get("config", {})

    wb  = openpyxl.Workbook()
    hdr_fill   = PatternFill("solid", fgColor="0D1520")
    hdr_font   = Font(bold=True, color="00D68F", size=10)
    pass_fill  = PatternFill("solid", fgColor="0A2018")
    fail_fill  = PatternFill("solid", fgColor="200808")
    skip_fill  = PatternFill("solid", fgColor="0D1520")
    pass_font  = Font(color="00D68F", bold=True)
    fail_font  = Font(color="FF4444", bold=True)
    skip_font  = Font(color="444444")
    thin       = Side(style="thin", color="1D2730")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def header_row(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill   = hdr_fill
            cell.font   = hdr_font
            cell.border = border
            cell.alignment = center
        ws.row_dimensions[row].height = 20

    def autofit(ws, max_w=40):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2, 8), max_w)

    # ── Sheet 1: Results ──────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Results"
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:M1")
    ws1["A1"] = f"SmartStock / AlphaScreener — Results ({len(results)} stocks)"
    ws1["A1"].font = Font(bold=True, color="00D68F", size=12)
    ws1["A1"].fill = PatternFill("solid", fgColor="050B10")
    ws1["A1"].alignment = left

    hdrs1 = ["Ticker","Company","Sector","Exchange","Price","Mo.Δ%","Wk.Δ%",
             "⚡ Swing","📊 Value","RS","RSI","EMA Trend","Vol Ratio","P/E","Div%","AI Verdict"]
    header_row(ws1, hdrs1, 2); ws1.freeze_panes = "A3"

    for i, r in enumerate(results, 3):
        et = {2:"Bull ▲▲", 1:"Bull ▲", 0:"Mixed", -1:"Bear ▼"}.get(r.get("ema_trend"), "—")
        vals = [
            r.get("ticker",""), r.get("name",""), r.get("sector",""), r.get("exchange",""),
            r.get("price_now"), r.get("monthly_chg_pct"), r.get("weekly_chg_pct"),
            r.get("swing_score"), r.get("value_score"),
            r.get("rs_rating"), r.get("rsi"), et, r.get("vol_ratio"),
            r.get("pe"), r.get("div_yield_pct"), r.get("analyst_label",""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=c, value=v)
            cell.fill   = PatternFill("solid", fgColor="0A1A12") if i%2==0 else PatternFill("solid", fgColor="050B10")
            cell.border = border
            cell.alignment = center

    autofit(ws1)

    # ── Sheet 2: Rejects — Full Scorecard ────────────────────────────────────
    ws2 = wb.create_sheet("Rejects — Full Scorecard")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:E1")
    ws2["A1"] = f"Rejected Stocks — Full Scorecard ({len(rejections)} stocks)"
    ws2["A1"].font = Font(bold=True, color="FF7055", size=12)
    ws2["A1"].fill = PatternFill("solid", fgColor="1A0808")
    ws2["A1"].alignment = left

    # Collect all unique filter names across all rejections for dynamic columns
    all_filters = []
    for r in rejections:
        for s in (r.get("scorecard") or []):
            if s.get("filter") not in all_filters:
                all_filters.append(s["filter"])

    hdrs2 = ["Ticker","Company","Sector","Price","Rejected At Stage","Failed Filters"] + all_filters
    header_row(ws2, hdrs2, 2); ws2.freeze_panes = "A3"

    for i, r in enumerate(rejections, 3):
        sc       = {s["filter"]: s for s in (r.get("scorecard") or [])}
        failed   = r.get("failed_filters") or []
        n_failed = len([s for s in (r.get("scorecard") or []) if not s.get("passed") and s.get("active") != False])
        stage    = r.get("_agent_label") or ("Stage " + str(r.get("_agent_rejected","?")))

        base_vals = [
            r.get("ticker",""), r.get("name",""), r.get("sector",""),
            r.get("price_now"), stage, ", ".join(failed)[:80],
        ]
        row_bg = PatternFill("solid", fgColor="180808") if n_failed >= 3 else PatternFill("solid", fgColor="100808")

        for c, v in enumerate(base_vals, 1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.fill = row_bg; cell.border = border; cell.alignment = left

        # Filter scorecard columns
        for j, fname in enumerate(all_filters, len(base_vals)+1):
            s = sc.get(fname)
            if s is None:
                cell = ws2.cell(row=i, column=j, value="—")
                cell.fill = skip_fill; cell.font = skip_font
            elif s.get("active") == False:
                cell = ws2.cell(row=i, column=j, value=f"SKIP")
                cell.fill = skip_fill; cell.font = skip_font
            elif s.get("passed"):
                cell = ws2.cell(row=i, column=j, value=f"✓ {s.get('actual','')}")
                cell.fill = pass_fill; cell.font = pass_font
            else:
                cell = ws2.cell(row=i, column=j, value=f"✗ {s.get('actual','')} (need {s.get('threshold','')})")
                cell.fill = fail_fill; cell.font = fail_font
            cell.border = border; cell.alignment = center

    autofit(ws2)

    # ── Sheet 3: Rejects Summary (sorted by failed count) ────────────────────
    ws3 = wb.create_sheet("Rejects Summary")
    ws3.sheet_view.showGridLines = False
    hdrs3 = ["Ticker","Company","Sector","Stage Rejected","# Failed","# Passed","Failed Filters"]
    header_row(ws3, hdrs3, 1)

    rej_summary = []
    for r in rejections:
        sc     = r.get("scorecard") or []
        passed = len([s for s in sc if s.get("passed") and s.get("active") != False])
        failed = len([s for s in sc if not s.get("passed") and s.get("active") != False])
        rej_summary.append((r, failed, passed))

    rej_summary.sort(key=lambda x: x[1], reverse=True)
    for i, (r, n_fail, n_pass) in enumerate(rej_summary, 2):
        sc     = {s["filter"]: s for s in (r.get("scorecard") or [])}
        failed_names = ", ".join(r.get("failed_filters") or [])
        stage  = r.get("_agent_label") or ("Stage " + str(r.get("_agent_rejected","?")))
        vals   = [r.get("ticker",""), r.get("name",""), r.get("sector",""),
                  stage, n_fail, n_pass, failed_names]
        bg = PatternFill("solid", fgColor="180808" if n_fail>=3 else "0D1520")
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=i, column=c, value=v)
            cell.fill = bg; cell.border = border; cell.alignment = left

    autofit(ws3)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import send_file
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="alphascreener_results.xlsx")



# ── NOVICE SCREENER ──────────────────────────────────────────────────────────
import threading as _nv_threading
import json as _nv_json
import time as _nv_time

_novice_jobs = {}
_novice_lock = _nv_threading.Lock()

@app.route("/api/novice_search", methods=["POST"])
def novice_search():
    import requests as _req
    query = (request.get_json() or {}).get("query", "").strip()
    if not query or len(query) < 2:
        return jsonify({"results": []})
    try:
        url = ("https://query2.finance.yahoo.com/v1/finance/search"
               "?q=" + _req.utils.quote(query) + "&quotesCount=10&newsCount=0&listsCount=0")
        resp = _req.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=6)
        quotes = resp.json().get("quotes", [])
        results = []
        for q in quotes:
            if q.get("quoteType") != "EQUITY": continue
            if q.get("exchange") not in ("NYQ","NMS","NGM","NCM","NYSEArca"): continue
            sym = q.get("symbol", "")
            if not sym or "." in sym: continue
            price = None
            try:
                raw = q.get("regularMarketPrice") or q.get("regularMarketPreviousClose")
                if raw: price = round(float(raw), 2)
            except Exception: pass
            results.append({"symbol": sym,
                            "name": q.get("longname") or q.get("shortname") or sym,
                            "price": price})
            if len(results) >= 5: break
        return jsonify({"results": results})
    except Exception as e:
        return jsonify({"results": [], "error": str(e)[:80]})

@app.route("/api/novice", methods=["POST"])
def novice_run():
    data = request.get_json() or {}
    sym  = data.get("symbol", "").upper().strip()
    name = data.get("name", sym)
    if not sym: return jsonify({"error": "symbol required"}), 400
    import uuid
    jid = uuid.uuid4().hex[:8]
    with _novice_lock:
        _novice_jobs[jid] = {"log": [], "result": None, "done": False}
    _nv_threading.Thread(target=_run_novice_job, args=(jid, sym, name), daemon=True).start()
    return jsonify({"job_id": jid})

@app.route("/api/novice_progress/<jid>")
def novice_progress(jid):
    def _safe_json(obj):
        """Convert numpy types to plain Python before JSON serialisation."""
        import math
        if isinstance(obj, dict):
            return {k: _safe_json(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [_safe_json(v) for v in obj]
        # numpy floats / ints
        try:
            import numpy as np
            if isinstance(obj, (np.floating, np.integer)):
                v = obj.item()
                return None if (math.isnan(v) or math.isinf(v)) else v
            if isinstance(obj, np.bool_):
                return bool(obj)
        except ImportError:
            pass
        # plain float NaN/Inf
        if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
            return None
        return obj

    def gen():
        sent = 0
        while True:
            with _novice_lock:
                job = _novice_jobs.get(jid, {})
            if not job:
                yield "data:" + _nv_json.dumps({"error": "not found"}) + "\n\n"
                break
            while sent < len(job.get("log", [])):
                yield "data:" + _nv_json.dumps(job["log"][sent]) + "\n\n"
                sent += 1
            if job.get("done"):
                try:
                    result_clean = _safe_json(job.get("result"))
                    payload = _nv_json.dumps({"done": True, "result": result_clean})
                except Exception as je:
                    payload = _nv_json.dumps({"done": True, "result": None,
                                              "error": "serialise error: " + str(je)[:80]})
                yield "data:" + payload + "\n\n"
                break
            # keepalive comment so browser doesn't drop the connection
            yield ": keepalive\n\n"
            _nv_time.sleep(0.3)
    return Response(gen(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

def _run_novice_job(jid, symbol, company_name):
    import yfinance as yf
    import numpy as np
    from datetime import date, timedelta
    import requests as _req2
    import xml.etree.ElementTree as ET

    def log(msg, kind="info"):
        with _novice_lock:
            _novice_jobs[jid]["log"].append({"msg": msg, "kind": kind})

    try:
        log("Fetching data for {} ({})...".format(company_name, symbol), "phase")
        df = yf.Ticker(symbol).history(period="2y", interval="1d", auto_adjust=True)
        if df is None or df.empty or len(df) < 10:
            log("No data found for {}.".format(symbol), "error")
            with _novice_lock: _novice_jobs[jid]["done"] = True
            return

        cl = df["Close"]; hi = df["High"]; lo = df["Low"]; vol = df["Volume"]
        price = round(float(cl.iloc[-1]), 2)

        def pct(a, b): return round((a/b-1)*100, 2) if b else 0
        r1d = pct(cl.iloc[-1], cl.iloc[-2])  if len(cl)>=2   else 0
        r5d = pct(cl.iloc[-1], cl.iloc[-6])  if len(cl)>=6   else 0
        r1m = pct(cl.iloc[-1], cl.iloc[-22]) if len(cl)>=22  else 0
        r6m = pct(cl.iloc[-1], cl.iloc[-126])if len(cl)>=126 else 0
        w52h = round(float(hi.max()), 2); w52l = round(float(lo.min()), 2)

        ema20  = round(float(cl.ewm(span=20,  adjust=False).mean().iloc[-1]), 2)
        ema50  = round(float(cl.ewm(span=50,  adjust=False).mean().iloc[-1]), 2)
        ema200 = round(float(cl.ewm(span=200, adjust=False).mean().iloc[-1]), 2) if len(cl)>=200 else None

        delta = cl.diff()
        gain  = delta.clip(lower=0).rolling(14).mean()
        loss  = (-delta.clip(upper=0)).rolling(14).mean()
        rsi   = round(float(100 - 100/(1 + gain.iloc[-1]/max(loss.iloc[-1], 1e-9))), 1)

        avg_vol20 = float(vol.iloc[-20:].mean()) if len(vol)>=20 else float(vol.mean())
        vol_ratio = round(float(vol.iloc[-1])/avg_vol20, 2) if avg_vol20>0 else 1.0

        try:
            info     = yf.Ticker(symbol).info or {}
            pe_raw   = info.get("trailingPE") or info.get("forwardPE")
            pe       = round(float(pe_raw), 1) if pe_raw else None
            raw_div  = float(info.get("dividendYield", 0) or 0)
            div_yield= round(raw_div*100, 2) if raw_div < 0.20 else round(raw_div, 2)
            mktcap_b = round(info.get("marketCap",0)/1e9, 1) if info.get("marketCap") else None
            sector   = info.get("sector", "Unknown")
            eps      = float(info.get("trailingEps") or 0)
        except Exception as ie:
            log("Info note: {}".format(ie), "info")
            pe = div_yield = mktcap_b = None; sector = "Unknown"; eps = 0

        SECTOR_PE = {"Technology":28,"Communication Services":22,"Consumer Discretionary":25,
                     "Consumer Staples":20,"Health Care":22,"Financials":14,"Industrials":21,
                     "Energy":12,"Utilities":17,"Real Estate":35,"Materials":18}
        sector_pe_avg = SECTOR_PE.get(sector, 20)

        pe_vs_sector = ""
        if pe and sector_pe_avg:
            diff_pct = round((pe - sector_pe_avg)/sector_pe_avg*100, 1)
            if diff_pct > 20:
                pe_vs_sector = "PE {} vs {} sector avg {} - stock is {}% MORE expensive than peers".format(pe, sector, sector_pe_avg, diff_pct)
            elif diff_pct < -15:
                pe_vs_sector = "PE {} vs {} sector avg {} - stock is {}% CHEAPER than peers".format(pe, sector, sector_pe_avg, abs(diff_pct))
            else:
                pe_vs_sector = "PE {} vs {} sector avg {} - priced in line with peers".format(pe, sector, sector_pe_avg)

        log("${} RSI:{} EMA20:{} PE:{} SectorPE:{}".format(
            round(float(price),2), round(float(rsi),1), round(float(ema20),2), pe, sector_pe_avg), "info")

        # Swing score
        sw=0; sw_good=[]; sw_warn=[]
        if price>ema20>ema50: sw+=2; sw_good.append("Price above both moving averages - uptrend")
        elif price>ema20:     sw+=1; sw_good.append("Price above short-term average - mild uptrend")
        else: sw_warn.append("Price below moving averages - downtrend short term")
        if 40<=rsi<=65:  sw+=2; sw_good.append("RSI {} in sweet spot".format(rsi))
        elif rsi<40:     sw+=1; sw_good.append("RSI {} - oversold, possible bounce".format(rsi))
        elif rsi>70: sw_warn.append("RSI {} - overbought, may pull back".format(rsi))
        if r5d>1:    sw+=1; sw_good.append("Up {:.1f}% this week - good momentum".format(r5d))
        elif r5d<-3: sw_warn.append("Down {:.1f}% this week - weak short-term".format(abs(r5d)))
        if vol_ratio>=1.2: sw+=1; sw_good.append("Volume {:.1f}x normal - strong activity".format(vol_ratio))
        elif vol_ratio<0.7: sw_warn.append("Very low volume - limited conviction")
        dist_high = round((w52h-price)/w52h*100, 1)
        if dist_high<=8:   sw+=1; sw_good.append("Only {}% below 52-week high".format(dist_high))
        elif dist_high>=30: sw_warn.append("Down {}% from yearly high".format(dist_high))
        if r1m>2:   sw+=1; sw_good.append("Up {:.1f}% this month".format(r1m))
        elif r1m<-5: sw_warn.append("Down {:.1f}% this month".format(abs(r1m)))
        sw = min(sw, 10)
        sw_verdict = "GOOD" if sw>=7 else "MIXED" if sw>=4 else "RISKY"

        swing_buy_zone = None
        if sw_verdict in ("RISKY","MIXED"):
            cands = []
            if ema20<price: cands.append(("EMA20", round(ema20,2)))
            if ema50<price: cands.append(("EMA50", round(ema50,2)))
            if ema200 and ema200<price: cands.append(("EMA200", round(ema200,2)))
            valid = [(l,v) for l,v in cands if (price-v)/price*100>=0.5]
            if valid:
                valid.sort(key=lambda x: price-x[1])
                lb,lv = valid[0]
                swing_buy_zone = {"price":lv,"reason":lb,"desc":"${} ({}) - moving average acting as floor".format(lv,lb)}
            elif w52l>0:
                fb = round(w52l*1.05,2)
                swing_buy_zone = {"price":fb,"reason":"52w low","desc":"${} - near 52-week low of ${}, where buyers tend to step in".format(fb,w52l)}

        # Value score
        va=0; va_good=[]; va_warn=[]
        if pe is not None:
            if pe<15:   va+=2; va_good.append("PE {} - bargain (pay ${} per $1 the company earns)".format(pe,pe))
            elif pe<25: va+=1; va_good.append("PE {} - fairly priced (pay ${} per $1 earned)".format(pe,pe))
            elif pe<=40: va_warn.append("PE {} - pricey (paying ${} per $1 earned)".format(pe,pe))
            else: va_warn.append("PE {} - expensive".format(pe))
        else: va_warn.append("No PE ratio - company may not yet be profitable")
        if div_yield>=2:    va+=2; va_good.append("Dividend {}% - company pays you {}% per year just for holding!".format(div_yield,div_yield))
        elif div_yield>=0.5: va+=1; va_good.append("Small dividend {}%".format(div_yield))
        else: va_warn.append("No dividend - profits reinvested in growth")
        if ema200:
            if price>ema200: va+=2; va_good.append("Above 200-day average ${} - long-term uptrend".format(ema200))
            else: va_warn.append("Below 200-day average ${} - long-term trend down".format(ema200))
        if r6m>5:    va+=1; va_good.append("Up {:.1f}% in 6 months - growing".format(r6m))
        elif r6m<-15: va_warn.append("Down {:.1f}% in 6 months - tough stretch".format(abs(r6m)))
        if mktcap_b and mktcap_b>=50: va+=1; va_good.append("Market cap ${}B - large stable company".format(mktcap_b))
        elif mktcap_b and mktcap_b<5: va_warn.append("Small cap ${}B - higher risk and reward".format(mktcap_b))
        va = min(va, 10)
        va_verdict = "GOOD" if va>=7 else "MIXED" if va>=4 else "RISKY"

        value_buy = None
        if va_verdict in ("RISKY","MIXED"):
            vb_tech = vb_val = None
            tcands = []
            if ema200 and ema200<price: tcands.append(("200-day avg", round(ema200,2)))
            if ema50  and ema50 <price: tcands.append(("50-day avg",  round(ema50, 2)))
            if tcands:
                tcands.sort(key=lambda x: price-x[1])
                tl,tv = tcands[0]
                vb_tech = {"price":tv,"desc":"${} ({}) - long-term support where patient buyers enter".format(tv,tl)}
            if eps and eps>0 and sector_pe_avg:
                fv = round(eps*sector_pe_avg, 2)
                pd = round(abs(fv-price)/price*100, 1)
                vb_val = {"price":fv,"desc":"${} - at this price PE equals {} sector avg of {}x ({}% {} current)".format(fv,sector,sector_pe_avg,pd,"below" if fv<price else "above")}
            value_buy = {"technical":vb_tech,"valuation":vb_val}

        # Pivots
        def pv(h,l,c):
            p=round((h+l+c)/3,2)
            return {"pivot":p,"r1":round(2*p-l,2),"r2":round(p+(h-l),2),"s1":round(2*p-h,2),"s2":round(p-(h-l),2)}
        def near(px,lv,pct=1.5): return abs(px-lv)/lv*100<=pct if lv else False

        pivots = {}
        if len(hi)>=5:   pivots["weekly"]  = pv(float(hi.iloc[-5:].max()),  float(lo.iloc[-5:].min()),  float(cl.iloc[-5]))
        if len(hi)>=22:  pivots["monthly"] = pv(float(hi.iloc[-22:].max()), float(lo.iloc[-22:].min()), float(cl.iloc[-22]))
        if len(hi)>=252: pivots["yearly"]  = pv(float(hi.iloc[-252:].max()),float(lo.iloc[-252:].min()),float(cl.iloc[-252]))

        pivot_hits = []
        for tf,pvt in pivots.items():
            for lbl,lv in [("Pivot",pvt["pivot"]),("Support 1",pvt["s1"]),("Support 2",pvt["s2"]),("Resistance 1",pvt["r1"]),("Resistance 2",pvt["r2"])]:
                if not near(price,lv): continue
                is_sup = lbl.startswith("Support") or lbl=="Pivot"
                pivot_hits.append({"timeframe":tf,"level_name":lbl,"level":round(lv,2),"price":price,
                    "dist_pct":round(abs(price-lv)/lv*100,2),"vol_strong":vol_ratio>=1.15,
                    "vol_ratio":vol_ratio,"bounced_up":r1d>0 and is_sup,"r1d":r1d})

        ptxt = "\n".join("- {} {} at ${}: price {} this level, vol {:.1f}x".format(
            h["timeframe"],h["level_name"],h["level"],"bounced UP from" if h["bounced_up"] else "near",h["vol_ratio"])
            for h in pivot_hits) if pivot_hits else "Price is not near any key pivot level right now."

        # ── Chart data: 1 year of daily OHLCV + EMA20/EMA50 series + pivot levels ──
        log("Building price chart...", "phase")
        ema20_series = cl.ewm(span=20, adjust=False).mean()
        ema50_series = cl.ewm(span=50, adjust=False).mean()

        # Slice to last ~252 trading days (1 year) — EMAs computed on full 2y for accuracy
        chart_len = min(len(df), 252)
        chart_df  = df.iloc[-chart_len:]
        ema20_chart = ema20_series.iloc[-chart_len:]
        ema50_chart = ema50_series.iloc[-chart_len:]

        chart_candles = []
        chart_volume  = []
        chart_ema20   = []
        chart_ema50   = []
        for idx in range(len(chart_df)):
            row = chart_df.iloc[idx]
            t   = chart_df.index[idx].strftime("%Y-%m-%d")
            o,h_,l_,c_ = float(row["Open"]), float(row["High"]), float(row["Low"]), float(row["Close"])
            chart_candles.append({"time": t, "open": round(o,2), "high": round(h_,2),
                                   "low": round(l_,2), "close": round(c_,2)})
            chart_volume.append({"time": t, "value": float(row["Volume"]),
                                  "color": "#10b98155" if c_ >= o else "#ef444455"})
            chart_ema20.append({"time": t, "value": round(float(ema20_chart.iloc[idx]), 2)})
            chart_ema50.append({"time": t, "value": round(float(ema50_chart.iloc[idx]), 2)})

        # Pivot levels for horizontal lines — use monthly pivot (most relevant for 3mo view)
        chart_pivot_lines = []
        if "monthly" in pivots:
            pm = pivots["monthly"]
            chart_pivot_lines = [
                {"label": "R2", "value": pm["r2"], "color": "#ef4444"},
                {"label": "R1", "value": pm["r1"], "color": "#f59e0b"},
                {"label": "Pivot", "value": pm["pivot"], "color": "#818cf8"},
                {"label": "S1", "value": pm["s1"], "color": "#f59e0b"},
                {"label": "S2", "value": pm["s2"], "color": "#10b981"},
            ]

        chart_data = {
            "candles": chart_candles,
            "volume":  chart_volume,
            "ema20":   chart_ema20,
            "ema50":   chart_ema50,
            "pivot_lines": chart_pivot_lines,
        }

        # News
        headlines = []
        try:
            after = (date.today()-timedelta(days=7)).strftime("%Y-%m-%d")
            url = "https://news.google.com/rss/search?q="+_req2.utils.quote(company_name+" stock after:"+after)+"&hl=en-US&gl=US&ceid=US:en"
            nr = _req2.get(url, headers={"User-Agent":"Mozilla/5.0"}, timeout=6)
            if nr.ok:
                for item in ET.fromstring(nr.content).findall(".//item")[:4]:
                    t = item.findtext("title","").strip()
                    if len(t)>10: headlines.append(t)
        except Exception: pass
        if not headlines:
            try:
                for n in (yf.Ticker(symbol).news or [])[:3]:
                    if n.get("title"): headlines.append(n["title"])
            except Exception: pass
        news_block = "\n".join("- "+h for h in headlines[:4]) or "No recent news."

        log("Asking Groq AI for explanation...", "phase")

        def buy_txt(obj):
            if not obj: return "Not applicable."
            lines = []
            if obj.get("technical"):  lines.append("Technical: " + obj["technical"]["desc"])
            if obj.get("valuation"):  lines.append("Fair value: " + obj["valuation"]["desc"])
            return "\n".join(lines) if lines else "Not applicable."

        prompt = """You explain stocks to a teenager learning to invest. Be their cool older sibling - honest, encouraging, using real-life comparisons.

Company: {company} ({sym}) | Price: ${price} | Sector: {sector}

SWING (1 week-1 month): {sw}/10 -> {sw_v}
Good: {sw_good}
Warnings: {sw_warn}
Buy zone if waiting: {sw_buy}

VALUE (6 months-1 year): {va}/10 -> {va_v}
Good: {va_good}
Warnings: {va_warn}
Better buy prices: {va_buy}

PE vs Sector: {pe_cmp}
Key price levels: {pivots}
News: {news}

Write FOUR paragraphs for a smart 15-year-old:
1. SWING - should they buy for a quick 1 week-1 month trade? Use "X out of a possible 10 points" not "score of X". Use analogies. If RISKY/MIXED mention the buy zone.
2. VALUE - should they hold 6-12 months? Compare PE to sector average. Explain PE as "paying $X for every $1 the company earns". Mention dividend if any. If RISKY/MIXED mention the better buy prices.
3. KEY PRICE LEVELS - explain what a support/resistance level is (use a real-world analogy like a floor or magnet). Then explain what the levels above mean for this stock. Mention volume.
4. YOUR OVERALL TAKE - honest view for a young investor. Worth watching, hold-and-wait, or skip for now?

End with one KEY TAKEAWAY sentence.
Return ONLY valid JSON no markdown:
{{"swing_explanation":"...","value_explanation":"...","pivot_explanation":"...","overall_take":"...","key_takeaway":"..."}}""".format(
            company=company_name, sym=symbol, price=price, sector=sector,
            sw=sw, sw_v=sw_verdict, sw_good="; ".join(sw_good) or "none", sw_warn="; ".join(sw_warn) or "none",
            sw_buy=buy_txt(swing_buy_zone),
            va=va, va_v=va_verdict, va_good="; ".join(va_good) or "none", va_warn="; ".join(va_warn) or "none",
            va_buy=buy_txt(value_buy),
            pe_cmp=pe_vs_sector or "N/A", pivots=ptxt, news=news_block)

        sw_expl=va_expl=pv_expl=overall=takeaway=""
        try:
            from groq import Groq as _Groq
            gc = _Groq(api_key=GROQ_API_KEY)
            raw = gc.chat.completions.create(
                model=GROQ_MODEL,
                messages=[{"role":"user","content":prompt}],
                temperature=0.4, max_tokens=1100
            ).choices[0].message.content.strip()
            p = _nv_json.loads(raw.replace("```json","").replace("```","").strip())
            sw_expl  = p.get("swing_explanation","")
            va_expl  = p.get("value_explanation","")
            pv_expl  = p.get("pivot_explanation","")
            overall  = p.get("overall_take","")
            takeaway = p.get("key_takeaway","")
            log("Groq explanation ready!", "ok")
        except Exception as ge:
            log("Groq note: {}".format(ge), "info")
            sw_expl  = "Swing: {}/10 ({}). ".format(sw,sw_verdict) + (sw_good[0] if sw_good else "")
            va_expl  = "Value: {}/10 ({}). ".format(va,va_verdict) + (va_good[0] if va_good else "")
            pv_expl  = ptxt; overall=""; takeaway="Always research before you invest!"

        # Sanitise all values before storing — ensures JSON serialisability
        def _clean(v):
            import math
            try:
                import numpy as np
                if isinstance(v, (np.floating, np.integer)):
                    x = v.item(); return None if (math.isnan(x) or math.isinf(x)) else x
                if isinstance(v, np.bool_): return bool(v)
            except ImportError: pass
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)): return None
            if isinstance(v, dict):  return {kk: _clean(vv) for kk, vv in v.items()}
            if isinstance(v, list):  return [_clean(i) for i in v]
            return v

        result = _clean(dict(symbol=symbol, company_name=company_name, price=price, sector=sector,
            mktcap_b=mktcap_b, pe=pe, div_yield=div_yield, rsi=rsi,
            ema20=ema20, ema50=ema50, ema200=ema200,
            r1d=r1d, r5d=r5d, r1m=r1m, r6m=r6m, w52h=w52h, w52l=w52l, vol_ratio=vol_ratio,
            sector_pe_avg=sector_pe_avg, pe_vs_sector=pe_vs_sector,
            swing_score=sw, swing_verdict=sw_verdict, swing_signals=sw_good, swing_flags=sw_warn,
            swing_explanation=sw_expl, swing_buy_zone=swing_buy_zone,
            value_score=va, value_verdict=va_verdict, value_signals=va_good, value_flags=va_warn,
            value_explanation=va_expl, value_buy=value_buy,
            pivot_hits=pivot_hits, pivot_explanation=pv_expl,
            overall_take=overall, key_takeaway=takeaway, headlines=headlines,
            chart_data=chart_data))

        log("Swing {}/10 {} . Value {}/10 {} . {} pivot hit(s)".format(sw,sw_verdict,va,va_verdict,len(pivot_hits)), "ok")

        with _novice_lock:
            _novice_jobs[jid]["result"] = result
            _novice_jobs[jid]["done"]   = True

    except Exception as e:
        import traceback
        tb = traceback.format_exc().splitlines()[-1]
        with _novice_lock:
            _novice_jobs[jid]["log"].append({"msg":"Error: {} - {}".format(e,tb),"kind":"error"})
            _novice_jobs[jid]["done"] = True



# ── SHORT TERM PICKS / LONG TERM PICKS ───────────────────────────────────────
import uuid as _uuid_mod

_picks_jobs  = {}
_picks_lock  = threading.Lock()
_picks_last_run_ts = 0   # unix timestamp of last completed picks run (for rate-limit cooldown)

# Time estimates per stock in seconds at 12 workers
_SECS_PER_STOCK = 1.8
_WORKERS_PICKS  = 12

UNIVERSE_SIZES = {
    "SP500":   500,
    "NASDAQ": 3500,
    "NYSE":   2800,
}

def _picks_time_estimate(exchanges):
    total = sum(UNIVERSE_SIZES.get(e, 0) for e in exchanges)
    # deduplicate overlap ~10%
    if len(exchanges) > 1:
        total = int(total * 0.9)
    import math
    batches = math.ceil(total / _WORKERS_PICKS)
    secs    = batches * _SECS_PER_STOCK
    mins    = round(secs / 60, 1)
    return {"stocks": total, "minutes": mins, "seconds": int(secs)}


@app.route("/api/picks_estimate", methods=["POST"])
def picks_estimate():
    exchanges = (request.get_json() or {}).get("exchanges", [])
    return jsonify(_picks_time_estimate(exchanges))


@app.route("/api/picks_run", methods=["POST"])
def picks_run():
    data      = request.get_json() or {}
    mode      = data.get("mode", "short")      # "short" or "long"
    exchanges = data.get("exchanges", [])
    if not exchanges:
        return jsonify({"error": "Select at least one exchange"}), 400
    jid = _uuid_mod.uuid4().hex[:8]
    with _picks_lock:
        _picks_jobs[jid] = {"log": [], "results": [], "done": False, "total": 0, "scanned": 0}
    threading.Thread(target=_run_picks_job,
                     args=(jid, mode, exchanges), daemon=True).start()
    return jsonify({"job_id": jid})


@app.route("/api/picks_progress/<jid>")
def picks_progress(jid):
    def gen():
        import json as _j, math as _m
        sent = 0
        while True:
            with _picks_lock:
                job = _picks_jobs.get(jid, {})
            if not job:
                yield f"data:{_j.dumps({'error':'not found'})}\n\n"; break
            while sent < len(job.get("log", [])):
                yield f"data:{_j.dumps(job['log'][sent])}\n\n"
                sent += 1
            if job.get("done"):
                yield f"data:{_j.dumps({'done': True, 'results': job.get('results', []), 'scanned': job.get('scanned', 0), 'total': job.get('total', 0)})}\n\n"
                break
            # Send scanned progress on every keepalive so UI counter updates live
            yield f"data:{_j.dumps({'scanned': job.get('scanned', 0), 'total': job.get('total', 0), 'qualified': len(job.get('results', []))})}\n\n"
            time.sleep(0.3)
    return Response(gen(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


def _run_picks_job(jid, mode, exchanges):
    import math as _m

    def log(msg, kind="info"):
        with _picks_lock:
            _picks_jobs[jid]["log"].append({"msg": msg, "kind": kind})

    try:
        global _picks_last_run_ts
        from screener_engine import fetch_tickers_for_exchanges, screen_stock
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import yfinance as yf

        # If a previous scan finished recently, wait for Yahoo Finance rate limits to clear
        elapsed = time.time() - _picks_last_run_ts
        if _picks_last_run_ts > 0 and elapsed < 90:
            wait = int(90 - elapsed)
            log(f"Waiting {wait}s for Yahoo Finance rate limits to clear after previous scan...", "phase")
            for i in range(wait, 0, -10):
                log(f"  Starting in {i}s...", "info")
                time.sleep(min(10, i))

        # ── Build universe ────────────────────────────────────────────────────
        log(f"Fetching tickers for: {', '.join(exchanges)}...", "phase")
        tickers = fetch_tickers_for_exchanges(exchanges, log_cb=log)
        tickers = list(dict.fromkeys(
            t for t in tickers
            if t and isinstance(t, str) and len(t) <= 5 and "." not in t
        ))
        total = len(tickers)
        log(f"Universe: {total} unique tickers", "info")
        with _picks_lock:
            _picks_jobs[jid]["total"] = total
        if total == 0:
            log("No tickers found — check exchange selection", "error")
            with _picks_lock: _picks_jobs[jid]["done"] = True
            return

        # ── Sector reference tables ───────────────────────────────────────────
        # Source: S&P 500 historical medians, updated 2024.
        # Update every 6-12 months from:
        #   PE  → https://www.wsj.com/market-data/stocks/peyield
        #   D/E → https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/dbtfund.html
        SECTOR_AVG_PE = {
            "Technology": 28.0, "Communication Services": 22.0,
            "Consumer Discretionary": 25.0, "Consumer Staples": 20.0,
            "Health Care": 22.0, "Financials": 14.0, "Industrials": 21.0,
            "Energy": 12.0, "Utilities": 17.0, "Real Estate": 35.0,
            "Materials": 18.0, "Basic Materials": 18.0,
        }
        SECTOR_AVG_DE = {
            # Debt/Equity as percentage — sector medians
            "Technology": 40.0, "Communication Services": 80.0,
            "Consumer Discretionary": 120.0, "Consumer Staples": 90.0,
            "Health Care": 50.0, "Financials": 300.0,
            "Industrials": 100.0, "Energy": 70.0,
            "Utilities": 150.0, "Real Estate": 120.0,
            "Materials": 60.0, "Basic Materials": 60.0,
        }
        DEFAULT_PE = 20.0
        DEFAULT_DE = 100.0

        passed  = []
        scanned = 0

        # ════════════════════════════════════════════════════════════════════
        # SHORT TERM PICKS — uses screener_engine.screen_stock (swing filters)
        # ════════════════════════════════════════════════════════════════════
        if mode == "short":
            conditions = {
                "market_cap": True, "div_yield": False, "pe_ratio": False,
                "debt_equity": False, "monthly_trend": True, "weekly_trend": False,
                "analyst_rating": False, "roce": False, "roe": False,
                "pe_vs_industry": False,
                "swing_rsi": True, "swing_ema": True,
                "swing_volume": True, "swing_macd": False, "swing_atr": True,
            }
            thresholds = {
                "min_market_cap_b": 1.0, "min_monthly_chg_pct": -5.0,
                "swing_min_rsi": 35, "swing_max_rsi": 70,
                "swing_min_vol_ratio": 1.2, "swing_min_atr_pct": 1.5,
            }
            label = "SHORT TERM PICKS (Swing)"
            log(f"Running {label} on {total} stocks...", "phase")
            log(f"Est. {round(_m.ceil(total/_WORKERS_PICKS)*_SECS_PER_STOCK/60,1)} min. Results stream as stocks qualify.", "info")

            def _short_one(sym):
                return sym, screen_stock(sym, conditions, thresholds, None)

            with ThreadPoolExecutor(max_workers=_WORKERS_PICKS) as ex:
                futures = {ex.submit(_short_one, sym): sym for sym in tickers}
                for future in as_completed(futures):
                    scanned += 1
                    sym, (result, rejection) = future.result()
                    if result:
                        score   = result.get("swing_score") or 0
                        price   = result.get("price_now")
                        sector  = result.get("sector", "—")
                        name    = result.get("name", sym)
                        verdict = ("STRONG" if score >= 7 else "GOOD" if score >= 5
                                   else "MIXED" if score >= 3 else "WEAK")
                        passed.append({"ticker": sym, "name": name, "sector": sector,
                            "price": round(float(price), 2) if price else None,
                            "score": round(float(score), 1) if score else 0,
                            "verdict": verdict})
                        with _picks_lock:
                            _picks_jobs[jid]["results"] = sorted(passed, key=lambda x: -x["score"])
                            _picks_jobs[jid]["scanned"] = scanned
                        log(f"PASS {sym} score={score:.1f} {sector}", "match")
                    if scanned % 50 == 0 or scanned == total:
                        log(f"Progress: {scanned}/{total} ({int(scanned/total*100)}%) — {len(passed)} qualifying", "info")

        # ════════════════════════════════════════════════════════════════════
        # LONG TERM PICKS — custom value screener with strict filters
        # Filters: PE <= sector avg | D/E <= sector avg | Div >= 2% (soft)
        #          ROE >= 8% AND rising 3 consecutive years | Analyst <= 3.0
        # ════════════════════════════════════════════════════════════════════
        else:
            label = "LONG TERM PICKS (Value)"
            log(f"Running {label} on {total} stocks...", "phase")
            log("Filters: PE<=sector avg | D/E<=sector avg | Div>=2% | ROE>=8% rising 3yr | Analyst<=3.0", "info")
            log(f"Est. {round(_m.ceil(total/_WORKERS_PICKS)*_SECS_PER_STOCK/60,1)} min. Results stream as stocks qualify.", "info")

            def _long_one(sym):
                try:
                    import random
                    time.sleep(random.uniform(0.05, 0.15))  # spread requests slightly
                    tk   = yf.Ticker(sym)
                    info = None
                    for attempt in range(3):
                        try:
                            raw = tk.info
                            if raw and len(raw) > 5:
                                info = raw; break
                        except Exception:
                            time.sleep(1.0 * (attempt + 1))
                    if not info: return sym, None

                    # Try fast_info first for price (cheaper API call)
                    price = None
                    try:
                        fi = tk.fast_info
                        price = fi.last_price if hasattr(fi, 'last_price') else None
                    except Exception:
                        pass
                    if not price:
                        price = info.get("regularMarketPrice") or info.get("currentPrice")
                    if not price: return sym, None

                    sector = info.get("sector", "Unknown")

                    # Filter 1: Market cap >= $1B
                    mcap = info.get("marketCap") or 0
                    if mcap < 1e9: return sym, None

                    # Filter 2: PE <= sector average
                    pe = info.get("trailingPE") or info.get("forwardPE")
                    if pe is not None and pe > 0:
                        if pe > SECTOR_AVG_PE.get(sector, DEFAULT_PE):
                            return sym, None

                    # Filter 3: D/E <= sector average
                    de = info.get("debtToEquity")
                    if de is not None:
                        if de > SECTOR_AVG_DE.get(sector, DEFAULT_DE):
                            return sym, None

                    # Filter 4: Analyst <= 3.0
                    analyst = info.get("recommendationMean")
                    if analyst is not None and analyst > 3.0:
                        return sym, None

                    # Filter 5: ROE >= 8% current
                    roe_now = info.get("returnOnEquity")
                    if roe_now is None or roe_now < 0.08:
                        return sym, None

                    # Filter 5b: ROE rising all 3 consecutive years
                    # If financial data is unavailable (rate-limited), skip trend check
                    # for consistency — current ROE >= 8% already verified above
                    roe_trend_ok = False
                    roe_data_available = False
                    try:
                        fin = tk.financials
                        bs  = tk.balance_sheet
                        if fin is not None and bs is not None and not fin.empty and not bs.empty:
                            ni_row = eq_row = None
                            for lbl in ["Net Income", "Net Income Common Stockholders",
                                        "NetIncome", "Net Income From Continuing Operations"]:
                                if lbl in fin.index: ni_row = fin.loc[lbl]; break
                            for lbl in ["Stockholders Equity", "Total Stockholder Equity",
                                        "Common Stock Equity", "Total Equity Gross Minority Interest"]:
                                if lbl in bs.index: eq_row = bs.loc[lbl]; break
                            if ni_row is not None and eq_row is not None:
                                cols = sorted(fin.columns, reverse=True)[:4]
                                roes = []
                                for c in cols:
                                    ni = ni_row.get(c); eq = eq_row.get(c)
                                    if ni and eq and float(eq) != 0:
                                        roes.append(float(ni) / float(eq))
                                if len(roes) >= 4:
                                    roe_data_available = True
                                    roe_trend_ok = (roes[0]>roes[1] and roes[1]>roes[2] and roes[2]>roes[3])
                                elif len(roes) >= 2:
                                    roe_data_available = True
                                    roe_trend_ok = roes[0] > roes[-1]
                    except Exception:
                        roe_data_available = False

                    # Only reject if data was available AND trend failed
                    # If data unavailable (rate-limit/missing), pass through for consistency
                    if roe_data_available and not roe_trend_ok:
                        return sym, None

                    # Dividend (soft filter — reduces score if < 2%, doesn't reject)
                    # Yahoo returns dividendYield as decimal (0.0055 = 0.55%)
                    # but sometimes as already-converted percent (0.55 = 0.55%)
                    # A real yield of 50%+ is impossible, so use 0.20 as the cutoff
                    raw_div = info.get("dividendYield") or 0
                    div_pct = round(raw_div * 100, 2) if raw_div < 0.20 else round(float(raw_div), 2)

                    # ── Score 0-10 ────────────────────────────────────────────
                    score = 0
                    sp = SECTOR_AVG_PE.get(sector, DEFAULT_PE)
                    sd = SECTOR_AVG_DE.get(sector, DEFAULT_DE)

                    # Dividend (0-2 pts)
                    if div_pct >= 3.0:   score += 2
                    elif div_pct >= 2.0: score += 1
                    # PE vs sector avg (0-2 pts): lower = better
                    if pe and pe > 0:
                        ratio = pe / sp
                        if ratio <= 0.7:   score += 2
                        elif ratio <= 0.9: score += 1
                    # D/E vs sector (0-1 pt)
                    if de is not None and de <= sd * 0.6: score += 1
                    # ROE level (0-2 pts)
                    if roe_now >= 0.20:   score += 2
                    elif roe_now >= 0.12: score += 1
                    # ROE trend (1 pt)
                    if roe_trend_ok: score += 1
                    # Analyst (0-2 pts)
                    if analyst is not None:
                        if analyst <= 1.5:   score += 2
                        elif analyst <= 2.5: score += 1
                    # Large cap stability (0-1 pt)
                    if mcap >= 50e9: score += 1
                    score = min(score, 10)

                    verdict = ("STRONG" if score >= 7 else "GOOD" if score >= 5
                               else "MIXED" if score >= 3 else "WEAK")
                    flags = []
                    if div_pct < 2.0:           flags.append("Low/no dividend")
                    if pe is None:              flags.append("PE unavailable")
                    if not roe_data_available:  flags.append("ROE trend data unavailable")
                    elif not roe_trend_ok:      flags.append("ROE trend not confirmed")

                    return sym, {
                        "ticker": sym, "name": info.get("shortName") or sym,
                        "sector": sector, "price": round(float(price), 2),
                        "score": round(float(score), 1), "verdict": verdict,
                        "pe": round(float(pe), 1) if pe else None,
                        "sector_pe": sp,
                        "de": round(float(de), 1) if de is not None else None,
                        "sector_de": sd,
                        "roe": round(float(roe_now) * 100, 1),
                        "div": round(float(div_pct), 2),
                        "analyst": round(float(analyst), 2) if analyst else None,
                        "mcap_b": round(mcap / 1e9, 1),
                        "flags": flags,
                    }
                except Exception:
                    return sym, None

            with ThreadPoolExecutor(max_workers=_WORKERS_PICKS) as ex:
                futures = {ex.submit(_long_one, sym): sym for sym in tickers}
                for future in as_completed(futures):
                    scanned += 1
                    sym, result = future.result()
                    if result:
                        passed.append(result)
                        with _picks_lock:
                            _picks_jobs[jid]["results"] = sorted(passed, key=lambda x: -x["score"])
                            _picks_jobs[jid]["scanned"] = scanned
                        log(f"PASS {sym} score={result['score']} ROE={result['roe']}% div={result['div']}%", "match")
                    if scanned % 50 == 0 or scanned == total:
                        log(f"Progress: {scanned}/{total} ({int(scanned/total*100)}%) — {len(passed)} qualifying", "info")

        passed_sorted = sorted(passed, key=lambda x: -x["score"])
        log(f"COMPLETE — {len(passed_sorted)} qualified from {total} scanned", "ok")
        _picks_last_run_ts = time.time()  # record completion for rate-limit cooldown
        with _picks_lock:
            _picks_jobs[jid]["results"] = passed_sorted
            _picks_jobs[jid]["done"]    = True

    except Exception as e:
        import traceback
        tb = traceback.format_exc().splitlines()[-1]
        log(f"Error: {e} — {tb}", "error")
        with _picks_lock:
            _picks_jobs[jid]["done"] = True



# ── ETF Lookup — top N funds holding each Final 5 stock ──────────────────────
@app.route("/api/etf_lookup", methods=["POST"])
def etf_lookup():
    """
    POST { "tickers": ["AAMI","ASX","ABX","APH","BDJ"], "n": 3 }
    Returns { "etfs": { "AAMI": [{"name":"IJR","full_name":"...","pct":0.089},...], ... } }
    """
    import concurrent.futures as _cf
    data    = request.get_json() or {}
    tickers = data.get("tickers", [])
    n       = int(data.get("n", 3))
    etf_map = {}
    with _cf.ThreadPoolExecutor(max_workers=5) as ex:
        futures = {ex.submit(get_top_etfs_for_ticker, t, n): t for t in tickers}
        for future in _cf.as_completed(futures):
            ticker          = futures[future]
            try:
                etf_map[ticker] = future.result()
            except Exception:
                etf_map[ticker] = []
    return Response(_dumps({"etfs": etf_map}), mimetype="application/json")


if __name__ == "__main__":
    print("\n  ╔═══════════════════════════════════════════╗")
    print("  ║   SmartStock  —  http://localhost:5000    ║")
    print("  ╚═══════════════════════════════════════════╝\n")
    app.run(debug=False, port=5000, threaded=True)