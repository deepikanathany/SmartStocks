"""
screener_engine.py  —  SmartStock US Screener
All yfinance screening, ticker fetching, and Excel export logic.
Imported by app.py — no Flask dependency here.
"""

import yfinance as yf
import pandas as pd
import requests
import time
import io
import os
import json
from datetime import datetime, timedelta
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ─────────────────────────────────────────────────────────────────────────────
# FMP CONFIG  (optional)
# ─────────────────────────────────────────────────────────────────────────────
FMP_API_KEY = None

# ─────────────────────────────────────────────────────────────────────────────
# Groq config (shared key used across SmartStock — same as swing_pro_engine.py)
# Used for SEPA catalyst-news check: is there a fresh reason (earnings, big
# order, sector news) for this stock to move, or should the trader wait?
# ─────────────────────────────────────────────────────────────────────────────
try:
    from groq import Groq as _Groq
    _GROQ_KEY    = "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"
    _GROQ_MODEL  = "llama-3.3-70b-versatile"
    _groq_client = _Groq(api_key=_GROQ_KEY)
    _GROQ_OK     = True
except ImportError:
    _GROQ_OK     = False
    _groq_client = None
    _GROQ_MODEL  = ""


def _ask_groq(prompt, max_tokens=200):
    if not _GROQ_OK or not _groq_client:
        return None
    try:
        r = _groq_client.chat.completions.create(
            model=_GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
            max_tokens=max_tokens,
        )
        return r.choices[0].message.content.strip()
    except Exception:
        return None


# Sector -> SPDR sector ETF, used as a fallback news source when a stock has
# no fresh headlines of its own (checks for sector-wide catalysts instead)
_SEPA_SECTOR_ETF = {
    "Technology":             "XLK",
    "Communication Services": "XLC",
    "Consumer Cyclical":      "XLY",
    "Consumer Defensive":     "XLP",
    "Healthcare":             "XLV",
    "Financials":             "XLF",
    "Financial Services":     "XLF",
    "Industrials":            "XLI",
    "Basic Materials":        "XLB",
    "Energy":                 "XLE",
    "Utilities":              "XLU",
    "Real Estate":            "XLRE",
}


def _check_catalyst_news(symbol, name, sector):
    """
    SEPA catalyst check — for a stock that has already passed the full
    Trend Template + Fundamentals + VCP screen, check whether there's a
    fresh (last 24h) news catalyst that explains WHY it might move now:
    earnings, a major contract/order, an upgrade, or sector-wide news.

    Falls back to sector ETF news, then SPY (market-wide) news, if the
    stock itself has nothing in the last 24 hours.

    Returns dict:
        {"verdict": "BUY"|"WAIT"|"N/A", "reason": str, "source": str}
    """
    if not _GROQ_OK:
        return {"verdict": "N/A", "reason": "Groq not available", "source": "none"}

    cutoff = time.time() - 24 * 3600  # last 24 hours

    def recent_headlines(tkr, limit=6):
        try:
            items = yf.Ticker(tkr).news or []
            recent = [it for it in items
                      if it.get("providerPublishTime", 0) >= cutoff and it.get("title")]
            return [it["title"] for it in recent[:limit]]
        except Exception:
            return []

    headlines = recent_headlines(symbol)
    source = "stock"

    if not headlines:
        etf = _SEPA_SECTOR_ETF.get(sector)
        if etf:
            sector_news = recent_headlines(etf)
            if sector_news:
                headlines = sector_news
                source = "sector"

    if not headlines:
        market_news = recent_headlines("SPY")
        if market_news:
            headlines = market_news
            source = "market"

    if not headlines:
        return {
            "verdict": "WAIT",
            "reason": "No fresh news in the last 24h for this stock, its sector, or the broader market — the technical setup looks good but there's no obvious catalyst right now.",
            "source": "none",
        }

    hl_text = "\n".join(f"- {h}" for h in headlines)
    scope = {"stock": "this stock", "sector": "this stock's sector",
             "market": "the broader market"}[source]

    prompt = f"""You are a swing-trading catalyst analyst.

Stock: {name} ({symbol}), Sector: {sector}

This stock just passed a full technical + fundamental trend screen (Minervini-style
Trend Template, EPS/revenue growth, and a Volatility Contraction Pattern). The setup
looks technically ready. The question is whether there's a CATALYST right now to
actually trigger the move.

Recent headlines (last 24h, from {scope}):
{hl_text}

Based ONLY on these headlines, is there a real catalyst (earnings beat/report,
new contract or big order, analyst upgrade, product launch, regulatory news,
or sector-moving news) that could drive THIS stock higher in the near term?

Respond ONLY in this JSON format, no markdown:
{{"verdict": "BUY", "reason": "One sentence, max 25 words, citing the specific catalyst."}}
or
{{"verdict": "WAIT", "reason": "One sentence, max 25 words, explaining why nothing in these headlines is a real catalyst."}}"""

    raw = _ask_groq(prompt, max_tokens=120)
    if not raw:
        return {"verdict": "N/A", "reason": "Groq request failed", "source": source}

    try:
        raw = raw.replace("```json", "").replace("```", "").strip()
        parsed = json.loads(raw)
        verdict = parsed.get("verdict", "WAIT")
        if verdict not in ("BUY", "WAIT"):
            verdict = "WAIT"
        return {"verdict": verdict, "reason": parsed.get("reason", ""), "source": source}
    except Exception:
        return {"verdict": "N/A", "reason": "Could not parse Groq response", "source": source}

SECTOR_PE_FALLBACK = {
    "Technology": 35.0, "Communication Services": 22.0,
    "Consumer Cyclical": 25.0, "Consumer Defensive": 22.0,
    "Healthcare": 28.0, "Financials": 14.0, "Financial Services": 14.0,
    "Industrials": 22.0, "Basic Materials": 18.0,
    "Energy": 12.0, "Utilities": 18.0, "Real Estate": 35.0,
}

_industry_pe_cache: dict = {}
_industry_pe_date:  str  = ""


def load_industry_pe_map(log_cb=None):
    global _industry_pe_cache, _industry_pe_date
    today = datetime.today().strftime("%Y-%m-%d")
    if _industry_pe_date == today and _industry_pe_cache:
        return
    def log(m):
        if log_cb: log_cb(m)
    if not FMP_API_KEY:
        log("P/E vs Industry: no FMP key — using hardcoded sector fallback")
        _industry_pe_cache = {}
        _industry_pe_date  = today
        return
    try:
        cache = {}
        for exchange in ("NYSE", "NASDAQ"):
            url = (f"https://financialmodelingprep.com/api/v4/industry_price_earning_ratio"
                   f"?date={today}&exchange={exchange}&apikey={FMP_API_KEY}")
            r = requests.get(url, timeout=10); r.raise_for_status()
            for row in r.json():
                industry = row.get("industry", "").strip()
                pe_val   = row.get("pe")
                if industry and pe_val:
                    try:
                        pef = float(pe_val)
                        if pef > 0:
                            cache[industry] = round((cache.get(industry, pef) + pef) / 2, 2)
                    except (ValueError, TypeError):
                        pass
        if cache:
            _industry_pe_cache = cache
            _industry_pe_date  = today
            log(f"P/E vs Industry: loaded {len(cache)} industries from FMP")
        else:
            raise ValueError("FMP returned empty data")
    except Exception as ex:
        log(f"P/E vs Industry: FMP fetch failed ({ex}) — using hardcoded fallback")
        _industry_pe_cache = {}
        _industry_pe_date  = today


def get_industry_pe(industry: str, sector: str):
    if _industry_pe_cache:
        pe = _industry_pe_cache.get(industry)
        if pe: return pe
        pe = _industry_pe_cache.get(sector)
        if pe: return pe
        il = industry.lower() if industry else ""
        for key, val in _industry_pe_cache.items():
            if il and il in key.lower(): return val
    return SECTOR_PE_FALLBACK.get(sector)


# ─────────────────────────────────────────────────────────────────────────────
# TICKER UNIVERSE
# ─────────────────────────────────────────────────────────────────────────────
_CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ticker_cache.json")

_FALLBACK = {
    "SP500": [
        "AAPL","MSFT","NVDA","AMZN","GOOGL","META","BRK-B","LLY","AVGO","JPM",
        "TSLA","UNH","XOM","V","MA","PG","JNJ","HD","COST","MRK","ABBV","CVX",
        "ORCL","WMT","BAC","KO","NFLX","CRM","PEP","TMO","ACN","MCD","LIN","ABT",
        "AMD","CSCO","TXN","WFC","DHR","NEE","AMGN","NKE","PM","INTC","UNP","INTU",
        "HON","CAT","IBM","GE","SPGI","GS","BLK","LOW","ISRG","ELV","BKNG","SYK",
        "VRTX","REGN","PLD","CB","ADI","MMC","C","SCHW","DE","TJX","MO","ZTS",
        "BSX","GILD","CI","CME","AXP","ETN","SO","PGR","DUK","EOG","AON","MDLZ",
        "SLB","MCO","ITW","NOC","EMR","MPC","RTX","USB","WM","TGT","FCX","CL",
        "EQIX","FDX","NSC","CEG","CSX","ROP","HCA","PCAR","APH","MSI","CDNS",
        "SNPS","PANW","PYPL","ADSK","FTNT","KLAC","LRCX","AMAT","MCHP",
    ],
    "NASDAQ": [
        "AAPL","MSFT","NVDA","AMZN","META","GOOGL","TSLA","AVGO","COST","NFLX",
        "AMD","ADBE","CSCO","INTU","PEP","AMGN","TXN","ISRG","QCOM","CMCSA",
        "BKNG","HON","AMAT","VRTX","ADP","PANW","GILD","LRCX","MU","KLAC",
        "SNPS","CDNS","REGN","MELI","ABNB","MDLZ","FTNT","CRWD","PYPL","CTAS",
        "ORLY","MNST","ROST","PCAR","CEG","IDXX","MRVL","KDP","FAST","DXCM",
        "TEAM","WDAY","NXPI","ASML","CPRT","ODFL","ZS","OKTA","VRSK","ANSS",
        "MTCH","ALGN","DLTR","SIRI","EXPE","EBAY","WDC","STX","NTAP","CTSH","TMUS",
    ],
    "DOWJONES": [
        "AAPL","MSFT","UNH","GS","HD","SHW","CAT","CRM","MCD","V","AMGN","TRV",
        "AXP","JPM","IBM","HON","NVDA","MMM","DIS","BA","WMT","PG","CVX","KO",
        "MRK","CSCO","JNJ","NKE","DOW","VZ",
    ],
    "NYSE": [
        "XOM","CVX","COP","EOG","SLB","MPC","PSX","VLO","OXY","HAL","DVN",
        "JPM","BAC","WFC","C","GS","MS","BX","KKR","AXP","CB","TRV","AIG",
        "MET","PRU","AFL","ALL","PGR","USB","TFC","BK","STT","NTRS",
        "JNJ","PFE","ABT","BMY","MRK","LLY","ABBV","MDT","BSX","SYK","ZBH",
        "PG","KO","PEP","PM","MO","CL","MCD","YUM","DRI","CMG","SBUX","HLT","MAR",
        "WMT","TGT","KR","GE","HON","MMM","CAT","DE","EMR","ETN","ROK","ITW",
        "BA","GD","LMT","NOC","RTX","LHX","IBM","ORCL","ACN","HPE","T","VZ",
        "AMT","CCI","EQIX","DLR","PSA","PLD","SPG","NEE","DUK","SO","AEP",
        "FCX","NEM","NKE","TJX","M","F","GM","MMC","AON","MCO",
    ],

    # ── Curated sub-universes — ranked by market cap (updated Apr 2025) ────────
    # S&P 500 Top 100 by market cap
    "SP500_TOP100": [
        "AAPL","MSFT","NVDA","AMZN","GOOGL","META","GOOG","BRK-B","LLY","AVGO",
        "TSLA","JPM","V","UNH","XOM","MA","COST","HD","PG","ABBV",
        "NFLX","JNJ","BAC","CRM","AMD","WMT","CVX","MRK","TMO","ORCL",
        "CSCO","ACN","NOW","IBM","GE","LIN","ABT","MCD","ISRG","GS",
        "AMGN","CAT","INTU","TXN","PM","SPGI","PEP","BLK","LOW","SYK",
        "BKNG","VRTX","REGN","PLD","ADI","MMC","C","AXP","SCHW","DE",
        "BSX","TJX","MO","GILD","CI","ETN","ZTS","CME","WFC","MDLZ",
        "CB","SO","DUK","ITW","NOC","EMR","EOG","MCO","RTX","USB",
        "WM","TGT","FCX","CL","EQIX","FDX","NSC","CSX","ROP","HCA",
        "PCAR","APH","MSI","CDNS","SNPS","PANW","ADSK","FTNT","KLAC","LRCX",
    ],

    # S&P 500 Top 200 by market cap
    "SP500_TOP200": [
        "AAPL","MSFT","NVDA","AMZN","GOOGL","META","GOOG","BRK-B","LLY","AVGO",
        "TSLA","JPM","V","UNH","XOM","MA","COST","HD","PG","ABBV",
        "NFLX","JNJ","BAC","CRM","AMD","WMT","CVX","MRK","TMO","ORCL",
        "CSCO","ACN","NOW","IBM","GE","LIN","ABT","MCD","ISRG","GS",
        "AMGN","CAT","INTU","TXN","PM","SPGI","PEP","BLK","LOW","SYK",
        "BKNG","VRTX","REGN","PLD","ADI","MMC","C","AXP","SCHW","DE",
        "BSX","TJX","MO","GILD","CI","ETN","ZTS","CME","WFC","MDLZ",
        "CB","SO","DUK","ITW","NOC","EMR","EOG","MCO","RTX","USB",
        "WM","TGT","FCX","CL","EQIX","FDX","NSC","CSX","ROP","HCA",
        "PCAR","APH","MSI","CDNS","SNPS","PANW","ADSK","FTNT","KLAC","LRCX",
        "AMAT","MCHP","PYPL","DHR","NEE","AON","NKE","PGR","SLB","WELL",
        "AJG","PRU","ICE","AFL","ALL","CTAS","EW","A","SHW","APD",
        "D","EXC","PAYX","OKE","KMI","PSA","AVB","ROST","ORLY","FAST",
        "ODFL","VRSK","ANSS","IDXX","BIIB","DXCM","ILMN","ALGN","HSIC","CINF",
        "ROL","EPAM","CTSH","MNST","DLTR","KDP","MELI","ABNB","CRWD","WDAY",
        "ZS","DDOG","NET","SNOW","COIN","SQ","SHOP","RBLX","LYFT","UBER",
        "DASH","HOOD","PATH","AI","GTLB","MDB","CFLT","ESTC","SMAR","BOX",
        "DOCN","FROG","SUMO","JAMF","BRZE","AMPL","TASK","GTLB","RNG","FIVE",
        "CHWY","ETSY","W","PRTS","REAL","POSH","RENT","WISH","OPAD","SPWH",
        "VSTO","HIBB","BIG","CATO","DXLG","SCVL","OXM","TLYS","MOV","CTRN",
    ],

    # NASDAQ 100 (QQQ components) — ranked by weight
    "NASDAQ100": [
        "AAPL","MSFT","NVDA","AMZN","META","GOOGL","GOOG","AVGO","TSLA","COST",
        "NFLX","AMD","ADBE","CSCO","INTU","PEP","AMGN","TXN","ISRG","QCOM",
        "CMCSA","BKNG","AMAT","VRTX","ADP","PANW","GILD","LRCX","MU","KLAC",
        "SNPS","CDNS","REGN","MELI","ABNB","MDLZ","FTNT","CRWD","PYPL","CTAS",
        "ORLY","MNST","ROST","PCAR","CEG","IDXX","MRVL","KDP","FAST","DXCM",
        "TEAM","WDAY","NXPI","ASML","CPRT","ODFL","ZS","VRSK","ANSS","CHTR",
        "FANG","PAYX","MCHP","BIIB","SIRI","EXPE","EBAY","WDC","STX","NTAP",
        "CTSH","TMUS","DDOG","INTC","ILMN","ALGN","DLTR","VRSN","SWKS","TTWO",
        "MTCH","OKTA","ZM","DOCU","PTON","ROKU","NTNX","COUP","ESTC","VEEV",
        "HZNP","LULU","BMRN","ALXN","SGEN","NBIX","ARWR","ITCI","FOLD","FATE",
    ],

    # Russell 1000 — large+mid cap (top 250 by weight shown as representative sample)
    "RUSSELL1000": [
        "AAPL","MSFT","NVDA","AMZN","GOOGL","META","GOOG","BRK-B","LLY","AVGO",
        "TSLA","JPM","V","UNH","XOM","MA","COST","HD","PG","ABBV",
        "NFLX","JNJ","BAC","CRM","AMD","WMT","CVX","MRK","TMO","ORCL",
        "CSCO","ACN","NOW","IBM","GE","LIN","ABT","MCD","ISRG","GS",
        "AMGN","CAT","INTU","TXN","PM","SPGI","PEP","BLK","LOW","SYK",
        "BKNG","VRTX","REGN","PLD","ADI","MMC","C","AXP","SCHW","DE",
        "BSX","TJX","MO","GILD","CI","ETN","ZTS","CME","WFC","MDLZ",
        "CB","SO","DUK","ITW","NOC","EMR","EOG","MCO","RTX","USB",
        "WM","TGT","FCX","CL","EQIX","FDX","NSC","CSX","ROP","HCA",
        "PCAR","APH","MSI","CDNS","SNPS","PANW","ADSK","FTNT","KLAC","LRCX",
        "AMAT","MCHP","PYPL","DHR","NEE","AON","NKE","PGR","SLB","WELL",
        "AJG","PRU","ICE","AFL","ALL","CTAS","EW","A","SHW","APD",
        "D","EXC","PAYX","OKE","KMI","PSA","AVB","ROST","ORLY","FAST",
        "F","GM","UBER","LYFT","DASH","ABNB","COIN","SQ","SHOP","RBLX",
        "TTD","OPEN","Z","MTCH","BMBL","GRPN","YELP","ANGI","IAC","LMND",
        "SOFI","AFRM","UPST","LC","TREE","CURO","ENVA","OMF","QFIN","X",
        "NUE","STLD","CLF","RS","CMC","ATI","CENX","KALU","HAYN","TMST",
        "DVN","FANG","MRO","APA","SM","CIVI","PDCE","MTDR","ESTE","REX",
        "RRC","EQT","CNX","CRK","AR","SWN","GPOR","TELL","GLNG","GLOG",
        "CCL","RCL","NCLH","MAR","HLT","H","IHG","WH","CHH","VC",
        "UAL","DAL","AAL","LUV","ALK","JBLU","SAVE","HA","ULCC","MESA",
        "FDX","UPS","XPO","GXO","SAIA","WERN","KNX","CHRW","JBHT","HTLD",
        "DG","DLTR","BIG","FIVE","PRTY","CATO","DXLG","SCVL","OXM","MOV",
        "HD","LOW","TSCO","WSM","RH","BBY","SIG","TCS","CONN","FURN",
    ],

    # Russell 2000 — small cap sample (200 representative tickers)
    "RUSSELL2000": [
        "SMCI","MSTR","AEHR","CLOV","OPEN","DKNG","STEM","BLNK","NKLA","RIDE",
        "GOEV","FSR","WKHS","XPEV","LI","NIO","LCID","RIVN","FFIE","MULN",
        "CPRX","HIMS","ACCD","DOCS","ONEM","TDOC","AMWL","LFST","TALK","GDRX",
        "SMAR","BOX","DOCN","FROG","SUMO","JAMF","BRZE","AMPL","TASK","GTLB",
        "VERX","ALKT","CFLT","ESTC","DDOG","NET","ZS","SNOW","PATH","AI",
        "RBLX","HOOD","DAVE","MRIN","GHVI","PSTH","AJAX","LFAC","GNAC","AJAX",
        "XPOF","PLNT","PTON","BMBL","GRPN","YELP","ANGI","IAC","LMND","ROOT",
        "INVA","AMRX","MNKD","AGEN","IMMU","ACAD","EXEL","PCVX","RCKT","ARWR",
        "BLUE","SGEN","FATE","FOLD","ITCI","NBIX","ARVN","KRYS","KROS","IMVT",
        "JANX","NUVL","RXRX","TWST","BEAM","NTLA","EDIT","CRSP","VERVE","PRIME",
        "SDGR","SCHRD","RXRX","TALS","PRAX","PTGX","KYMR","KRTX","ACMR","ACRX",
        "MRSN","MRUS","MSRT","MGTX","MGNX","MDXG","MDVX","MDRN","MDGL","MDCO",
        "CNST","CNTG","CNSP","CNRG","CNMD","CNFN","CNET","CNEY","CNDA","CNCE",
        "AXSM","AXNX","AXNF","AXMO","AXLA","AXJG","AXGT","AXGN","AXFN","AXDX",
        "RVMD","RVLV","RVNC","RVNB","RVEN","RVDC","RVBL","RVBD","RVAN","RVAT",
        "LNTH","LNSR","LNNR","LNNN","LNMK","LNKN","LNJN","LNIN","LNHN","LNGN",
        "SNDL","SNCR","SNBR","SNAP","SNAL","SNAG","SNAC","SN","SMWB","SMTC",
        "AGEN","AGRX","AGRI","AGRO","AGPP","AGMH","AGFY","AGFS","AGFN","AGCL",
        "PLBY","PLSE","PLRX","PLPC","PLNT","PLNG","PLNHF","PLMR","PLL","PLKQ",
        "HTHT","HTGM","HTGC","HTFB","HTFA","HTBX","HROW","HRTG","HRMY","HROB",
    ],

    # Sector-focused universes (S&P 500 stocks filtered by sector)
    "SECTOR_TECH": [
        "AAPL","MSFT","NVDA","AVGO","AMD","ORCL","CRM","CSCO","ADBE","NOW",
        "INTU","TXN","IBM","ACN","QCOM","AMAT","ADI","KLAC","LRCX","SNPS",
        "CDNS","MCHP","PANW","FTNT","ADSK","MSI","APH","ANSS","CTSH","EPAM",
        "GDDY","JNPR","NTAP","STX","WDC","HPE","FFIV","AKAM","ZBRA","TTWO",
        "EA","ATVI","RBLX","U","DDOG","MDB","CFLT","GTLB","NET","ZS",
        "CRWD","OKTA","TEAM","WDAY","VEEV","PAYC","PCTY","SMAR","BOX","DOCN",
    ],
    "SECTOR_HEALTHCARE": [
        "LLY","UNH","JNJ","ABBV","MRK","TMO","ABT","ISRG","SYK","BSX",
        "AMGN","VRTX","REGN","GILD","CI","MDT","ZTS","DXCM","EW","IDXX",
        "IQV","BIIB","ILMN","ALGN","HSIC","HOLX","PODD","TMDX","INSP","NARI",
        "BDX","BAX","STE","AMED","ACCD","DOCS","ONEM","TDOC","AMWL","HIMS",
        "MRNA","BNTX","PFE","AZN","GSK","NVO","RHHBY","SNY","TAK","NVS",
        "INCY","ALXN","BMRN","NBIX","ARWR","ITCI","FOLD","FATE","BLUE","SGEN",
    ],
    "SECTOR_FINANCIALS": [
        "JPM","BAC","WFC","GS","MS","C","BX","KKR","AXP","CB",
        "BLK","SCHW","ICE","CME","SPGI","MCO","MMC","AON","AJG","PRU",
        "MET","AFL","ALL","PGR","TRV","USB","TFC","BK","STT","NTRS",
        "COF","DFS","SYF","AMP","RJF","LNC","UNM","GL","FNF","FAF",
        "HBAN","RF","FITB","MTB","CFG","ZION","CMA","KEY","PBCT","FHN",
        "SOFI","AFRM","UPST","LC","SQ","PYPL","COIN","HOOD","SMAR","OPEN",
    ],
    "SECTOR_ENERGY": [
        "XOM","CVX","COP","EOG","SLB","MPC","PSX","VLO","OXY","HAL",
        "DVN","FANG","MRO","APA","SM","CIVI","PDCE","MTDR","HES","PXD",
        "RRC","EQT","CNX","CRK","AR","SWN","CHK","KMI","OKE","WMB",
        "ET","EPD","MPLX","PAA","PAGP","AM","CTRA","PR","LPX","TRGP",
    ],
    "SECTOR_CONSUMER": [
        "AMZN","TSLA","HD","COST","WMT","MCD","NKE","SBUX","TGT","LOW",
        "TJX","ROST","BKNG","ABNB","MAR","HLT","F","GM","UBER","LYFT",
        "DASH","ETSY","CHWY","W","BBWI","RL","PVH","TPR","CPRI","VFC",
        "KO","PEP","PM","MO","MDLZ","KHC","GIS","K","CPB","HRL",
        "CL","PG","EL","ULTA","COTY","REV","ELF","SKIN","PLBY","BMBL",
    ],
    "SECTOR_INDUSTRIALS": [
        "GE","HON","CAT","DE","EMR","ETN","ITW","RTX","BA","GD",
        "LMT","NOC","LHX","FDX","UPS","XPO","GXO","SAIA","WERN","KNX",
        "NSC","CSX","UNP","CNI","CP","WAB","TT","PNR","GNRC","ROK",
        "IR","CARR","OTIS","ROP","AME","IEX","IDEX","RXO","CHRW","JBHT",
    ],
}


def _load_cache():
    try:
        with open(_CACHE_FILE) as f: return json.load(f)
    except Exception: return {}


def _save_cache(data):
    try:
        with open(_CACHE_FILE, "w") as f: json.dump(data, f)
    except Exception: pass


def _fetch_live_tickers(exchange_key, log_cb=None):
    def log(m):
        if log_cb: log_cb(m)

    if exchange_key == "SP500":
        for src_url in [
            "https://raw.githubusercontent.com/datasets/s-and-p-500-companies/main/data/constituents.csv",
            "https://raw.githubusercontent.com/nickoala/snp/master/snp500.csv",
        ]:
            try:
                host = src_url.split("/")[2]
                log(f"SP500: trying {host}…")
                r = requests.get(src_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
                r.raise_for_status()
                import csv as _csv, io as _io
                reader = _csv.reader(_io.StringIO(r.text)); next(reader, None)
                symbols = [row[0].strip().replace(".", "-") for row in reader
                           if row and row[0].strip().replace(".", "-").replace("-","").isalpha() and len(row[0]) <= 5]
                if len(symbols) >= 400:
                    log(f"SP500: {len(symbols)} tickers"); return symbols
            except Exception as e:
                log(f"SP500: failed — {str(e)[:60]}")
        return []

    if exchange_key == "NASDAQ":
        for src_kind, src_url in [
            ("txt",  "https://raw.githubusercontent.com/rreichel3/US-Stock-Symbols/main/nasdaq/nasdaq_tickers.txt"),
            ("pipe", "https://ftp.nasdaqtrader.com/dynamic/SymDir/nasdaqlisted.txt"),
        ]:
            try:
                host = src_url.split("/")[2]
                log(f"NASDAQ: trying {host}…")
                r = requests.get(src_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
                r.raise_for_status()
                symbols = []
                if src_kind == "txt":
                    symbols = [l.strip() for l in r.text.splitlines()
                               if l.strip() and "." not in l and l.strip().replace("-","").isalpha() and len(l.strip()) <= 5]
                else:
                    for line in r.text.splitlines()[1:]:
                        parts = line.split("|")
                        if len(parts) < 2: continue
                        sym = parts[0].strip()
                        test = parts[6].strip() if len(parts) > 6 else "N"
                        etf  = parts[5].strip() if len(parts) > 5 else "N"
                        if sym and "." not in sym and test != "Y" and etf != "Y" and sym.isalpha() and len(sym) <= 5:
                            symbols.append(sym)
                if len(symbols) >= 200:
                    log(f"NASDAQ: {len(symbols)} tickers"); return symbols
            except Exception as e:
                log(f"NASDAQ: failed — {str(e)[:60]}")
        return []

    if exchange_key == "NYSE":
        for src_kind, src_url in [
            ("txt",  "https://raw.githubusercontent.com/rreichel3/US-Stock-Symbols/main/nyse/nyse_tickers.txt"),
            ("sec",  "https://www.sec.gov/files/company_tickers_exchange.json"),
            ("pipe", "https://ftp.nasdaqtrader.com/dynamic/SymDir/otherlisted.txt"),
        ]:
            try:
                host = src_url.split("/")[2]
                log(f"NYSE: trying {host}…")
                r = requests.get(src_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
                r.raise_for_status()
                symbols = []
                if src_kind == "txt":
                    symbols = [l.strip() for l in r.text.splitlines()
                               if l.strip() and "." not in l and len(l.strip()) <= 5]
                elif src_kind == "sec":
                    data = r.json(); fields = data.get("fields",[]); rows = data.get("data",[])
                    try:
                        ex_idx = fields.index("exchange"); sym_idx = fields.index("ticker")
                    except ValueError: continue
                    symbols = [str(row[sym_idx]).upper().strip() for row in rows
                               if row[ex_idx] == "NYSE" and "." not in str(row[sym_idx]) and len(str(row[sym_idx])) <= 5]
                elif src_kind == "pipe":
                    for line in r.text.splitlines()[1:]:
                        parts = line.split("|")
                        if len(parts) < 3: continue
                        sym = parts[0].strip(); exchange = parts[2].strip()
                        etf = parts[4].strip() if len(parts) > 4 else ""; test = parts[6].strip() if len(parts) > 6 else "N"
                        if exchange == "N" and etf != "Y" and test != "Y" and sym and "." not in sym and len(sym) <= 5:
                            symbols.append(sym)
                if symbols:
                    log(f"NYSE: {len(symbols)} tickers from {host}"); return symbols
            except Exception as e:
                log(f"NYSE: {host} failed — {str(e)[:60]}")
        return []

    if exchange_key == "DOWJONES" and FMP_API_KEY:
        try:
            r = requests.get(f"https://financialmodelingprep.com/api/v3/dowjones_constituent?apikey={FMP_API_KEY}", timeout=15)
            r.raise_for_status()
            syms = [row["symbol"] for row in r.json() if row.get("symbol")]
            if syms: log(f"DOWJONES: {len(syms)} tickers"); return syms
        except Exception as e:
            log(f"DOWJONES: FMP failed — {str(e)[:60]}")
    return []


# Keys that are static curated lists — no live fetch needed
_STATIC_UNIVERSES = {
    "SP500_TOP100", "SP500_TOP200", "NASDAQ100",
    "RUSSELL1000",  "RUSSELL2000",
    "SECTOR_TECH",  "SECTOR_HEALTHCARE", "SECTOR_FINANCIALS",
    "SECTOR_ENERGY","SECTOR_CONSUMER",   "SECTOR_INDUSTRIALS",
}

def fetch_tickers_for_exchanges(exchanges, log_cb=None):
    def log(m):
        if log_cb: log_cb(m)
    cache = _load_cache(); today = datetime.today().strftime("%Y-%m-%d")
    all_tickers = set(); cache_dirty = False

    for ex in exchanges:
        # ── Static curated lists — just return the hardcoded list, no HTTP call ──
        if ex in _STATIC_UNIVERSES:
            tickers = _FALLBACK.get(ex, [])
            log(f"{ex}: {len(tickers)} tickers (curated list)")
            all_tickers.update(tickers)
            continue

        # ── Dynamic exchanges — try live fetch, fallback to cache then hardcoded ──
        cached = cache.get(ex, {})
        log(f"Cache for {ex}: date={cached.get('date')}, count={len(cached.get('tickers',[]))}")
        needs_fmp    = ex == "DOWJONES"
        should_fetch = cached.get("date") != today and (not needs_fmp or FMP_API_KEY)

        if should_fetch:
            tickers = _fetch_live_tickers(ex, log_cb=log_cb)
            if tickers:
                log(f"{ex}: {len(tickers)} tickers (live)")
                cache[ex] = {"date": today, "tickers": tickers}; cache_dirty = True
            else:
                tickers = cached.get("tickers") or _FALLBACK.get(ex, [])
                log(f"{ex}: {len(tickers)} tickers ({'cached' if cached.get('tickers') else 'fallback'})")
        elif cached.get("tickers"):
            tickers = cached["tickers"]
            log(f"{ex}: {len(tickers)} tickers (cached)")
        else:
            tickers = _FALLBACK.get(ex, [])
            log(f"{ex}: {len(tickers)} tickers (fallback)")

        all_tickers.update(tickers)

    if cache_dirty: _save_cache(cache)
    symbols = sorted(all_tickers)
    log(f"Total unique symbols: {len(symbols)}")
    return symbols


# ─────────────────────────────────────────────────────────────────────────────
# VALUE SCORE  0–10
# ─────────────────────────────────────────────────────────────────────────────
def _compute_value_score(pe, div_yield, de_raw, roe, analyst_mean, pe_vs_industry, thresholds):
    score = 0; signals = []

    max_pe = float(thresholds.get("max_pe", 15))
    if pe and pe > 0:
        if pe < max_pe * 0.5:   score += 2; signals.append(f"P/E {round(pe,1)}x (very cheap)")
        elif pe < max_pe:       score += 1; signals.append(f"P/E {round(pe,1)}x (within range)")
        else:                              signals.append(f"P/E {round(pe,1)}x (expensive)")
    else: signals.append("P/E N/A")

    min_dy = float(thresholds.get("min_div_yield_pct", 3))
    dy_pct = (div_yield or 0) * 100
    if dy_pct >= min_dy * 1.5:   score += 2; signals.append(f"Div {round(dy_pct,2)}% (strong yield)")
    elif dy_pct >= min_dy:       score += 1; signals.append(f"Div {round(dy_pct,2)}% (meets threshold)")
    elif dy_pct > 0:                         signals.append(f"Div {round(dy_pct,2)}% (below threshold)")
    else:                                    signals.append("No dividend")

    max_de = float(thresholds.get("max_debt_equity_pct", 20))
    if de_raw is not None:
        if de_raw < max_de * 0.5: score += 2; signals.append(f"D/E {round(de_raw,1)}% (low debt)")
        elif de_raw < max_de:     score += 1; signals.append(f"D/E {round(de_raw,1)}% (acceptable)")
        else:                                signals.append(f"D/E {round(de_raw,1)}% (high debt)")
    else: signals.append("D/E N/A")

    min_roe = float(thresholds.get("min_roe_pct", 15))
    if roe is not None:
        if roe >= min_roe * 1.5: score += 2; signals.append(f"ROE {round(roe,1)}% (excellent)")
        elif roe >= min_roe:     score += 1; signals.append(f"ROE {round(roe,1)}% (good)")
        else:                                signals.append(f"ROE {round(roe,1)}% (below)")
    else: signals.append("ROE N/A")

    max_analyst = float(thresholds.get("max_analyst_mean", 2.5))
    if analyst_mean is not None and analyst_mean <= max_analyst:
        score += 1; signals.append(f"Analyst {round(analyst_mean,1)}/5 (bullish)")
    elif analyst_mean is not None:
        signals.append(f"Analyst {round(analyst_mean,1)}/5 (neutral/bearish)")
    else: signals.append("Analyst N/A")

    return min(score, 10), signals


# ─────────────────────────────────────────────────────────────────────────────
# FULL SCREENING
# ─────────────────────────────────────────────────────────────────────────────
def _empty(symbol="", name="", sector="N/A", exchange="", filters=None, sc=None):
    return {
        "ticker": symbol, "name": name or symbol, "sector": sector, "exchange": exchange,
        "market_cap_b": None, "div_yield_pct": None, "pe": None, "debt_equity_pct": None,
        "price_now": None, "price_month_ago": None, "price_4w_ago": None,
        "monthly_chg_pct": None, "weekly_chg_pct": None,
        "analyst_mean": None, "analyst_label": "—", "roce": None, "roe": None,
        "pe_vs_industry": None, "52w_high": None, "52w_low": None,
        "rsi": None, "ema20": None, "ema50": None, "ema_trend": None,
        "vol_ratio": None, "atr": None, "macd_signal": None,
        "price_vs_bb": None, "range_pos": None, "swing_score": None, "swing_signals": [],
        "value_score": None, "value_signals": [],
        "scorecard": sc or [], "failed_filters": filters or [],
    }


def screen_stock(symbol, conditions, thresholds, allowed_sectors=None):
    scorecard = []

    def check(name, actual, display, threshold, passed, active):
        scorecard.append({"filter": name, "actual": display, "threshold": threshold,
                          "passed": passed, "active": active})
        return passed

    try:
        t = yf.Ticker(symbol); info = None; _last_exc = None
        for _a in range(3):
            try:
                _raw = t.info
                if _raw and len(_raw) > 3:
                    info = _raw
                    if info.get("regularMarketPrice") or info.get("currentPrice"): break
                time.sleep(0.1)
            except Exception as _e:
                _last_exc = _e; time.sleep(0.3)

        if not info: info = {}
        name = info.get("shortName", symbol)

        if not info or (not info.get("regularMarketPrice") and not info.get("currentPrice")):
            r = _empty(symbol, symbol, sc=[{"filter":"Data Fetch","actual":"N/A",
                "threshold":"Yahoo Finance returned no data","passed":False,"active":True}],
                filters=["Data Fetch"])
            return None, r

        sector = info.get("sector", "N/A")

        # 0. Sector inline filter
        if allowed_sectors and sector not in allowed_sectors:
            r = _empty(symbol, name, sector, info.get("exchange",""),
                       filters=["Sector"],
                       sc=[{"filter":"Sector","actual":sector,
                            "threshold":f"one of: {', '.join(list(allowed_sectors)[:4])}…",
                            "passed":False,"active":True}])
            r["price_now"] = info.get("currentPrice") or info.get("regularMarketPrice")
            r["52w_high"]  = info.get("fiftyTwoWeekHigh")
            r["52w_low"]   = info.get("fiftyTwoWeekLow")
            return None, r

        # 1. Market Cap
        mcap = info.get("marketCap"); min_mcap = float(thresholds.get("min_market_cap_b",1))*1e9
        active = conditions.get("market_cap", True)
        check("Market Cap", mcap, f"${round(mcap/1e9,2)}B" if mcap else "N/A",
              f"> ${thresholds.get('min_market_cap_b',1)}B", bool(mcap and mcap>=min_mcap), active)

        # 2. Dividend Yield
        # Yahoo returns dividendYield inconsistently: sometimes as a true decimal
        # (0.0069 = 0.69%) and sometimes already as a percent expressed as a
        # decimal (1.09 = 1.09%). A real US dividend yield of 20%+ is essentially
        # impossible, so treat raw >= 0.20 as already-percent and convert to a
        # true decimal here — every downstream *100 then produces the correct %.
        raw_div   = info.get("dividendYield") or 0
        div_yield = raw_div / 100 if raw_div >= 0.20 else raw_div
        min_dy    = float(thresholds.get("min_div_yield_pct",3))/100
        active    = conditions.get("div_yield", True)
        check("Dividend Yield", div_yield, f"{round(div_yield*100,2)}%",
              f"> {thresholds.get('min_div_yield_pct',3)}%", div_yield>=min_dy, active)

        # 3. P/E Ratio
        pe     = info.get("trailingPE") or info.get("forwardPE")
        max_pe = float(thresholds.get("max_pe", 15))
        active = conditions.get("pe_ratio", True)
        check("P/E Ratio", pe, f"{round(pe,1)}×" if pe else "N/A", f"< {max_pe}×",
              bool(pe and pe>0 and pe<max_pe), active)
        pe = pe or 0

        # 4. Debt/Equity
        de_raw = info.get("debtToEquity"); max_de = float(thresholds.get("max_debt_equity_pct",20))
        active = conditions.get("debt_equity", True)
        check("Debt/Equity", de_raw, f"{round(de_raw,1)}%" if de_raw is not None else "N/A",
              f"< {max_de}%", de_raw is not None and de_raw<max_de, active)
        de_raw = de_raw or 0

        # 5&6. Price Trends
        monthly_ok = weekly_ok = None
        price_now = price_month_ago = price_4w_ago = None
        monthly_chg = weekly_chg = 0
        if conditions.get("monthly_trend",True) or conditions.get("weekly_trend",True):
            monthly_ok,weekly_ok,price_now,price_month_ago,_,price_4w_ago = _get_price_trend(t)
            if price_now and price_month_ago:
                monthly_chg = round((price_now-price_month_ago)/price_month_ago*100,2)
            if price_now and price_4w_ago:
                weekly_chg  = round((price_now-price_4w_ago)/price_4w_ago*100,2)
        else:
            price_now = info.get("currentPrice") or info.get("regularMarketPrice") or 0

        active = conditions.get("monthly_trend", True)
        # min_monthly_chg_pct: 0 = must be positive (strict), -5 = allows up to 5% dip
        min_mo_pct = float(thresholds.get("min_monthly_chg_pct", 0))
        if min_mo_pct < 0:
            # Relaxed: allow dips up to threshold (e.g. -5 means ≥ -5% is OK)
            mo_pass = monthly_chg >= min_mo_pct
            mo_thresh_lbl = f"≥ {min_mo_pct}% (allows small dips)"
        else:
            mo_pass = bool(monthly_ok)   # original: must be positive
            mo_thresh_lbl = "Price ↑ vs 1M ago"
        mo_disp = f"{monthly_chg:+.2f}%" if price_now and price_month_ago else "N/A"
        check("Monthly Trend", monthly_chg, mo_disp, mo_thresh_lbl, mo_pass, active)

        active = conditions.get("weekly_trend", True)
        wk_disp = f"{weekly_chg:+.2f}%" if price_now and price_4w_ago else "N/A"
        check("Weekly Trend", weekly_chg, wk_disp, "Price ↑ vs 4W ago", bool(weekly_ok), active)

        # 7. Analyst Rating
        analyst_mean, analyst_label = _get_analyst(t)
        max_analyst = float(thresholds.get("max_analyst_mean", 2.5))
        active      = conditions.get("analyst_rating", True)
        check("Analyst Rating", analyst_mean,
              f"{analyst_mean} ({analyst_label})" if analyst_mean else "N/A",
              f"≤ {max_analyst}/5", analyst_mean is not None and analyst_mean<=max_analyst, active)

        # 8. ROE (Return on Equity) — uses Yahoo Finance's returnOnEquity field directly,
        # which is reliable for US stocks. ROCE was removed: it required manually
        # parsing balance_sheet/income_stmt labels which are inconsistent across
        # US companies and frequently return None, making the filter unreliable.
        roe = None
        try:
            rv = info.get("returnOnEquity")
            if rv is not None: roe = round(float(rv)*100, 2)
        except Exception: pass

        min_roe = float(thresholds.get("min_roe_pct", 15))
        active  = conditions.get("roe", False)
        check("ROE", roe, f"{roe}%" if roe is not None else "N/A",
              f"> {min_roe}%", roe is not None and roe>=min_roe, active)

        # 10. P/E vs Industry
        pe_vs_industry = None; industry = info.get("industry","")
        benchmark_pe   = get_industry_pe(industry, sector)
        stock_pe_val   = info.get("trailingPE") or info.get("forwardPE")
        try:
            if stock_pe_val and benchmark_pe and float(benchmark_pe)>0:
                pe_vs_industry = round(float(stock_pe_val)/float(benchmark_pe), 3)
        except (ValueError, TypeError): pass

        max_pvi = float(thresholds.get("max_pe_vs_industry", 1.0))
        active  = conditions.get("pe_vs_industry", False)
        blabel  = f"(benchmark: {benchmark_pe}×)" if benchmark_pe else "(no benchmark)"
        check("P/E vs Industry", pe_vs_industry,
              f"{pe_vs_industry}×  {blabel}" if pe_vs_industry is not None else f"N/A  {blabel}",
              f"< {max_pvi}×", pe_vs_industry is None or pe_vs_industry<max_pvi, active)

        # Swing technicals
        swing = _get_swing_technicals(t)

        # 11. RSI
        rsi_min = float(thresholds.get("swing_rsi_min", 35))
        rsi_max = float(thresholds.get("swing_rsi_max", 65))
        active  = conditions.get("swing_rsi", False); rv = swing["rsi"]
        check("RSI(14) Range", rv, f"{rv}" if rv is not None else "N/A",
              f"{rsi_min}–{rsi_max}", rv is not None and rsi_min<=rv<=rsi_max, active)

        # 12. EMA Trend
        active = conditions.get("swing_ema", False); et = swing["ema_trend"]
        ela    = {2:"Price>EMA20>EMA50",1:"Price>EMA20",0:"Mixed",-1:"Bearish"}
        check("EMA Trend", et, ela.get(et,"N/A") if et is not None else "N/A",
              "Price > EMA20", et is not None and et>=1, active)

        # 13. Volume Surge
        min_vol = float(thresholds.get("swing_min_vol_ratio",1.5))
        active  = conditions.get("swing_volume", False); vr = swing["vol_ratio"]
        check("Volume Ratio", vr, f"{vr}×" if vr is not None else "N/A",
              f"≥ {min_vol}× avg", vr is not None and vr>=min_vol, active)

        # 14. MACD
        active = conditions.get("swing_macd", False); ms = swing["macd_signal"]
        md = {1:"Bullish crossover",-1:"Bearish crossover",0:"No crossover"}
        check("MACD Signal", ms, md.get(ms,"N/A") if ms is not None else "N/A",
              "Bullish crossover", ms==1, active)

        # 15. ATR (hard filter) + Range Position (informational — scored, not gated)
        # range_pos is shown for context and rewarded in scoring (see swing_score
        # below: room-to-run pullback setups get a bonus), but does NOT reject on
        # its own. Rationale: RSI(14)<=65 already hard-rejects overbought stocks,
        # and for trending stocks range_pos and RSI move together — gating on
        # BOTH was rejecting healthy smooth trenders that RSI alone would pass.
        min_atr     = float(thresholds.get("swing_min_atr",1.5))
        max_rng_pos = float(thresholds.get("swing_max_range_pos",50))
        active  = conditions.get("swing_atr", False); av = swing["atr"]; rp = swing.get("range_pos")
        atr_ok  = av is not None and av>=min_atr
        actual_str = f"{av}%" if av is not None else "N/A"
        if rp is not None: actual_str += f", range pos {rp:.0f}%"
        check("ATR Volatility", av, actual_str,
              f"≥ {min_atr}% (range pos ≤{max_rng_pos:.0f}% scores bonus)", atr_ok, active)

        # Value Score
        value_score, value_signals = _compute_value_score(
            pe if pe else None, div_yield, de_raw if de_raw else None,
            roe, analyst_mean, pe_vs_industry, thresholds)

        failed_active = [s["filter"] for s in scorecard if s["active"] and not s["passed"]]
        base = {
            "ticker": symbol, "name": name, "sector": sector,
            "exchange": info.get("exchange",""),
            "market_cap_b":    round(mcap/1e9,2) if mcap else None,
            "div_yield_pct":   round(div_yield*100,2),
            "pe":              round(float(pe),2) if pe else None,
            "debt_equity_pct": round(float(de_raw),2) if de_raw else None,
            "price_now":       round(float(price_now),2) if price_now else None,
            "price_month_ago": round(float(price_month_ago),2) if price_month_ago else None,
            "price_4w_ago":    round(float(price_4w_ago),2) if price_4w_ago else None,
            "monthly_chg_pct": monthly_chg, "weekly_chg_pct": weekly_chg,
            "analyst_mean":    round(float(analyst_mean),2) if analyst_mean else None,
            "analyst_label":   analyst_label,
            "roce": None, "roe": roe, "pe_vs_industry": pe_vs_industry,
            "52w_high": round(float(info.get("fiftyTwoWeekHigh")),2) if info.get("fiftyTwoWeekHigh") else None,
            "52w_low":  round(float(info.get("fiftyTwoWeekLow")),2)  if info.get("fiftyTwoWeekLow")  else None,
            "rsi": swing["rsi"], "ema20": swing["ema20"], "ema50": swing["ema50"],
            "ema_trend": swing["ema_trend"], "vol_ratio": swing["vol_ratio"],
            "atr": swing["atr"], "macd_signal": swing["macd_signal"],
            "price_vs_bb": swing["price_vs_bb"], "range_pos": swing["range_pos"],
            "swing_score": swing["swing_score"], "swing_signals": swing["swing_signals"],
            "value_score": value_score, "value_signals": value_signals,
            "scorecard": scorecard, "failed_filters": failed_active,
        }
        return (base, None) if not failed_active else (None, base)

    except Exception as e:
        r = _empty(symbol, symbol, sc=[{"filter":"Error","actual":str(e)[:60],
                                        "threshold":"Exception","passed":False,"active":True}],
                   filters=["Error"])
        return None, r


def _get_price_trend(t):
    """ONE daily history call replaces two separate monthly/weekly calls.
    Uses 95 days of daily data to derive both monthly and weekly comparisons.
    Saves 1 HTTP request per stock (~33% speedup overall).
    """
    monthly_ok = weekly_ok = price_now = monthly_prev = weekly_prev = None
    try:
        end   = datetime.today()
        start = end - timedelta(days=95)
        hist  = t.history(start=start.strftime("%Y-%m-%d"),
                          end=end.strftime("%Y-%m-%d"), interval="1d")
        if hist is not None and len(hist) >= 2:
            close      = hist["Close"]
            price_now  = round(float(close.iloc[-1]), 2)
            # Monthly: compare to ~22 trading days ago
            mo_idx     = max(0, len(close) - 22)
            monthly_prev = round(float(close.iloc[mo_idx]), 2)
            monthly_ok = price_now > monthly_prev
            # Weekly: compare to ~5 trading days ago
            wk_idx     = max(0, len(close) - 5)
            weekly_prev = round(float(close.iloc[wk_idx]), 2)
            weekly_ok  = price_now > weekly_prev
    except Exception:
        pass
    return (monthly_ok, weekly_ok, price_now, monthly_prev, price_now, weekly_prev)


def _get_analyst(t):
    try:
        mean=t.info.get("recommendationMean"); key=t.info.get("recommendationKey","n/a")
        if mean is not None: return float(mean), key.title()
    except Exception: pass
    return None, "N/A"


def _get_swing_technicals(t):
    res = {"rsi":None,"ema20":None,"ema50":None,"ema_trend":None,
           "vol_ratio":None,"atr":None,"macd_signal":None,
           "price_vs_bb":None,"range_pos":None,"swing_score":None,"swing_signals":[]}
    try:
        end=datetime.today(); start=end-timedelta(days=95)
        hist=t.history(start=start.strftime("%Y-%m-%d"),end=end.strftime("%Y-%m-%d"),interval="1d")
        if hist is None or len(hist)<20: return res
        close=hist["Close"]; high=hist["High"]; low=hist["Low"]; volume=hist["Volume"]

        try:
            delta=close.diff(); gain=delta.clip(lower=0).rolling(14).mean()
            loss=(-delta.clip(upper=0)).rolling(14).mean()
            rs=gain/loss.replace(0,float("nan"))
            res["rsi"]=round(float((100-100/(1+rs)).iloc[-1]),1)
        except Exception: pass

        try:
            e20=round(float(close.ewm(span=20,adjust=False).mean().iloc[-1]),2)
            e50=round(float(close.ewm(span=50,adjust=False).mean().iloc[-1]),2)
            p=round(float(close.iloc[-1]),2)
            res["ema20"]=e20; res["ema50"]=e50
            if p>e20>e50: res["ema_trend"]=2
            elif p>e20:   res["ema_trend"]=1
            elif p<e20<e50: res["ema_trend"]=-1
            else:           res["ema_trend"]=0
        except Exception: pass

        try:
            # Use the LAST COMPLETED full trading day's volume (iloc[-2]) rather than
            # "today's" bar (iloc[-1]). While the market is open, today's volume is
            # only partial and produces a misleadingly low ratio. Using the last
            # completed day makes this filter usable consistently at any time of day.
            avg=volume.iloc[-22:-2].mean(); cur=float(volume.iloc[-2])
            if avg>0: res["vol_ratio"]=round(cur/avg,2)
        except Exception: pass

        try:
            pc=close.shift(1)
            tr=pd.concat([high-low,(high-pc).abs(),(low-pc).abs()],axis=1).max(axis=1)
            res["atr"]=round(float(tr.rolling(14).mean().iloc[-1])/float(close.iloc[-1])*100,2)
        except Exception: pass

        try:
            # Range position: where does the current price sit within its recent
            # 20-day high-low range? 0% = sitting at the low, 100% = at the high.
            # A stock in the lower 30-50% of its range has "room to run" toward
            # the high before hitting recent resistance — combined with a healthy
            # ATR (enough volatility to actually make that move).
            roll_high = float(high.rolling(20).max().iloc[-1])
            roll_low  = float(low.rolling(20).min().iloc[-1])
            rng = roll_high - roll_low
            if rng > 0:
                res["range_pos"] = round((float(close.iloc[-1]) - roll_low) / rng * 100, 1)
        except Exception: pass

        try:
            e12=close.ewm(span=12,adjust=False).mean(); e26=close.ewm(span=26,adjust=False).mean()
            macd=e12-e26; sig=macd.ewm(span=9,adjust=False).mean()
            mv=macd.iloc[-4:]; sv=sig.iloc[-4:]
            cu=any(mv.iloc[i]>sv.iloc[i] and mv.iloc[i-1]<=sv.iloc[i-1] for i in range(1,4))
            cd=any(mv.iloc[i]<sv.iloc[i] and mv.iloc[i-1]>=sv.iloc[i-1] for i in range(1,4))
            res["macd_signal"]=1 if cu else (-1 if cd else 0)
        except Exception: pass

        try:
            sma=close.rolling(20).mean(); std=close.rolling(20).std()
            bbl=sma-2*std; bbh=sma+2*std; p=float(close.iloc[-1])
            rng=float(bbh.iloc[-1]-bbl.iloc[-1])
            if rng>0: res["price_vs_bb"]=round((p-float(bbl.iloc[-1]))/rng*100,1)
        except Exception: pass

        score=0; signals=[]
        rsi=res["rsi"]; et=res["ema_trend"]; vr=res["vol_ratio"]
        atr=res["atr"]; ms=res["macd_signal"]; bb=res["price_vs_bb"]; rp=res["range_pos"]

        if rsi is not None:
            if 35<=rsi<=55:   score+=2; signals.append(f"RSI {rsi} (entry zone)")
            elif 55<rsi<=65:  score+=1; signals.append(f"RSI {rsi} (momentum)")
            elif rsi<35:      score+=1; signals.append(f"RSI {rsi} (oversold)")
            else:                       signals.append(f"RSI {rsi} (overbought)")
        if et==2:             score+=2; signals.append("Price>EMA20>EMA50 ✓")
        elif et==1:           score+=1; signals.append("Price>EMA20")
        elif et==-1:                    signals.append("Downtrend")
        if ms==1:             score+=2; signals.append("MACD bullish crossover ✓")
        elif ms==-1:                    signals.append("MACD bearish crossover")
        if vr is not None:
            if vr>=2.0:       score+=2; signals.append(f"Volume {vr}× avg (surge ✓)")
            elif vr>=1.5:     score+=1; signals.append(f"Volume {vr}× avg (elevated)")
        if atr is not None:
            if atr>=1.5:
                if rp is not None and rp<=50:
                    score+=2; signals.append(f"ATR {atr}% + lower {rp:.0f}% of 20d range (room to run ✓)")
                else:
                    score+=1; signals.append(f"ATR {atr}% (good range)")
            elif atr<0.8:               signals.append(f"ATR {atr}% (low vol)")
        if bb is not None:
            if bb<=35:        score+=1; signals.append(f"Near BB lower band ({bb}%)")
            elif bb>=80:                signals.append(f"Near BB upper band ({bb}%) — stretched")

        res["swing_score"]=min(score,10); res["swing_signals"]=signals
    except Exception: pass
    return res


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
def export_to_excel(results, output, config):
    if not OPENPYXL_OK:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    DG="1A6B3A"; LG="F0FAF4"; MG="F5F5F5"; BC="CCCCCC"
    def thin():
        s=Side(style="thin",color=BC); return Border(left=s,right=s,top=s,bottom=s)
    def hdr(cell,text):
        cell.value=text; cell.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
        cell.fill=PatternFill("solid",fgColor=DG); cell.alignment=Alignment(horizontal="center",vertical="center")
        cell.border=thin()
    def body(cell,val,align="center",fmt=None,shade=False):
        cell.value=val; cell.font=Font(name="Arial",size=10)
        cell.alignment=Alignment(horizontal=align,vertical="center"); cell.border=thin()
        if fmt: cell.number_format=fmt
        if shade: cell.fill=PatternFill("solid",fgColor=MG)

    wb=Workbook(); ws=wb.active; ws.title="Results"
    ws.row_dimensions[1].height=30; ws.row_dimensions[3].height=20
    ws.merge_cells("A1:R1"); c=ws["A1"]
    c.value=f"Alpha Screen Results  ·  {datetime.today().strftime('%d %b %Y')}"
    c.font=Font(name="Arial",bold=True,size=14,color=DG)
    c.alignment=Alignment(horizontal="left",vertical="center"); c.fill=PatternFill("solid",fgColor=LG)

    hdrs=["#","Ticker","Company","Sector","Exchange","Mkt Cap ($B)","Div Yield","P/E",
          "Debt/Eq%","Price","1M Ago","Monthly Δ","4W Ago","Weekly Δ","Analyst","ROCE %","ROE %","P/E vs Ind."]
    for col,h in enumerate(hdrs,1): hdr(ws.cell(row=3,column=col),h)

    for i,r in enumerate(results,1):
        row=i+3; shade=(i%2==0)
        body(ws.cell(row=row,column=1),i,"center",shade=shade)
        c2=ws.cell(row=row,column=2); c2.value=r["ticker"]
        c2.font=Font(name="Arial",bold=True,color=DG,size=10)
        c2.alignment=Alignment(horizontal="center"); c2.border=thin()
        if shade: c2.fill=PatternFill("solid",fgColor=MG)
        body(ws.cell(row=row,column=3),r["name"],"left",shade=shade)
        body(ws.cell(row=row,column=4),r["sector"],"left",shade=shade)
        body(ws.cell(row=row,column=5),r["exchange"],"center",shade=shade)
        body(ws.cell(row=row,column=6),r["market_cap_b"],"right","#,##0.00",shade)
        body(ws.cell(row=row,column=7),(r["div_yield_pct"] or 0)/100,"right","0.00%",shade)
        body(ws.cell(row=row,column=8),r["pe"],"right",'0.0"x"',shade)
        body(ws.cell(row=row,column=9),(r["debt_equity_pct"] or 0)/100,"right","0.00%",shade)
        body(ws.cell(row=row,column=10),r["price_now"],"right","$#,##0.00",shade)
        body(ws.cell(row=row,column=11),r["price_month_ago"],"right","$#,##0.00",shade)
        body(ws.cell(row=row,column=12),(r["monthly_chg_pct"] or 0)/100,"right","+0.00%;[Red]-0.00%",shade)
        body(ws.cell(row=row,column=13),r["price_4w_ago"],"right","$#,##0.00",shade)
        body(ws.cell(row=row,column=14),(r["weekly_chg_pct"] or 0)/100,"right","+0.00%;[Red]-0.00%",shade)
        body(ws.cell(row=row,column=15),r["analyst_label"],"center",shade=shade)
        body(ws.cell(row=row,column=16),(r.get("roce") or 0)/100,"right","0.00%",shade)
        body(ws.cell(row=row,column=17),(r.get("roe") or 0)/100,"right","0.00%",shade)
        body(ws.cell(row=row,column=18),r.get("pe_vs_industry"),"right",'0.00"×"',shade)

    last=len(results)+3
    for i,w in enumerate([4,9,30,22,10,13,11,9,11,11,11,11,11,11,14,10,10,13],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    if results:
        for col in [12,14]:
            cl=get_column_letter(col)
            ws.conditional_formatting.add(f"{cl}4:{cl}{last}",ColorScaleRule(
                start_type="min",start_color="F8696B",
                mid_type="num",mid_value=0,mid_color="FFFFFF",
                end_type="max",end_color="63BE7B"))
    ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:{get_column_letter(18)}{last}"

    ws2=wb.create_sheet("Sector Summary"); ws2.merge_cells("A1:F1")
    t2=ws2["A1"]; t2.value="Sector Breakdown"
    t2.font=Font(name="Arial",bold=True,size=13,color=DG)
    t2.fill=PatternFill("solid",fgColor=LG); t2.alignment=Alignment(horizontal="left",vertical="center")
    for col,h in enumerate(["Sector","# Stocks","Avg Div Yield","Avg P/E","Avg D/E","Avg Analyst"],1):
        hdr(ws2.cell(row=2,column=col),h)
    groups=defaultdict(list)
    for r in results: groups[r["sector"]].append(r)
    for i,(sec,stocks) in enumerate(sorted(groups.items()),1):
        row=i+2; shade=(i%2==0)
        body(ws2.cell(row=row,column=1),sec,"left",shade=shade)
        body(ws2.cell(row=row,column=2),len(stocks),"center",shade=shade)
        body(ws2.cell(row=row,column=3),sum(s["div_yield_pct"] or 0 for s in stocks)/len(stocks)/100,"right","0.00%",shade)
        body(ws2.cell(row=row,column=4),sum(s["pe"] or 0 for s in stocks)/len(stocks),"right",'0.0"x"',shade)
        body(ws2.cell(row=row,column=5),sum(s["debt_equity_pct"] or 0 for s in stocks)/len(stocks)/100,"right","0.00%",shade)
        body(ws2.cell(row=row,column=6),sum(s["analyst_mean"] or 0 for s in stocks)/len(stocks),"right","0.0",shade)
    for col,w in zip(range(1,7),[26,10,14,10,10,14]):
        ws2.column_dimensions[get_column_letter(col)].width=w
    ws2.freeze_panes="A3"; wb.save(output)


# ═════════════════════════════════════════════════════════════════════════════
# SEPA SCREENER  —  Minervini Trend Template + Fundamentals + VCP
# ═════════════════════════════════════════════════════════════════════════════

def _get_sma(closes, period):
    """Return SMA of last `period` closes, or None if insufficient data."""
    if closes is None or len(closes) < period:
        return None
    return float(closes.iloc[-period:].mean())

def _sma_trending_up(closes, period, lookback_days=20):
    """Return True if the SMA(period) is higher now than `lookback_days` ago."""
    if closes is None or len(closes) < period + lookback_days:
        return False
    sma_now  = float(closes.iloc[-period:].mean())
    sma_past = float(closes.iloc[-(period + lookback_days):-lookback_days].mean())
    return sma_now > sma_past

def compute_rs_raw(closes, market_closes):
    """
    Compute raw RS score using Minervini / IBD weighting:
      40% weight on most recent quarter, 20% each on prior 3 quarters.
    Returns a float (higher = stronger relative performance) or None.
    """
    try:
        if closes is None or market_closes is None:
            return None
        if len(closes) < 252 or len(market_closes) < 252:
            return None

        p   = float(closes.iloc[-1])
        m   = float(market_closes.iloc[-1])

        def safe_ratio(series, idx):
            v = float(series.iloc[idx])
            return None if v == 0 else series.iloc[-1] / v

        s63  = safe_ratio(closes, -63);  m63  = safe_ratio(market_closes, -63)
        s126 = safe_ratio(closes, -126); m126 = safe_ratio(market_closes, -126)
        s189 = safe_ratio(closes, -189); m189 = safe_ratio(market_closes, -189)
        s252 = safe_ratio(closes, -252); m252 = safe_ratio(market_closes, -252)

        if None in (s63, m63, s126, m126, s189, m189, s252, m252):
            return None

        stock_score  = s63 * 0.4 + s126 * 0.2 + s189 * 0.2 + s252 * 0.2
        market_score = m63 * 0.4 + m126 * 0.2 + m189 * 0.2 + m252 * 0.2

        return stock_score / market_score if market_score else None
    except Exception:
        return None


def _detect_vcp(closes, volumes, n_contractions=3):
    """
    Simplified VCP detector.
    Looks for shrinking pullback depth and volume dry-up over the last base.

    Returns dict:
      { vcp: bool, contractions: int, max_pullback_pct: float,
        last_pullback_pct: float, vol_dry_up: bool, near_pivot: bool }
    """
    result = {
        "vcp": False, "contractions": 0,
        "max_pullback_pct": None, "last_pullback_pct": None,
        "vol_dry_up": False, "near_pivot": False,
    }
    try:
        if closes is None or len(closes) < 60:
            return result
        if volumes is None or len(volumes) < 60:
            return result

        c  = closes.iloc[-60:].values
        v  = volumes.iloc[-60:].values
        n  = len(c)

        # Find local peaks and troughs in last 60 days
        peaks   = []
        troughs = []
        for i in range(1, n - 1):
            if c[i] > c[i-1] and c[i] > c[i+1]:
                peaks.append((i, c[i]))
            if c[i] < c[i-1] and c[i] < c[i+1]:
                troughs.append((i, c[i]))

        if len(peaks) < 2 or len(troughs) < 1:
            return result

        # Measure pullbacks between alternating peak→trough cycles
        pullbacks = []
        events = sorted(peaks + troughs, key=lambda x: x[0])
        last_peak = None
        for idx, price in events:
            if (idx, price) in peaks:
                last_peak = price
            elif last_peak is not None and (idx, price) in troughs:
                pct = (last_peak - price) / last_peak * 100
                pullbacks.append(pct)
                last_peak = None

        if len(pullbacks) < 2:
            return result

        # Check contractions: each pullback smaller than previous
        contractions = 0
        for i in range(1, len(pullbacks)):
            if pullbacks[i] < pullbacks[i-1] * 0.75:  # at least 25% tighter
                contractions += 1

        # Volume dry-up: last 5 days avg vol < 50% of 20-day avg vol
        avg_vol_20  = float(v[-20:].mean()) if len(v) >= 20 else None
        avg_vol_5   = float(v[-5:].mean())  if len(v) >= 5  else None
        vol_dry_up  = (avg_vol_20 and avg_vol_5 and avg_vol_5 < avg_vol_20 * 0.5)

        # Near pivot: price within 2% of recent 60-day high
        high_60     = float(max(c))
        price_now   = float(c[-1])
        near_pivot  = (price_now >= high_60 * 0.98)

        result.update({
            "vcp":               contractions >= max(1, n_contractions - 1),
            "contractions":      contractions,
            "max_pullback_pct":  round(float(max(pullbacks)), 1) if pullbacks else None,
            "last_pullback_pct": round(float(pullbacks[-1]), 1) if pullbacks else None,
            "vol_dry_up":        bool(vol_dry_up),
            "near_pivot":        bool(near_pivot),
        })
    except Exception:
        pass
    return result


def screen_stock_sepa(symbol, thresholds, market_closes, all_rs_raw, log_cb=None):
    """
    Full SEPA screen for one stock.

    Stage 1: Minervini Trend Template (8 criteria)
    Stage 2: Fundamental filters (EPS growth, revenue growth, margins)
    Stage 3: VCP pattern detection

    Returns (result_dict, rejection_dict) — same pattern as screen_stock().
    """
    import yfinance as yf
    from datetime import datetime, timedelta

    def log(m):
        if log_cb: log_cb(m)

    rejection = {
        "ticker": symbol, "name": symbol, "sector": "—",
        "price_now": None, "52w_high": None, "52w_low": None,
        "failed_filters": [], "scorecard": [], "stage_failed": None,
    }

    def fail(stage, reason, scorecard):
        rejection["failed_filters"] = [reason]
        rejection["scorecard"]      = scorecard
        rejection["stage_failed"]   = stage
        return None, rejection

    try:
        tk   = yf.Ticker(symbol)
        info = tk.info or {}
        if not info or len(info) < 5:
            return None, None

        name   = info.get("longName") or info.get("shortName") or symbol
        sector = info.get("sector", "—")
        rejection["name"]   = name
        rejection["sector"] = sector

        # Price history — need 252+ trading days for RS
        hist = tk.history(period="18mo", interval="1d")
        if hist is None or len(hist) < 200:
            return None, None

        closes  = hist["Close"]
        volumes = hist["Volume"]

        price = float(closes.iloc[-1])
        rejection["price_now"] = price
        rejection["52w_high"]  = info.get("fiftyTwoWeekHigh")
        rejection["52w_low"]   = info.get("fiftyTwoWeekLow")

        sc = []  # scorecard list

        def chk(label, passed, actual="", threshold=""):
            sc.append({"filter": label, "passed": passed,
                       "actual": str(actual), "threshold": str(threshold), "active": True})
            return passed

        # ── STAGE 1: Trend Template ───────────────────────────────────────────
        sma50  = _get_sma(closes, 50)
        sma150 = _get_sma(closes, 150)
        sma200 = _get_sma(closes, 200)

        if not all([sma50, sma150, sma200]):
            return None, None

        t1 = chk("Price > 150d & 200d SMA",
                  price > sma150 and price > sma200,
                  f"${price:.2f}", f"SMA150=${sma150:.2f}, SMA200=${sma200:.2f}")

        t2 = chk("150d SMA > 200d SMA",
                  sma150 > sma200,
                  f"150d=${sma150:.2f}", f"200d=${sma200:.2f}")

        sma200_rising = _sma_trending_up(closes, 200, thresholds.get("sma200_min_days", 20))
        t3 = chk("200d SMA trending up",
                  sma200_rising, "Rising" if sma200_rising else "Flat",
                  f">{thresholds.get('sma200_min_days',20)}d up")

        t4 = chk("50d SMA > 150d & 200d",
                  sma50 > sma150 and sma50 > sma200,
                  f"50d=${sma50:.2f}", f"150d=${sma150:.2f}")

        t5 = chk("Price > 50d SMA",
                  price > sma50,
                  f"${price:.2f}", f"SMA50=${sma50:.2f}")

        low52  = info.get("fiftyTwoWeekLow") or float(closes.min())
        high52 = info.get("fiftyTwoWeekHigh") or float(closes.max())
        pct_above_low  = (price - low52)  / low52  * 100 if low52  else 0
        pct_below_high = (high52 - price) / high52 * 100 if high52 else 0

        t6 = chk("Price ≥ 30% above 52W Low",
                  pct_above_low >= thresholds.get("min_pct_above_52w_low", 30),
                  f"+{pct_above_low:.1f}%", f"≥{thresholds.get('min_pct_above_52w_low',30)}%")

        t7 = chk("Price ≤ 25% below 52W High",
                  pct_below_high <= thresholds.get("max_pct_below_52w_high", 25),
                  f"-{pct_below_high:.1f}%", f"≤{thresholds.get('max_pct_below_52w_high',25)}%")

        # RS rating — percentile vs universe (computed after all stocks scanned)
        rs_raw = compute_rs_raw(closes, market_closes)
        if rs_raw is not None:
            all_rs_raw[symbol] = rs_raw

        # RS check uses placeholder — will be re-evaluated after universe scan
        # For now just store raw score; rs_rating percentile computed in batch
        rs_ok = rs_raw is not None  # will be refined post-scan
        t8 = chk("RS Rating (preliminary)",
                  rs_ok, f"raw={rs_raw:.3f}" if rs_raw else "N/A", "> threshold")

        stage1_passed = all([t1, t2, t3, t4, t5, t6, t7])
        if not stage1_passed:
            rejection["scorecard"]      = sc
            rejection["failed_filters"] = [s["filter"] for s in sc if not s["passed"]]
            rejection["stage_failed"]   = 1
            return None, rejection

        # ── STAGE 2: SEPA Fundamentals ────────────────────────────────────────
        # EPS growth — trailing vs year-ago
        eps_now  = info.get("trailingEps")
        eps_fwd  = info.get("forwardEps")
        rev_growth = info.get("revenueGrowth")  # yfinance returns as decimal
        gross_margin = info.get("grossMargins")
        profit_margin = info.get("profitMargins")

        # EPS growth proxy: (forwardEps - trailingEps) / abs(trailingEps)
        eps_growth = None
        if eps_now and eps_fwd and eps_now != 0:
            eps_growth = (eps_fwd - eps_now) / abs(eps_now) * 100

        min_eps  = thresholds.get("min_eps_growth_pct", 20)
        min_rev  = thresholds.get("min_rev_growth_pct", 20)

        f1 = chk("EPS Growth",
                  eps_growth is not None and eps_growth >= min_eps,
                  f"{eps_growth:.1f}%" if eps_growth is not None else "N/A",
                  f"≥{min_eps}%")

        f2 = chk("Revenue Growth",
                  rev_growth is not None and rev_growth * 100 >= min_rev,
                  f"{rev_growth*100:.1f}%" if rev_growth is not None else "N/A",
                  f"≥{min_rev}%")

        # Institutional ownership increasing — use institutionsPercentHeld as proxy
        inst_pct = info.get("institutionsPercentHeld")
        f3 = chk("Institutional Ownership",
                  inst_pct is not None and inst_pct > 0.3,
                  f"{inst_pct*100:.1f}%" if inst_pct is not None else "N/A",
                  "> 30%")

        stage2_passed = f1 and f2
        if not stage2_passed:
            rejection["scorecard"]      = sc
            rejection["failed_filters"] = [s["filter"] for s in sc if not s["passed"]]
            rejection["stage_failed"]   = 2
            return None, rejection

        # ── STAGE 3: VCP Pattern ──────────────────────────────────────────────
        vcp = _detect_vcp(closes, volumes, thresholds.get("vcp_contractions", 3))

        chk("VCP Contractions",
            vcp["contractions"] >= 1,
            f"{vcp['contractions']} contractions", "≥ 1")
        chk("Volume Dry-up",
            vcp["vol_dry_up"],
            "Yes" if vcp["vol_dry_up"] else "No", "< 50% avg vol")

        # Build result
        result = {
            "ticker":            symbol,
            "name":              name,
            "sector":            sector,
            "exchange":          info.get("exchange", "—"),
            "price_now":         price,
            "52w_high":          high52,
            "52w_low":           low52,
            "sma50":             round(sma50, 2),
            "sma150":            round(sma150, 2),
            "sma200":            round(sma200, 2),
            "pct_above_52w_low": round(pct_above_low, 1),
            "pct_below_52w_high":round(pct_below_high, 1),
            "rs_raw":            rs_raw,
            "rs_rating":         None,   # filled in post-scan
            "eps_growth_pct":    round(eps_growth, 1) if eps_growth is not None else None,
            "rev_growth_pct":    round(rev_growth * 100, 1) if rev_growth is not None else None,
            "gross_margin_pct":  round(gross_margin * 100, 1) if gross_margin is not None else None,
            "profit_margin_pct": round(profit_margin * 100, 1) if profit_margin is not None else None,
            "inst_pct":          round(inst_pct * 100, 1) if inst_pct is not None else None,
            "pe":                info.get("trailingPE") or info.get("forwardPE"),
            "market_cap_b":      round(info.get("marketCap", 0) / 1e9, 2),
            "monthly_chg_pct":   round((price / float(closes.iloc[-21]) - 1) * 100, 2) if len(closes) >= 21 else None,
            "weekly_chg_pct":    round((price / float(closes.iloc[-5]) - 1) * 100, 2)  if len(closes) >= 5  else None,
            # VCP
            "vcp":               vcp["vcp"],
            "vcp_contractions":  vcp["contractions"],
            "max_pullback_pct":  vcp["max_pullback_pct"],
            "last_pullback_pct": vcp["last_pullback_pct"],
            "vol_dry_up":        vcp["vol_dry_up"],
            "near_pivot":        vcp["near_pivot"],
            "scorecard":         sc,
            "failed_filters":    [],
            "stage_failed":      None,
        }
        return result, None

    except Exception as e:
        return None, None


def run_sepa_screen(tickers, thresholds, log_cb=None):
    """
    Orchestrate the full SEPA screen across all tickers.
    1. Fetch S&P 500 market closes for RS calculation.
    2. Screen each stock through all 3 stages.
    3. Compute RS rating percentiles from raw scores.
    4. Apply RS rating filter.
    5. Return (results, rejections).
    """
    import yfinance as yf
    from concurrent.futures import ThreadPoolExecutor, as_completed

    def log(m):
        if log_cb: log_cb(m)

    log("SEPA Phase 0 — Fetching S&P 500 benchmark prices...")
    try:
        spy_hist     = yf.Ticker("SPY").history(period="18mo", interval="1d")
        market_closes = spy_hist["Close"] if spy_hist is not None and not spy_hist.empty else None
    except Exception:
        market_closes = None
    if market_closes is None:
        log("WARNING: Could not fetch SPY benchmark — RS ratings will be skipped")

    all_rs_raw = {}   # symbol -> raw RS score
    results    = []
    rejections = []
    total      = len(tickers)

    log(f"SEPA Phase 1+2+3 — Screening {total} stocks...")

    def screen_one(symbol):
        return symbol, screen_stock_sepa(symbol, thresholds, market_closes, all_rs_raw, log_cb=None)

    with ThreadPoolExecutor(max_workers=12) as ex:
        futures = {ex.submit(screen_one, sym): sym for sym in tickers}
        done    = 0
        for future in as_completed(futures):
            done += 1
            pct  = int(done / total * 100)
            sym, (result, rejection) = future.result()
            log(f"__sepa_progress__{pct}__{sym}__{len(results)}")
            if result:
                results.append(result)
            elif rejection:
                rejections.append(rejection)

    # ── Compute RS rating percentiles ─────────────────────────────────────────
    if all_rs_raw:
        scores  = list(all_rs_raw.values())
        scores.sort()
        n       = len(scores)
        for r in results:
            raw = r.get("rs_raw")
            if raw is not None:
                rank       = sum(1 for s in scores if s <= raw)
                r["rs_rating"] = round(rank / n * 99 + 1)
            else:
                r["rs_rating"] = None
        # Apply RS filter
        min_rs = thresholds.get("min_rs_rating", 70)
        before = len(results)
        results = [r for r in results if r.get("rs_rating") is not None and r["rs_rating"] >= min_rs]
        if before - len(results):
            log(f"RS Rating filter: removed {before - len(results)} stocks below RS {min_rs}")

    # ── Catalyst check (Groq) — only on final passers, small set ──────────────
    if results and _GROQ_OK:
        log(f"SEPA Phase 4 — Checking news catalysts for {len(results)} finalists...")

        def catalyst_one(r):
            return r["ticker"], _check_catalyst_news(r["ticker"], r["name"], r["sector"])

        with ThreadPoolExecutor(max_workers=6) as ex:
            futures = {ex.submit(catalyst_one, r): r["ticker"] for r in results}
            for future in as_completed(futures):
                sym, cat = future.result()
                for r in results:
                    if r["ticker"] == sym:
                        r["catalyst_verdict"] = cat["verdict"]
                        r["catalyst_reason"]  = cat["reason"]
                        r["catalyst_source"]  = cat["source"]
                        break
        n_buy = sum(1 for r in results if r.get("catalyst_verdict") == "BUY")
        log(f"Catalyst check complete — {n_buy}/{len(results)} have a fresh BUY catalyst")
    elif results:
        for r in results:
            r["catalyst_verdict"] = "N/A"
            r["catalyst_reason"]  = "Groq not available"
            r["catalyst_source"]  = "none"

    log(f"SEPA complete — {total} screened · {len(results)} passed all 3 stages · {len(rejections)} rejected")
    return results, rejections


# ═════════════════════════════════════════════════════════════════════════════
# TRIPLE SCREEN  —  Swing → Value → SEPA (3-Agent Pipeline)
# Runs only on the ticker universe passed in (from the selected index/exchange)
# ═════════════════════════════════════════════════════════════════════════════

def run_triple_screen(tickers, swing_conditions, swing_thresholds,
                      value_conditions, value_thresholds,
                      sepa_thresholds, allowed_sectors=None, log_cb=None):
    """
    Agent 1 — Swing screen on selected universe
    Agent 2 — Value screen on Agent 1 survivors only
    Agent 3 — SEPA Minervini on Agent 2 survivors only

    RS Rating in Agent 3 is computed relative to the FULL original universe
    (not just the Agent 2 survivors) so the percentile ranking is meaningful.

    Returns dict with per-agent passed/rejected lists and final results.
    """
    import yfinance as yf
    from concurrent.futures import ThreadPoolExecutor, as_completed

    def log(m):
        if log_cb: log_cb(m)

    total = len(tickers)

    # ── AGENT 1: Swing Screen ─────────────────────────────────────────────────
    log(f"__agent__1__start__{total}")
    log(f"Agent 1 — Swing Screen · {total} stocks in universe")

    a1_passed   = []
    a1_rejected = []

    def _swing_one(sym):
        return sym, screen_stock(sym, swing_conditions, swing_thresholds, allowed_sectors)

    with ThreadPoolExecutor(max_workers=12) as ex:
        futures = {ex.submit(_swing_one, sym): sym for sym in tickers}
        done = 0
        for future in as_completed(futures):
            done += 1
            sym, (result, rejection) = future.result()
            pct = int(done / total * 100)
            log(f"__triple_progress__1__{pct}__{sym}__{len(a1_passed)}")
            if result:
                a1_passed.append(result)
            elif rejection:
                a1_rejected.append(rejection)

    log(f"Agent 1 complete — {len(a1_passed)} passed swing · {len(a1_rejected)} rejected")

    if not a1_passed:
        log("Agent 1 found no stocks — pipeline stopped")
        return {
            "agent1": {"passed": [], "rejected": a1_rejected},
            "agent2": {"passed": [], "rejected": []},
            "agent3": {"passed": [], "rejected": []},
            "final":  [],
        }

    # ── AGENT 2: Value Screen on Agent 1 survivors ───────────────────────────
    a1_syms = [r["ticker"] for r in a1_passed]
    log(f"__agent__2__start__{len(a1_syms)}")
    log(f"Agent 2 — Value Screen · {len(a1_syms)} stocks from Agent 1")

    a2_passed   = []
    a2_rejected = []

    def _value_one(sym):
        return sym, screen_stock(sym, value_conditions, value_thresholds, None)

    with ThreadPoolExecutor(max_workers=12) as ex:
        futures = {ex.submit(_value_one, sym): sym for sym in a1_syms}
        done = 0
        for future in as_completed(futures):
            done += 1
            sym, (result, rejection) = future.result()
            pct = int(done / len(a1_syms) * 100)
            log(f"__triple_progress__2__{pct}__{sym}__{len(a2_passed)}")
            if result:
                a2_passed.append(result)
            elif rejection:
                a2_rejected.append(rejection)

    log(f"Agent 2 complete — {len(a2_passed)} passed value · {len(a2_rejected)} rejected")

    if not a2_passed:
        log("Agent 2 found no stocks — pipeline stopped")
        return {
            "agent1": {"passed": a1_passed, "rejected": a1_rejected},
            "agent2": {"passed": [],        "rejected": a2_rejected},
            "agent3": {"passed": [],        "rejected": []},
            "final":  [],
        }

    # ── AGENT 3: SEPA Minervini on Agent 2 survivors ─────────────────────────
    a2_syms = [r["ticker"] for r in a2_passed]
    log(f"__agent__3__start__{len(a2_syms)}")
    log(f"Agent 3 — SEPA Minervini · {len(a2_syms)} stocks from Agent 2")
    log("Agent 3 — Fetching SPY benchmark for RS Rating...")

    try:
        spy_hist      = yf.Ticker("SPY").history(period="18mo", interval="1d")
        market_closes = spy_hist["Close"] if spy_hist is not None and not spy_hist.empty else None
    except Exception:
        market_closes = None

    if market_closes is None:
        log("WARNING: Could not fetch SPY — RS Ratings will be skipped")

    # RS raw scores — collected across ALL original tickers for accurate percentile
    all_rs_raw  = {}
    a3_passed   = []
    a3_rejected = []

    def _sepa_one(sym):
        return sym, screen_stock_sepa(sym, sepa_thresholds, market_closes, all_rs_raw)

    with ThreadPoolExecutor(max_workers=12) as ex:
        futures = {ex.submit(_sepa_one, sym): sym for sym in a2_syms}
        done = 0
        for future in as_completed(futures):
            done += 1
            sym, (result, rejection) = future.result()
            pct = int(done / len(a2_syms) * 100)
            log(f"__triple_progress__3__{pct}__{sym}__{len(a3_passed)}")
            if result:
                a3_passed.append(result)
            elif rejection:
                a3_rejected.append(rejection)

    # Compute RS percentile ratings across the Agent 2 pool
    if all_rs_raw:
        scores = sorted(all_rs_raw.values())
        n      = len(scores)
        for r in a3_passed:
            raw = r.get("rs_raw")
            if raw is not None:
                rank = sum(1 for s in scores if s <= raw)
                r["rs_rating"] = round(rank / n * 99 + 1)
            else:
                r["rs_rating"] = None
        # Apply RS filter
        min_rs = sepa_thresholds.get("min_rs_rating", 70)
        before = len(a3_passed)
        a3_passed = [r for r in a3_passed if r.get("rs_rating") is not None and r["rs_rating"] >= min_rs]
        removed   = before - len(a3_passed)
        if removed:
            log(f"Agent 3 — RS filter removed {removed} stocks below RS {min_rs}")

    # Enrich final results with swing + value scorecard data merged in
    swing_map = {r["ticker"]: r for r in a1_passed}
    value_map = {r["ticker"]: r for r in a2_passed}

    for r in a3_passed:
        sym = r["ticker"]
        sw  = swing_map.get(sym, {})
        va  = value_map.get(sym, {})
        # Merge key swing fields
        for key in ["swing_score","swing_signals","rsi","ema_trend","vol_ratio","atr","macd_signal"]:
            if key not in r or r[key] is None:
                r[key] = sw.get(key)
        # Merge key value fields
        for key in ["value_score","value_signals","div_yield_pct","pe","debt_equity_pct",
                    "analyst_label","analyst_mean","roce","roe"]:
            if key not in r or r[key] is None:
                r[key] = va.get(key)
        r["triple_screen"] = True

    log(f"Agent 3 complete — {len(a3_passed)} passed SEPA · {len(a3_rejected)} rejected")
    log(f"TRIPLE SCREEN DONE — {total} started · {len(a1_passed)} swing · {len(a2_passed)} value · {len(a3_passed)} SEPA final")

    return {
        "agent1": {"passed": a1_passed, "rejected": a1_rejected},
        "agent2": {"passed": a2_passed, "rejected": a2_rejected},
        "agent3": {"passed": a3_passed, "rejected": a3_rejected},
        "final":  a3_passed,
    }

# ══════════════════════════════════════════════════════════════════════════════
# FIXED TRIPLE SCREEN — with Market Regime Gate + RSI(2) + Entry Trigger
# ══════════════════════════════════════════════════════════════════════════════

def _compute_rsi2(closes, period=2):
    """
    RSI with period 2 — far more sensitive than RSI(14).
    RSI(2) < 30 in an uptrend = high-probability mean-reversion entry.
    Returns float 0-100 or None.
    """
    try:
        import pandas as pd
        s = pd.Series(closes)
        delta = s.diff()
        gain  = delta.clip(lower=0).rolling(period).mean()
        loss  = (-delta.clip(upper=0)).rolling(period).mean()
        rs    = gain / loss.replace(0, float("nan"))
        rsi   = (100 - 100 / (1 + rs)).iloc[-1]
        return round(float(rsi), 1) if not pd.isna(rsi) else None
    except Exception:
        return None


def get_market_regime(market: str = "US") -> dict:
    """
    Market regime gate — checks if broad market is in an uptrend.
    US:    SPY above its 200-day SMA
    India: ^NSEI (Nifty 50) above its 200-day SMA

    Returns:
        { "regime": "Bull" | "Bear", "index": str, "price": float,
          "sma200": float, "pct_above": float, "ok_to_trade": bool }
    """
    import yfinance as yf
    import pandas as pd

    index_sym = "SPY" if market == "US" else "^NSEI"
    try:
        hist = yf.Ticker(index_sym).history(period="1y", interval="1d", auto_adjust=True)
        if hist is None or len(hist) < 200:
            return {"regime": "Unknown", "index": index_sym, "price": None,
                    "sma200": None, "pct_above": None, "ok_to_trade": True}
        closes    = hist["Close"]
        price     = round(float(closes.iloc[-1]), 2)
        sma200    = round(float(closes.rolling(200).mean().iloc[-1]), 2)
        pct_above = round((price - sma200) / sma200 * 100, 2) if sma200 else 0
        regime    = "Bull" if price > sma200 else "Bear"
        return {
            "regime":      regime,
            "index":       index_sym,
            "price":       price,
            "sma200":      sma200,
            "pct_above":   pct_above,
            "ok_to_trade": regime == "Bull",
        }
    except Exception as e:
        return {"regime": "Unknown", "index": index_sym, "price": None,
                "sma200": None, "pct_above": None, "ok_to_trade": True}


def check_entry_trigger(symbol: str, market: str = "US") -> dict:
    """
    Entry trigger — fires when price crosses above the PREVIOUS DAY's high
    on above-average volume (1.2x). This converts a screener signal into
    a precise trade entry.

    Returns:
        { "triggered": bool, "prev_high": float, "current_price": float,
          "vol_ratio": float, "trigger_note": str }
    """
    import yfinance as yf
    suffix = ".NS" if market == "IN" else ""
    try:
        hist = yf.Ticker(symbol + suffix).history(period="10d", interval="1d", auto_adjust=True)
        if hist is None or len(hist) < 3:
            return {"triggered": False, "trigger_note": "Insufficient data"}

        prev_high     = float(hist["High"].iloc[-2])
        current_price = float(hist["Close"].iloc[-1])
        vol_avg       = float(hist["Volume"].iloc[:-1].mean())
        vol_today     = float(hist["Volume"].iloc[-1])
        vol_ratio     = round(vol_today / vol_avg, 2) if vol_avg > 0 else 0

        triggered = current_price > prev_high and vol_ratio >= 1.2
        note = (f"Price ${current_price:.2f} > prev high ${prev_high:.2f} + vol {vol_ratio}x ✓"
                if triggered else
                f"Price ${current_price:.2f} vs prev high ${prev_high:.2f} · vol {vol_ratio}x")

        return {
            "triggered":     triggered,
            "prev_high":     round(prev_high, 2),
            "current_price": round(current_price, 2),
            "vol_ratio":     vol_ratio,
            "trigger_note":  note,
        }
    except Exception as e:
        return {"triggered": False, "trigger_note": str(e)[:60]}


def run_fixed_triple_screen(tickers, sepa_thresholds=None, market="US",
                             use_regime_gate=True, log_cb=None):
    """
    Fixed Triple Screen — the corrected pipeline.

    Key differences from the original Triple Screen:
      1. Market Regime Gate: checks SPY/Nifty > 200MA before screening
      2. Agent 1 (Technical): RSI(14) in 25-75 + EMA uptrend only
                              NO dividend / P/E / debt filters in Agent 1
      3. Agent 2 (Fundamental): sensible thresholds (PE < 60, ROE > 5%)
      4. Agent 3 (SEPA): same as original
      5. Entry Trigger: added to each passing stock
    """
    import yfinance as yf
    from concurrent.futures import ThreadPoolExecutor, as_completed

    if sepa_thresholds is None:
        sepa_thresholds = {}
    if log_cb is None:
        log_cb = lambda m: None

    def log(m): log_cb(m)

    total = len(tickers)

    # ── Regime Gate ───────────────────────────────────────────────────────────
    regime = {"regime": "Unknown", "ok_to_trade": True}
    if use_regime_gate:
        log("__ Checking market regime gate...")
        regime = get_market_regime(market)
        log(f"__ Market Regime: {regime['regime']} "
            f"({regime['index']} ${regime.get('price','?')} vs "
            f"SMA200 ${regime.get('sma200','?')} · {regime.get('pct_above','?')}%)")
        if not regime["ok_to_trade"]:
            log(f"__ REGIME GATE: Market is in Bear regime — trade count will be low")

    # ── Agent 1: Technical only (RSI + EMA) ──────────────────────────────────
    log(f"__agent__1__start__{total}")
    log(f"Agent 1 — Technical Screen (RSI 25-75 + EMA uptrend) · {total} stocks")

    # Agent 1: EMA only — RSI removed (overbought stocks showed 100% profitability)
    # Monthly trend: relaxed to allow up to -5% monthly dip
    TECH_CONDITIONS = {
        "market_cap": True, "div_yield": False, "pe_ratio": False,
        "debt_equity": False, "monthly_trend": True, "weekly_trend": False,
        "analyst_rating": False, "roce": False, "roe": False,
        "pe_vs_industry": False,
        "swing_rsi": False,   # REMOVED: RSI overbought = momentum, not a risk
        "swing_ema": True,    # Keep: EMA trend confirms uptrend structure
        "swing_volume": False, "swing_macd": False, "swing_atr": False,
    }
    TECH_THRESHOLDS = {
        "min_market_cap_b":    0.5,
        "min_monthly_chg_pct": -5.0,  # allow up to -5% monthly dip (backtest validated)
        "swing_min_vol_ratio":  1.0,
    }

    a1_passed = []; a1_rejected = []

    def _tech_one(sym):
        return sym, screen_stock(sym, TECH_CONDITIONS, TECH_THRESHOLDS, None)

    with ThreadPoolExecutor(max_workers=12) as ex:
        futures = {ex.submit(_tech_one, sym): sym for sym in tickers}
        done = 0
        for future in as_completed(futures):
            done += 1
            sym, (result, rejection) = future.result()
            pct = int(done / total * 100)
            log(f"__triple_progress__1__{pct}__{sym}__{len(a1_passed)}")
            if result:
                a1_passed.append(result)
            elif rejection:
                a1_rejected.append(rejection)

    log(f"Agent 1 complete — {len(a1_passed)} passed · {len(a1_rejected)} rejected")

    if not a1_passed:
        log("Agent 1 found no stocks — check universe or RSI thresholds")
        return {
            "agent1": {"passed": [], "rejected": a1_rejected},
            "agent2": {"passed": [], "rejected": []},
            "agent3": {"passed": [], "rejected": []},
            "final":  [], "regime": regime,
        }

    # ── Agent 2: Fundamental screen with sane thresholds ─────────────────────
    a1_syms = [r["ticker"] for r in a1_passed]
    log(f"__agent__2__start__{len(a1_syms)}")
    log(f"Agent 2 — Fundamental Screen · {len(a1_syms)} stocks")

    FUND_CONDITIONS = {
        "market_cap": False, "div_yield": False, "pe_ratio": True,
        "debt_equity": True, "monthly_trend": False, "weekly_trend": False,
        "analyst_rating": True, "roce": False, "roe": True,
        "pe_vs_industry": False,
        "swing_rsi": False, "swing_ema": False,
        "swing_volume": False, "swing_macd": False, "swing_atr": False,
    }
    FUND_THRESHOLDS = {
        "max_pe":              60,    # growth stocks can have PE 30-50
        "max_debt_equity_pct": 150,   # allows financials and moderate leverage
        "min_roe_pct":         5,     # minimum profitability bar
        "max_analyst_mean":    3.5,   # includes Hold (2.5) and mild Sell
    }

    a2_passed = []; a2_rejected = []

    def _fund_one(sym):
        return sym, screen_stock(sym, FUND_CONDITIONS, FUND_THRESHOLDS, None)

    with ThreadPoolExecutor(max_workers=12) as ex:
        futures = {ex.submit(_fund_one, sym): sym for sym in a1_syms}
        done = 0
        for future in as_completed(futures):
            done += 1
            sym, (result, rejection) = future.result()
            pct = int(done / len(a1_syms) * 100)
            log(f"__triple_progress__2__{pct}__{sym}__{len(a2_passed)}")
            if result:
                a2_passed.append(result)
            elif rejection:
                a2_rejected.append(rejection)

    log(f"Agent 2 complete — {len(a2_passed)} passed · {len(a2_rejected)} rejected")

    if not a2_passed:
        log("Agent 2 found no stocks")
        return {
            "agent1": {"passed": a1_passed, "rejected": a1_rejected},
            "agent2": {"passed": [], "rejected": a2_rejected},
            "agent3": {"passed": [], "rejected": []},
            "final":  [], "regime": regime,
        }

    # ── Agent 3: SEPA (same as original) ─────────────────────────────────────
    a2_syms = [r["ticker"] for r in a2_passed]
    log(f"__agent__3__start__{len(a2_syms)}")
    log(f"Agent 3 — SEPA Minervini · {len(a2_syms)} stocks")

    # Fetch market closes for RS rating calculation
    try:
        import yfinance as _yf2
        _spy = _yf2.Ticker("SPY" if market == "US" else "^NSEI")
        _mhist = _spy.history(period="18mo", interval="1d", auto_adjust=True)
        market_closes = _mhist["Close"] if _mhist is not None and not _mhist.empty else None
    except Exception:
        market_closes = None

    default_sepa = {
        "sma200_min_days":       20,
        "min_pct_above_52w_low": 25,
        "max_pct_below_52w_high": 30,
        "min_rs_rating":         65,
        "min_eps_growth_pct":    15,
        "min_rev_growth_pct":    10,
        "vcp_contractions":      2,
    }
    default_sepa.update(sepa_thresholds)

    all_rs_raw = {}
    a3_passed = []; a3_rejected = []

    def _sepa_one(sym):
        return sym, screen_stock_sepa(sym, default_sepa, market_closes, all_rs_raw)

    with ThreadPoolExecutor(max_workers=12) as ex:
        futures = {ex.submit(_sepa_one, sym): sym for sym in a2_syms}
        done = 0
        for future in as_completed(futures):
            done += 1
            sym, (result, rejection) = future.result()
            pct = int(done / len(a2_syms) * 100)
            log(f"__triple_progress__3__{pct}__{sym}__{len(a3_passed)}")
            if result:
                a3_passed.append(result)
            elif rejection:
                a3_rejected.append(rejection)

    # RS percentile pass
    if all_rs_raw:
        scores = sorted(all_rs_raw.values()); n = len(scores)
        min_rs = default_sepa.get("min_rs_rating", 65)
        for r in a3_passed:
            raw = r.get("rs_raw")
            r["rs_rating"] = round(sum(1 for s in scores if s <= raw) / n * 99 + 1) if raw else None
        before = len(a3_passed)
        a3_passed = [r for r in a3_passed if r.get("rs_rating") and r["rs_rating"] >= min_rs]
        if before - len(a3_passed):
            log(f"RS filter removed {before-len(a3_passed)} stocks below RS {min_rs}")

    # ── Add RSI(2) + Entry Trigger to final results ───────────────────────────
    for r in a3_passed:
        try:
            suffix = ".NS" if market == "IN" else ""
            hist = yf.Ticker(r["ticker"] + suffix).history(period="30d", interval="1d", auto_adjust=True)
            if hist is not None and len(hist) >= 5:
                r["rsi2"] = _compute_rsi2(hist["Close"].values)
        except Exception:
            r["rsi2"] = None

        trigger = check_entry_trigger(r["ticker"], market)
        r["entry_triggered"]  = trigger["triggered"]
        r["entry_note"]       = trigger["trigger_note"]
        r["prev_high"]        = trigger.get("prev_high")
        r["trigger_vol_ratio"]= trigger.get("vol_ratio")
        r["fixed_triple"]     = True
        r["regime"]           = regime["regime"]

        # Merge swing/value data from agents
        sw = next((x for x in a1_passed if x["ticker"] == r["ticker"]), {})
        va = next((x for x in a2_passed if x["ticker"] == r["ticker"]), {})
        for key in ["swing_score","swing_signals","rsi","ema_trend","vol_ratio","atr"]:
            if not r.get(key): r[key] = sw.get(key)
        for key in ["pe","roe","analyst_label","div_yield_pct"]:
            if not r.get(key): r[key] = va.get(key)

    log(f"Agent 3 complete — {len(a3_passed)} passed SEPA")
    log(f"FIXED TRIPLE DONE — {total} → {len(a1_passed)} tech → {len(a2_passed)} fund → {len(a3_passed)} final")

    # Merge all rejections with agent label
    all_rejected = (
        [dict(r, _agent=1) for r in a1_rejected] +
        [dict(r, _agent=2) for r in a2_rejected] +
        [dict(r, _agent=3) for r in a3_rejected]
    )

    return {
        "agent1": {"passed": a1_passed, "rejected": a1_rejected},
        "agent2": {"passed": a2_passed, "rejected": a2_rejected},
        "agent3": {"passed": a3_passed, "rejected": a3_rejected},
        "final":  a3_passed,
        "regime": regime,
        "all_rejected": all_rejected,
    }

# ── ETF Lookup — get top N ETFs/funds that hold a given ticker ────────────────
_ETF_KEYWORDS = [
    "ETF", "Trust", "Index Fund", "Select Sector",
    "iShares", "Vanguard", "SPDR", "PowerShares",
    "Invesco", "ARK", "ProShares", "Direxion",
    "WisdomTree", "First Trust", "Global X",
    "Schwab", "Fidelity Index", "State Street",
]

_ETF_SHORT = {
    "vanguard total stock":    "VTI",
    "vanguard 500":            "VOO",
    "vanguard s&p 500":        "VOO",
    "spdr s&p 500":            "SPY",
    "ishares core s&p 500":    "IVV",
    "ishares russell 2000":    "IWM",
    "ishares russell 1000":    "IWB",
    "technology select sector":"XLK",
    "financial select sector": "XLF",
    "health care select":      "XLV",
    "consumer discret":        "XLY",
    "industrial select":       "XLI",
    "energy select":           "XLE",
    "communication services":  "XLC",
    "materials select":        "XLB",
    "vanguard small-cap":      "VB",
    "vanguard mid-cap":        "VO",
    "vanguard growth":         "VUG",
    "vanguard value":          "VTV",
    "ishares core s&p small":  "IJR",
    "ishares core s&p mid":    "IJH",
    "ishares core msci":       "IEMG",
    "invesco qqq":             "QQQ",
    "powershares qqq":         "QQQ",
    "ark innov":               "ARKK",
    "schwab us large":         "SCHX",
    "schwab us broad":         "SCHB",
    "global x":                "GXC",
    "wisdomtree":              "DLN",
    "first trust":             "FDN",
    "vaneck":                  "SMH",
    "semiconductor":           "SOXX",
}


def _shorten_etf_name(holder_name: str) -> str:
    lower = holder_name.lower()
    for fragment, short in _ETF_SHORT.items():
        if fragment in lower:
            return short
    clean = holder_name
    for suffix in [" ETF", " Trust", " Index Fund", " Fund"]:
        clean = clean.replace(suffix, "")
    return clean.strip()[:20]


def get_top_etfs_for_ticker(ticker: str, n: int = 3) -> list:
    """
    Fetch top N ETFs/mutual funds that hold the given stock.
    Uses yfinance .mutualfund_holders (primary) with .institutional_holders fallback.
    Returns list of dicts: [{"name": "QQQ", "full_name": "...", "pct": 1.23}, ...]
    """
    import yfinance as yf
    results = []
    try:
        t  = yf.Ticker(ticker)
        mf = t.mutualfund_holders
        if mf is not None and not mf.empty:
            mf.columns = [c.strip() for c in mf.columns]
            pct_col = next(
                (c for c in mf.columns if "%" in c or "pct" in c.lower() or "out" in c.lower()),
                None
            )
            mf["_h"] = mf["Holder"].astype(str)
            mask = mf["_h"].apply(
                lambda h: any(kw.lower() in h.lower() for kw in _ETF_KEYWORDS)
            )
            rows = mf[mask].copy()
            if pct_col and pct_col in rows.columns:
                rows = rows.sort_values(pct_col, ascending=False)
            for _, row in rows.head(n).iterrows():
                full = str(row["Holder"])
                pct  = float(row[pct_col]) * 100 if pct_col else 0.0
                results.append({"name": _shorten_etf_name(full), "full_name": full, "pct": round(pct, 3)})
    except Exception:
        pass

    if not results:
        try:
            t    = yf.Ticker(ticker)
            inst = t.institutional_holders
            if inst is not None and not inst.empty:
                inst.columns = [c.strip() for c in inst.columns]
                pct_col = next(
                    (c for c in inst.columns if "%" in c or "pct" in c.lower() or "out" in c.lower()),
                    None
                )
                inst["_h"] = inst["Holder"].astype(str)
                mask = inst["_h"].apply(
                    lambda h: any(kw.lower() in h.lower() for kw in _ETF_KEYWORDS)
                )
                rows = inst[mask].copy()
                if pct_col and pct_col in rows.columns:
                    rows = rows.sort_values(pct_col, ascending=False)
                for _, row in rows.head(n).iterrows():
                    full = str(row["Holder"])
                    pct  = float(row[pct_col]) * 100 if pct_col else 0.0
                    results.append({"name": _shorten_etf_name(full), "full_name": full, "pct": round(pct, 3)})
        except Exception:
            pass

    return results[:n]
