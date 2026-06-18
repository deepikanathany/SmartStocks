"""
backtest_rejections.py
======================
Backtest the 52 stocks that failed exactly 1 filter in the Fixed Triple Screen.
Tests 5-day and 10-day holding periods over the last 30 trading days.
$10,000 total investment, equal-weight per profitable stock.

Run from your terminal:
    python backtest_rejections.py

Requirements: pip install yfinance pandas openpyxl
"""

import yfinance as yf
import pandas as pd
from datetime import datetime
from collections import defaultdict
import json

# ── The 52 one-fail stocks with their failed filter ───────────────────────────

MONTHLY_TREND_FAILS = {
    "ABBV": ("Healthcare",              -0.82),
    "AES":  ("Utilities",               -0.21),
    "ALL":  ("Financial Services",      -0.82),
    "AMT":  ("Real Estate",             -1.15),
    "BG":   ("Consumer Defensive",      -0.68),
    "BK":   ("Financial Services",      -0.30),
    "BKR":  ("Energy",                  -6.37),
    "BX":   ("Financial Services",      -2.25),
    "CARR": ("Industrials",             -2.39),
    "CSX":  ("Industrials",             -1.98),
    "DE":   ("Industrials",             -0.12),
    "DOW":  ("Basic Materials",         -2.62),
    "ES":   ("Utilities",               -2.79),
    "GEV":  ("Industrials",             -1.06),
    "HAL":  ("Energy",                  -1.42),
    "IEX":  ("Industrials",             -1.84),
    "INVH": ("Real Estate",             -1.08),
    "KDP":  ("Consumer Defensive",      -0.78),
    "KIM":  ("Real Estate",             -0.38),
    "KVUE": ("Consumer Defensive",      -0.34),
    "LNT":  ("Utilities",               -0.40),
    "LYB":  ("Basic Materials",         -0.39),
    "MAS":  ("Industrials",             -2.81),
    "MSCI": ("Financial Services",      -1.74),
    "NSC":  ("Industrials",             -1.10),
    "NTRS": ("Financial Services",      -2.40),
    "NWS":  ("Communication Services",  -0.23),
}

RSI_OVERBOUGHT_FAILS = {
    "AAPL":  ("Technology",             76.0),
    "ADM":   ("Consumer Defensive",     84.8),
    "CNC":   ("Healthcare",             93.3),
    "CVS":   ("Healthcare",             85.1),
    "DVA":   ("Healthcare",             87.0),
    "EQR":   ("Real Estate",            78.0),
    "GNRC":  ("Industrials",            88.3),
    "GOOG":  ("Communication Services", 78.1),
    "GOOGL": ("Communication Services", 79.3),
    "HUM":   ("Healthcare",             87.8),
    "INTC":  ("Technology",             81.1),
    "IVZ":   ("Financial Services",     76.2),
    "KO":    ("Consumer Defensive",     75.1),
}

EMA_TREND_FAILS = {
    "AMCR":  ("Consumer Cyclical",      "Bearish"),
    "AON":   ("Financial Services",     "Bearish"),
    "BF-B":  ("Consumer Defensive",     "Mixed"),
    "CBRE":  ("Real Estate",            "Bearish"),
    "CRL":   ("Healthcare",             "Bearish"),
    "ERIE":  ("Financial Services",     "Bearish"),
    "GEHC":  ("Healthcare",             "Bearish"),
    "INTU":  ("Technology",             "Bearish"),
    "IP":    ("Consumer Cyclical",      "Bearish"),
    "LH":    ("Healthcare",             "Bearish"),
    "LMT":   ("Industrials",            "Bearish"),
    "LUV":   ("Industrials",            "Bearish"),
}

TOTAL_CAPITAL   = 10_000
HOLD_DAYS_SHORT = 5
HOLD_DAYS_LONG  = 10
LOOKBACK_DAYS   = 30   # last 30 trading days


def backtest_stock(sym, sector, fail_filter, fail_value):
    """Backtest a single stock over the last 30 trading days."""
    try:
        hist = yf.Ticker(sym).history(period="3mo", interval="1d", auto_adjust=True)
        if hist is None or len(hist) < HOLD_DAYS_LONG + 2:
            return None

        closes = hist["Close"].values
        dates  = [str(d.date()) for d in hist.index.tolist()]
        n      = len(closes)

        # Use last 30 trading days as entry window
        bt_start = max(0, n - LOOKBACK_DAYS - HOLD_DAYS_LONG)

        rets_5  = []
        rets_10 = []
        entry_rows = []

        for i in range(bt_start, n - HOLD_DAYS_SHORT):
            ep    = float(closes[i])
            p5    = float(closes[i + HOLD_DAYS_SHORT])  if i + HOLD_DAYS_SHORT  < n else None
            p10   = float(closes[i + HOLD_DAYS_LONG])   if i + HOLD_DAYS_LONG   < n else None
            r5    = round((p5  - ep) / ep * 100, 2) if p5  else None
            r10   = round((p10 - ep) / ep * 100, 2) if p10 else None
            if r5  is not None: rets_5.append(r5)
            if r10 is not None: rets_10.append(r10)
            entry_rows.append({"date": dates[i], "entry": round(ep,2), "r5": r5, "r10": r10})

        if not rets_5:
            return None

        avg5  = round(sum(rets_5)  / len(rets_5),  2)
        avg10 = round(sum(rets_10) / len(rets_10), 2) if rets_10 else None
        wr5   = round(sum(1 for r in rets_5  if r > 0) / len(rets_5)  * 100, 1)
        wr10  = round(sum(1 for r in rets_10 if r > 0) / len(rets_10) * 100, 1) if rets_10 else None
        best5 = round(max(rets_5), 2)
        worst5= round(min(rets_5), 2)
        best10= round(max(rets_10), 2) if rets_10 else None
        current_price = round(float(closes[-1]), 2)

        return {
            "symbol":       sym,
            "sector":       sector,
            "fail_filter":  fail_filter,
            "fail_value":   str(fail_value),
            "current":      current_price,
            "entries":      len(rets_5),
            "avg_5d":       avg5,
            "wr_5d":        wr5,
            "best_5d":      best5,
            "worst_5d":     worst5,
            "avg_10d":      avg10,
            "wr_10d":       wr10,
            "best_10d":     best10,
            "entry_rows":   entry_rows,
        }
    except Exception as e:
        print(f"  ERROR {sym}: {e}")
        return None


def print_results(results, group_name, hold=5):
    k    = "avg_5d" if hold == 5 else "avg_10d"
    wr_k = "wr_5d"  if hold == 5 else "wr_10d"
    key  = "best_5d" if hold == 5 else "best_10d"

    valid = [r for r in results if r.get(k) is not None]
    valid.sort(key=lambda x: x[k], reverse=True)

    print(f"\n{'═'*70}")
    print(f"  {group_name} — {hold}-DAY HOLD RESULTS")
    print(f"{'═'*70}")
    print(f"  {'Sym':<7} {'Sector':<26} {'Avg Ret':>8} {'Win%':>6} {'Best':>7} {'Worst':>7} {'Signal'}")
    print(f"  {'─'*70}")

    profitable = []
    for r in valid:
        avg   = r[k]
        wr    = r[wr_k] or 0
        best  = r[key] or 0
        worst = r.get("worst_5d", 0) if hold == 5 else r.get("worst_5d", 0)
        emoji = "✅" if avg > 0 and wr >= 55 else ("🔶" if avg > 0 else "❌")
        print(f"  {r['symbol']:<7} {r['sector'][:25]:<26} {avg:>+7}% {wr:>5}% {best:>+6}% {worst:>+6}% {emoji}")
        if avg > 0 and wr >= 55:
            profitable.append(r)

    print(f"\n  Profitable (avg>0, WR≥55%): {len(profitable)}/{len(valid)}")
    return profitable


def compute_portfolio(all_results, hold=5):
    """Equal-weight portfolio of all profitable stocks."""
    k = "avg_5d" if hold == 5 else "avg_10d"
    profitable = [r for r in all_results if r.get(k) is not None and r[k] > 0]
    if not profitable:
        return

    n_stocks   = len(profitable)
    per_stock  = round(TOTAL_CAPITAL / n_stocks, 0)
    total_pnl  = 0
    print(f"\n  {'─'*70}")
    print(f"  PORTFOLIO — ${TOTAL_CAPITAL:,} equally split across {n_stocks} profitable stocks")
    print(f"  ${per_stock:,.0f} per stock")
    print(f"  {'─'*70}")
    print(f"  {'Sym':<7} {'Investment':>12} {'Avg Ret':>8} {'Est P&L':>10}")

    profitable.sort(key=lambda x: x[k], reverse=True)
    for r in profitable:
        pnl = round(per_stock * r[k] / 100, 0)
        total_pnl += pnl
        print(f"  {r['symbol']:<7} ${per_stock:>10,.0f} {r[k]:>+7}% ${pnl:>+9,.0f}")

    print(f"  {'─'*70}")
    print(f"  {'TOTAL':<7} ${TOTAL_CAPITAL:>10,}         ${total_pnl:>+9,.0f}")
    print(f"  Total return: {round(total_pnl/TOTAL_CAPITAL*100, 2)}%")


def export_excel(all_results):
    """Export full results to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        wb = openpyxl.Workbook()

        # Summary sheet
        ws = wb.active; ws.title = "Summary"
        headers = ["Symbol","Sector","Filter Failed","Fail Value","Current Price",
                   "Avg 5d Return%","Win Rate 5d%","Best 5d%","Worst 5d%",
                   "Avg 10d Return%","Win Rate 10d%","Best 10d%","# Entries"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="00D68F")
            cell.fill = PatternFill("solid", fgColor="0D1520")

        for i, r in enumerate(sorted(all_results, key=lambda x: x.get("avg_5d",0), reverse=True), 2):
            vals = [r["symbol"], r["sector"], r["fail_filter"], r["fail_value"],
                    r["current"], r.get("avg_5d"), r.get("wr_5d"), r.get("best_5d"), r.get("worst_5d"),
                    r.get("avg_10d"), r.get("wr_10d"), r.get("best_10d"), r.get("entries")]
            for c, v in enumerate(vals, 1):
                ws.cell(row=i, column=c, value=v)

        fname = f"rejection_backtest_{datetime.today().strftime('%Y%m%d')}.xlsx"
        wb.save(fname)
        print(f"\n  Excel saved: {fname}")
    except ImportError:
        print("\n  (install openpyxl for Excel export: pip install openpyxl)")


# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    all_results = []
    groups = [
        (MONTHLY_TREND_FAILS,  "Monthly Trend",  "Monthly Trend"),
        (RSI_OVERBOUGHT_FAILS, "RSI Overbought", "RSI(14) Range"),
        (EMA_TREND_FAILS,      "EMA Trend",      "EMA Trend"),
    ]

    for stock_dict, group_name, fail_filter in groups:
        print(f"\nFetching data for {group_name} group ({len(stock_dict)} stocks)...")
        group_results = []
        for sym, info in stock_dict.items():
            sector = info[0]; fail_val = info[1]
            print(f"  {sym}...", end=" ", flush=True)
            r = backtest_stock(sym, sector, fail_filter, fail_val)
            if r:
                r["group"] = group_name
                group_results.append(r)
                all_results.append(r)
                print(f"5d:{r['avg_5d']:+.1f}% WR:{r['wr_5d']}%")
            else:
                print("no data")

        print_results(group_results, group_name, hold=5)
        print_results(group_results, group_name, hold=10)

    # Overall portfolio
    print(f"\n\n{'█'*70}")
    print(f"  OVERALL — ALL 52 STOCKS COMBINED")
    print(f"{'█'*70}")
    print_results(all_results, "All Groups", hold=5)
    print_results(all_results, "All Groups", hold=10)

    compute_portfolio(all_results, hold=5)
    compute_portfolio(all_results, hold=10)

    export_excel(all_results)

    print(f"\n{'═'*70}")
    print("  KEY RULES for using these results:")
    print("  1. Monthly Trend failures are highest probability (< -3% dip = avoid)")
    print("  2. RSI overbought stocks: only enter if RSI starts falling from peak")
    print("  3. EMA failures: wait for EMA20 reclaim before entering")
    print("  4. Never invest more than 2-3% of capital in any single trade")
    print(f"{'═'*70}")